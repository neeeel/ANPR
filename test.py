import pandas as pd
import numpy as np
import datetime
import openpyxl
from tkinter import messagebox,filedialog
import os.path
import mainwindow
import myDB
import pickle
import copy
import time
import csv
import threading
import shutil
import anprregex
import re


flag = False

def update_sites_and_movements(job):
    global df
    ###
    ### users can edit sites, changing movements and directions
    ### rather than have to make a new job, and reload all the plates etc,
    ### we want to be able to update the job in the database, and then propagate any changes through to
    ### the stored data
    ###
    print("updating",job)
    dataFolder = os.path.join(job["folder"], "data")
    job = myDB.load_job(job["jobno"],job["jobname"],datetime.datetime.strftime(job["surveydate"],"%d/%m/%y"))
    job["timeAdjustmentsDictionary"] = {}
    load_job(job)
    if not df is None:
        df["Site"] = df["Movement"].apply(convert_movement_to_site, args=(job,))
        df["newMovement"] = df["Movement"].apply(convert_old_movement_to_new, args=(job,))
        df["dir"] = df["newMovement"].apply(convert_movement_to_dir, args=(job,))
        save_job(job)
        try:
            os.remove(dataFolder + "/comparisondata.pkl")
        except Exception as e:
            pass
        threading.Thread(target=produce_full_routes, args=(job,)).start()

def convert_movement_to_site(val,job):
    ###
    ### function to convert one column of the pandas dataframe.
    ### val is the original movement number for a vehicle
    ### returns the site number for that specific original movement
    ###
    for siteNo,site in job["sites"].items():
        for moveNo, movement in site.items():
            if val in movement["originalmovements"]:
                #print("site",siteNo,"contains original movement",val)
                return siteNo
    return 0

def convert_old_movement_to_new(val,job):
    ###
    ### function to convert one column of the pandas dataframe
    ### val is the original movement number for a specific vehicle
    ### returns the new movement number
    ###



    for siteNo, site in job["sites"].items():
        for moveNo, movement in site.items():
            if val in movement["originalmovements"]:
                # print("site",siteNo,"contains original movement",val)
                return moveNo

def convert_movement_to_dir(val,job):
    ###
    ### function to convert a column of the pandas dataframe
    ### val is the new movement number for a specific vehicle
    ### returns the direction (0=in,1=out,2=both)
    ###
    for siteNo, site in job["sites"].items():
        for moveNo, movement in site.items():
            if val == moveNo:
                # print("site",siteNo,"contains original movement",val)
                return movement["dir"]

def load_unclassed_plates(job):
    global df
    ###
    ### Load unclassed vehicles file, sort out columns etc
    ###

    folder = job["folder"]
    dataFolder = os.path.join(folder,"data" )
    if not os.path.exists(dataFolder):
        os.makedirs(dataFolder)
    if os.path.isfile(dataFolder + "/data.pkl"):
        result =  messagebox.askquestion("Warning",
                                      "The unclassed plates have previously been loaded, do you want to reload them?")
        if result == "no":
            return
    result = messagebox.askquestion("Warning",message="This will reset all the files in the project, and wipe any progress, do you want to continue?")
    if result == "no":
        return
    reset_project(job)

    df = None
    file = filedialog.askopenfilename(initialdir=folder)
    if file == "":
        messagebox.showinfo(message = "No file selected,no plates loaded")
        return
    ext = file[file.rfind("."):]
    if ext not in (".xlsx",".csv",".xlsm",".xls"):
        messagebox.showinfo(message="Not valid CSV file, No plates loaded")
        return

    try:
        if ".csv" in file:
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file, converters={"VRN": str, "Direction": str, "Date": str, "Time": str})
        df = df[["Date","Time","Movement","VRN"]]
        print(df.head())
        df["Date"] = pd.to_datetime(df["Date"] + " " + df["Time"])
        df.drop(["Time"], inplace=True, axis=1)
        df["Class"] = ""


        ### TODO  : Check that we are doing the duplicates correctly

        ###
        ### set up the timediff and duplicates column
        ###
        print(df.head())
        df["Duplicates"] = "N"
        df["Site"] = df["Movement"].apply(convert_movement_to_site, args=(job,))
        df["newMovement"] = df["Movement"].apply(convert_old_movement_to_new, args=(job,))
        df["dir"] = df["newMovement"].apply(convert_movement_to_dir, args=(job,))
        df.sort_values(by=["VRN", "newMovement"], inplace=True, ascending=[True, True])
        df["timeDiff"] = df["Date"].diff()
        df.set_index(["Date"],inplace=True)
        mask = (df["VRN"] != df["VRN"].shift(-1)) | (df["newMovement"] != df["newMovement"].shift(-1))
        df["timeDiff"][mask] = np.nan
        df.to_pickle(dataFolder + "/data.pkl")
    except Exception as e:
        messagebox.showinfo(message="Error occured while loading csv file " + str(e))
        print(e)
        df = None
        return
    print(df.info())
    myDB.update_job_with_progress(job["id"],"unclassed")
    compute_comparison_data(job)
    load_job(job)

def reset_project(job):
    dataFolder = os.path.join(job["folder"], "data")
    try:
        os.remove(dataFolder+ "/data.pkl")
    except Exception as e:
        pass
    try:
        os.remove(dataFolder + "/classedData.pkl")
    except Exception as e:
        pass
    try:
        os.remove(dataFolder + "/comparisondata.pkl")
    except Exception as e:
        pass
    try:
        os.remove(dataFolder + "/OVData.pkl")
    except Exception as e:
        pass
    try:
        os.remove(dataFolder + "/durations.pkl")
    except Exception as e:
        pass

def load_completed_count(job):
    global overviewDf
    overviewDf = None
    dataframes = []
    dataFolder = os.path.join(job["folder"], "data")
    seen = set()
    classes = [x for i, x in enumerate(job["classification"].split(",")) if i % 2 == 0 and x not in seen and not seen.add(x)]
    classes.insert(0, "Time")
    numClasses = int(len(classes))
    file = filedialog.askopenfilename(initialdir=job["folder"])
    if file == "":
        return
    try:
        wb = openpyxl.load_workbook(file,data_only=True)
        for siteNo,siteDetails in job["sites"].items():
            print("loading site ",siteNo)
            print("site is",siteDetails)
            for movement,x in siteDetails.items():
                print("loading movement",movement)
                data = []
                print("movement details are ",x)
                try:
                    ws = wb.get_sheet_by_name("site " + str(siteNo))
                except openpyxl.utils.exceptions.SheetTitleException as e:
                    try:
                        ws = wb.get_sheet_by_name("Site " + str(siteNo))
                    except openpyxl.utils.exceptions.SheetTitleException as e:
                        print("bad sheet")
                        continue
                except KeyError as e:
                    try:
                        ws = wb.get_sheet_by_name("Site " + str(siteNo))
                    except KeyError as e:
                        print("couldnt find sheet")
                        print(e)
                        continue
                firstMovementOnSheet = ws["C3"].value
                if movement == firstMovementOnSheet:
                    offset = 1
                else:
                    offset = numClasses + 3

                for row in ws.iter_rows(column_offset=offset,row_offset=7):
                    #print("row is",row)
                    rowData = []
                    for cell in row[:numClasses]:
                        if isinstance(cell.value,datetime.time):
                            rowData.append(cell.value.strftime("%H:%M:%S"))
                        else:
                            rowData.append(cell.value)
                    if rowData != [] and "hr" not in str(rowData[0]).lower() and "total" not in str(rowData[0]).lower() :
                        data.append(rowData)
                data.insert(0,classes)
                countsDf = pd.DataFrame(data[1:],columns=data[0])
                countsDf.dropna(inplace =True)
                countsDf["Site"] = siteNo
                countsDf["Movement"] = movement
                countsDf["Date"] = datetime.datetime.strftime(job["surveydate"],"%d/%m/%y")
                countsDf["Date"] = pd.to_datetime(countsDf["Date"] + " " + countsDf["Time"],dayfirst=True)
                countsDf.drop(["Time"], inplace=True, axis=1)
                countsDf.set_index(["Date"], inplace=True)
                dataframes.append(countsDf)
    except Exception as e:
        print("oh phoo")
        print(e)
        messagebox.showinfo(message="Something went wrong when trying to read site "+ str(siteNo) + ", please check the format of the sheet")
        overviewDf=None
        #wb.close()
        return
    overviewDf = pd.concat(dataframes)
    for cl in classes[1:]:
        overviewDf[cl] = overviewDf[cl].astype(int)
    overviewDf.to_pickle(dataFolder + "/OVData.pkl")
    compute_comparison_data(job)
    print("Done")

def load_classes(job):
    global df
    dataFolder = os.path.join(job["folder"], "data")
    if not os.path.isfile(dataFolder + "/data.pkl"):
        messagebox.showinfo(message="You havent loaded any unclassed plates yet")
        return
    if os.path.isfile(dataFolder + "/classes.pkl"):
        if not messagebox.askquestion("Warning","The classes have previously been loaded, do you want to reload them?"):
            return
    file = filedialog.askopenfilename(initialdir=job["folder"])
    if file == "":
        messagebox.showinfo(message="No file selected,no classes loaded")
        return
    ext = file[file.rfind("."):]
    if ext not in (".xlsx", ".csv", ".xlsm", ".xls"):
        messagebox.showinfo(message="Not valid file, No classes loaded")
        return
    try:
        ###
        ### load the unclassed plates into the dataframe
        ###

        df = pd.read_pickle(dataFolder + "/data.pkl")

        ###
        ### load the classed plates into a temp data frame, save it as a pickle
        ###

        tempdf = pd.read_excel(file,converters={"VRN": str})
        tempdf.drop_duplicates(subset=["VRN"],inplace=True)
        tempdf = tempdf[["VRN","Class"]]
        tempdf.to_pickle(dataFolder + "/classes.pkl")
        df.drop("Class", inplace=True)
        df.reset_index(inplace=True)
        df = df.merge(tempdf, how="left",on="VRN")
        df.set_index(["Date"],inplace=True)
        df.drop("Class_x", axis=1, inplace=True)
        df.rename(columns={"Class_y": "Class"}, inplace=True)
        # df = df[pd.notnull(df["Class"])]
        df.to_pickle(dataFolder + "/classedData.pkl")
        df.to_csv("dumped.csv")
        myDB.update_job_with_progress(job["id"], "classed")
        compute_comparison_data(job)
        threading.Thread(target=produce_full_routes,args=(job,)).start()
    except FileNotFoundError as e:
        messagebox.showinfo(message="Something went wrong when trying to load the classes, please check that the file is a valid file")
        df = None
        return
    print("after loading classes, lenght is",len(df))
    compute_comparison_data(job)

def save_job(job):
    global df, overviewDf
    dataFolder = os.path.join(job["folder"], "data")
    ### save plates, classed and unclassed, from pickled dataframe file
    print("looking for",dataFolder + "/classedData.pkl")
    if  os.path.isfile(dataFolder + "/classedData.pkl"):
        df.to_pickle(dataFolder + "/classedData.pkl")
        print("saving c")
    else:
        df.to_pickle(dataFolder + "/data.pkl")

def load_job(job):
    global df,overviewDf
    ### load plates, classed or unclassed, from pickled dataframe file
    dataFolder = os.path.join(job["folder"], "data")
    job["platerestrictionpercentages"] = [0, 0, 0, 0]
    try:
        df = pd.read_pickle(dataFolder + "/classedData.pkl")
        #print(df.info())
    except FileNotFoundError as e:
        # messagebox(message="Data file is missing, you will need to load the unclassed plates")
        print("No classed data found, trying to load unclassed data")
        try:
            df = pd.read_pickle(dataFolder + "/data.pkl")
            #print(df.info())
        except FileNotFoundError as e:
            messagebox.showinfo(message="Data file is missing, you will need to load the unclassed plates")
            print("No unclassed data found")
            df = None
            return True
        print("Loaded unclassed plates, no of entries",len(df))


    ###
    ### load comparison
    ###
    try:
        overviewDf = pd.read_pickle(dataFolder + "/OVData.pkl")
        print("Loaded completed overview count")
    except FileNotFoundError as e:
        print("No comparison data found")
    except Exception as e:
        print(e)
        print("ERRRRRRRROR")
        df = None
        return False

    ###
    ### restrict plates based on length
    ###

    print("lenth of df BEFORE plate restrictions", len(df))
    job["platerestrictionpercentages"] = []

    if len(df) != 0:
        job["platerestrictionpercentages"].append(100)
        #print(len(df[(df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 7)]))
        job["platerestrictionpercentages"].append(
            float(
                "{0:.2f}".format(len(df[(df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 7)]) * 100 / len(df))))
        job["platerestrictionpercentages"].append(
            float(
                "{0:.2f}".format(len(df[(df["VRN"].str.len() >= 5) & (df["VRN"].str.len() <= 7)]) * 100 / len(df))))
        job["platerestrictionpercentages"].append(
            float(
                "{0:.2f}".format(len(df[(df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 8)]) * 100 / len(df))))
    else:
        job["platerestrictionpercentages"] = [0, 0, 0, 0]
        return True

    if job["platerestriction"] == 1:
        pass
    else:
        if job["platerestriction"] == 2:
            mask = (df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 7)
        if job["platerestriction"] == 3:
            mask = (df["VRN"].str.len() >= 5) & (df["VRN"].str.len() <= 7)
        if job["platerestriction"] == 4:
            mask = (df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 8)
        df = df[mask]
    print("lenth of df after plate restrictions", len(df))


    duplicates=[]
    for i in range(31):
        duplicates.append (len(df[df["timeDiff"] == pd.Timedelta(seconds=i)]))
    df["newMovement"] = df["newMovement"].astype(int,raise_on_error=False)
    for i in range(0, 465, 15):
        mask = (df["timeDiff"] >= pd.Timedelta(seconds=i)) & (df["timeDiff"] < pd.Timedelta(seconds=i + 15))
        duplicates.append (len(df[mask]))
    job["duplicateValues"] = duplicates
    set_duplicates(job["selectedduplicates"])
    print("duplicates are",duplicates)


    ###
    ### make any time adjustments
    ###



    df.reset_index(inplace=True)
    for k,v in job["timeAdjustmentsDictionary"].items():
        if v != 0:
            df.ix[df["newMovement"] == k, "Date"] += pd.Timedelta(seconds=v)
    df.set_index("Date",inplace=True)



    return True

def bin_time(t):
    ###
    ### takes a timedelta t
    ### put the timedelta t into a bin , basically flooring it to the nearest 15 seconds
    ### so 00:01:27 would return a bin of 00:01:15
    ###
    seconds = t.seconds - t.seconds%15
    return datetime.timedelta(seconds = seconds)

def get_comparison_data(job):
    ###
    ### retrieve and return the comparison data for a job
    ### If we have previously computed it, load it from the pickled file
    ### otherwise, compute it
    ###
    dataFolder = os.path.join(job["folder"], "data")
    if os.path.isfile(dataFolder + "/comparisondata.pkl"):
        with open(dataFolder + '/comparisondata.pkl', 'rb') as handle:
            data = pickle.load(handle)
            site = data[0]
            print("loaded site",site)
            l = [movement for key, movement in sorted(site["movements"].items())]
            movement = l[0]
            print("after loading, data is~",movement["data"])
    else:
        data = compute_comparison_data(job)
    print("finished getting comparison data")
    return data

def compute_comparison_data(job):
    ###
    ### compute the comparison data from the loaded plates and overview count
    ### do this for each site in the job.
    ### the data for each site is a dictionary, with sub-dictionaries for each movement
    ### each sub dictionary has a list of data, in the format
    ###[[original OV data],[Edited OV data],[unclassed original ANPR data],[unclassed ANPR data -duplicates removed],[classed original ANPR data],[classed ANPR data-duplicates removed]]
    ### each movement also has a subdictionary with summary details.
    ### each set of data is set up ready to display, with column and row totals, and timestamps already entered
    ###
    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")
    oldData = None
    if os.path.isfile(dataFolder + "/comparisondata.pkl"):
        with open(dataFolder + '/comparisondata.pkl', 'rb') as handle:
           oldData = pickle.load(handle)

    global df,overviewDf
    if overviewDf is None:
        messagebox.showinfo(message="No Overview counts loaded")
        return None
    if df is None:
        messagebox.showinfo(message="No Classed or Unclassed plates loaded")
        return None
    results = []
    seen = set()
    ANPRClasses = [x for i, x in enumerate(job["classification"].split(",")) if i % 2 == 1 and x not in seen and not seen.add(x)]
    seen = set()
    OVClasses = [x for i, x in enumerate(job["classification"].split(",")) if i % 2 == 0 and x not in seen and not seen.add(x)]
    ANPRtoOVdict = {} ### this will hold a dictionary of how we combine the OV classes into the ANPR classes
    for cl in ANPRClasses:
        ANPRtoOVdict[cl] = []
        for item in [i for i, x in enumerate(job["classification"].split(",")) if x.lower() == cl.lower() and i % 2 == 1]:
            ANPRtoOVdict[cl].append(OVClasses.index(job["classification"].split(",")[item-1]))
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    d = datetime.datetime.strftime(job["surveydate"],"%Y-%m-%d")
    for siteNo, siteDetails in job["sites"].items():
        site = {}
        #print("site details",siteDetails)
        site["siteNo"] = int(siteNo)
        site["movements"] = {}

        for movement, original in siteDetails.items():
            #print("processing site",siteNo,"movement",movement)
            mvt = {}
            mvt["movementNo"] = movement
            mvt["data"] = [[], [], [], [], [], []]
            mvt["summary"] = {}
            mvt["summary"]["OVTotal"] = 0
            mvt["summary"]["ANPRTotal"] = 0
            mvt["summary"]["AvgCapture"] = 0
            mvt["summary"]["MinCapture"] = 1000
            mvt["summary"]["MaxCapture"] = 0
            mvt["summary"]["TimeLessThan"] = 0
            site["movements"][movement] = mvt
            ANPRdata  = []
            OVdata = []
            for i in range(0, len(times) - 1, 2):
                result = get_OV_data(job,movement,times[i],times[i+1])
                for item in result:
                    OVdata.append(item)

            ###
            ### set up the OVdata for display
            ###

            newList = []
            first = True
            rowList = []
            for i, item in enumerate(OVdata):
                # print("processing item", item, "time fs", item[0])
                t = datetime.datetime.strptime(item[0], "%H:%M:%S")
                mvt["summary"]["OVTotal"] = mvt["summary"]["OVTotal"] + int(sum(item[1:]))
                if t.minute == 0 and not first:
                    rowList = [int(sum(r)) for r in zip(*rowList)]
                    rowList.insert(0, "1 Hr")
                    newList.append(list(rowList))
                    rowList = []
                first = False
                rowList.append(list(item[1:]))
                newList.append(list(item))
            rowList = [int(sum(r)) for r in zip(*rowList)]
            rowList.insert(0, "1 Hr")
            newList.append(list(rowList))
            OVdata = newList
            mvt["data"][0] = OVdata
            mvt["data"][1] = list(OVdata) # the edited version of OVdata, is same as original initally

            for index,combination in enumerate([(0,0),(1,0),(0,1),(1,1)]): #combinations of classed,unclassed,duplicates,etc
                a,b = combination
                print("site",siteNo,"combination",combination)
                ANPRdata = []
                for t in range(0, len(times) - 1, 2):
                    result = get_ANPR_data(job, movement, times[t], times[t + 1], a, b)
                    for item in result:
                        ANPRdata.append(item)

                ###
                ### set up ANPR data for display
                ###

                newList = []
                first = True
                rowList = []  ### holds the blocks of data that we want to sum by column
                for i, item in enumerate(ANPRdata):
                    # print("processing",item)
                    t = datetime.datetime.strptime(item[0], "%H:%M:%S")
                    if t.minute == 0 and not first:
                        rowList = [int(sum(r)) for r in zip(*rowList)]
                        rowList.insert(0, "1 Hr")
                        newList.append(list(rowList))
                        rowList = []
                    first = False
                    mvt["summary"]["ANPRTotal"] = mvt["summary"]["ANPRTotal"] + item[-1]
                    rowList.append(list(item[1:]))
                    newList.append(list(item))
                    # print("newlist is",newList)
                rowList = [int(sum(r)) for r in zip(*rowList)]
                rowList.insert(0, "1 Hr")
                newList.append(list(rowList))
                ANPRdata = newList
                mvt["data"][2+index] = ANPRdata
            mvt["summary"]["TimeLessThan"] = datetime.timedelta(
                seconds=mvt["summary"]["TimeLessThan"] * job["interval"] * 60)
            if mvt["summary"]["OVTotal"] != 0:
                # print("ovtotal",site["summary"]["OVTotal"])
                mvt["summary"]["AvgCapture"] = int(
                    mvt["summary"]["ANPRTotal"] * 100 / mvt["summary"]["OVTotal"])
        results.append(site)
    ###
    ### If there is any comparison data saved, get the edited portion of the data, and fill in the newly
    ### calculated site comparison data with the edited section of the saved comparison data.
    ###

    if not oldData is None:
        for index, oldSite in enumerate(oldData):
            newsite = results[index]
            for key, details in oldSite["movements"].items():
                newsite["movements"][key]["data"][1] = copy.deepcopy(details["data"][1])
    with open(dataFolder + '/comparisondata.pkl', 'wb') as handle:
        pickle.dump(results, handle)
    print("finished processing")
    return results

def get_ANPR_data(job,movement,startTime,endTime,classed,duplicates_removed):
    ANPRdata = []
    d = datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")
    mask = "newMovement == " + str(movement)
    seen = set()
    ANPRClasses = [x for i, x in enumerate(job["classification"].split(",")) if i % 2 == 1 and x not in seen and not seen.add(x)]
    ANPRFrames = []  # a list of lists, 1 list for each ANPR class, each list is the count of that class for each timeslice
    rng = pd.date_range(d + " " + startTime, d + " " + endTime, freq=str(job["interval"]) + "T", closed="left")
    indexDf = pd.DataFrame(index=rng)
    numRows = len(indexDf)
    tempdf = df[d].query(mask).between_time(startTime, endTime, include_end=False)
    tempdf.sort_index(inplace=True)
    if duplicates_removed == True:
        l= len(tempdf)
        tempdf = tempdf[tempdf["Duplicates"] == "N"]
        print("removed", l-len(tempdf))
    for cl in ANPRClasses:
        ###
        ### find all vehicles of a class cl
        ###
        if classed == True:
            try:
                count = tempdf[tempdf["Class"] == cl]
            except Exception as e:
                count = []
        else:
            count = []
        if len(count) == 0:
            ANPRFrames.append([0] * numRows)  ### if theres no classed data, just build a dummy list of 0's
        else:

            ###
            ### resample the vehicles, giving us a series with the number of vehicles of a given class
            ### for each time sample
            ###


            if duplicates_removed == True:
                #print("before removing", len(count))
                try:
                    count = count[count["Duplicates"] == "N"]
                except TypeError as e:
                    print()
                #print("after removing", len(count))

            count = count.resample(str(job["interval"]) + "T").count()

            ###
            ### we use the indexDf to make sure that any times that have no vehicles are registered
            ### as a 0 for that time period
            ###

            count = indexDf.copy().merge(count, how="left", left_index=True, right_index=True).fillna(0)
            ANPRFrames.append(count["VRN"].values.tolist())

    #print("after first stage, ANPR frames are",ANPRFrames)
    if not classed:
        ###
        ### we arent looking at classed plates, so just want the total number of vehicles for each time sample
        ### use the indexdf to fill in any time samples where there are no vehicles
        ###

        if duplicates_removed == True:
            #print("before removing", len(tempdf))
            try:
                result = tempdf[tempdf["Duplicates"] == "N"]
                #print("after removing", len(result))
                result = result.resample(str(job["interval"]) + "T").count()
            except TypeError as e:
                result = tempdf.resample(str(job["interval"]) + "T").count()
        else:
            result = tempdf.resample(str(job["interval"]) + "T").count()
        df_filtered = indexDf.copy().merge(result, how="left", left_index=True, right_index=True).fillna(0)
        df_filtered=df_filtered[["VRN"]]
    else:

        ###
        ### set up a dataframe based on the time series index, with a column for the total ANPR vehicle counts
        ### to give us the "Total" column for the comparison display
        ###
        sumList = [int(sum(r)) for r in zip(*ANPRFrames)]
        indexDf.reset_index(inplace=True)
        df_filtered = indexDf.copy().merge(pd.DataFrame(sumList), how="left", left_index=True, right_index=True).fillna(0)
        df_filtered.columns = ["Date", "total"]
        df_filtered.set_index(["Date"], inplace=True)


    if len(df_filtered) == 0:
        pass
        ### construct a dummy dataframe if the search returns empty

    ###
    ### insert columns corresponding to ANPR classes, remove unneeded columns
    ###

    for j, cl in enumerate(ANPRClasses):
        ###
        ### ANPRFrames is a list of lists, one list for each class of vehicle. This list has a count of
        ### that vehicle type for each time period. We want to insert this list as a column into the
        ### filtered_df, which currently holds the total number of vehicles for each time period
        ### We insert each list as a column labeled "cl" where cl is the class name, eg car, LGV etc
        ###
        #print(j, "th frame", "lenght", len(ANPRFrames[j]), ANPRFrames[j])
        df_filtered.insert(j, cl, ANPRFrames[j])

    ###
    ### now df_filtered is in the correct structure, a column for each class, and a column for total vehicles
    ### we remove unneeded columns, and produce a list of lists, one list for each time sample
    ###


    #print(df_filtered.head())
    df_filtered.index.name = "Date"
    df_filtered.reset_index(inplace=True)
    df_filtered["Date"] = df_filtered["Date"].dt.strftime("%H:%M:%S")
    for item in df_filtered.values.tolist():
        ANPRdata.append(item)
    return ANPRdata

def get_OV_data(job,movement,startTime,endTime):
    #print("trying to get OC data for movement ",movement)
    d = datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")
    rng = pd.date_range(d + " " + startTime, d + " " + endTime, freq=str(job["interval"]) + "T", closed="left")
    indexDf = pd.DataFrame(index=rng)
    seen = set()
    OVClasses = [x for i, x in enumerate(job["classification"].split(",")) if i % 2 == 0 and x not in seen and not seen.add(x)]
    OVCounts = overviewDf[d].query("Movement == " + str(movement))
    OVCounts.index.to_datetime(dayfirst=True)
    OVCounts = OVCounts.between_time(startTime, endTime, include_end=False)
    OVCounts.reset_index(inplace=True)
    OVCounts["Date"] = OVCounts["Date"].dt.strftime("%H:%M:%S")
    del OVCounts["Site"]
    del OVCounts["Movement"]
    OVdata = []
    if len(OVCounts) == 0:
        temp = indexDf.copy(deep=True)
        temp.reset_index(inplace=True)
        temp.rename(columns={"index": "Date"}, inplace=True)
        temp["Date"] = temp["Date"].dt.strftime("%H:%M:%S")
        for j, cl in enumerate(OVClasses):
            temp.insert(j + 1, cl, [0] * len(temp))
        temp.insert(j + 1, "total", [0] * len(temp))
        for item in temp.values.tolist():
            OVdata.append(item)
            #print("FIlled OVData is", OVdata)
    else:
        OVCounts["Date"] = pd.to_datetime(
            datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d") + " " + OVCounts["Date"])
        OVCounts.set_index(["Date"], inplace=True)
        OVCounts = indexDf.merge(OVCounts, how="left", left_index=True, right_index=True).fillna(0)
        OVCounts["total"] = OVCounts.sum(axis=1)
        OVCounts.reset_index(inplace=True)
        OVCounts.rename(columns={"index": "Date"}, inplace=True)
        OVCounts["Date"] = OVCounts["Date"].dt.strftime("%H:%M:%S")
        for item in OVCounts.values.tolist():
            OVdata.append(item)
    return OVdata

def set_duplicates(index):
    global df
    if index == -1:
        return
    if index <= 30:
        mask = (df["timeDiff"] <= pd.Timedelta(seconds = index))
    else:
        start = (index -31) * 15
        mask = (df["timeDiff"] < pd.Timedelta(seconds=start + 15))
    df["Duplicates"] = "N"
    df.ix[mask,"Duplicates"] = "Y"
    print("set no of duplicates",len(df[df["Duplicates"] == "Y"]))

def set_new_duplicates_value(index,job):
    ###
    ### index : index of which label was clicked in the window, relating to the time value we want to set for excluding
    ### any duplicates,eg index 0 refers to all duplicates with a time diff of 0, 1 refers to duplicates with a time diff
    ### of 1, etc. Any index over 30 refers to a 15 second chunk of time
    ###
    ### when we are setting a new duplicates value, we need to re-compute the comparison data
    set_duplicates(index)
    compute_comparison_data(job)

def format_timedelta(td):
    if pd.isnull(td):
        return 0
    minutes, seconds = divmod(td.seconds + td.days * 86400, 60)
    hours, minutes = divmod(minutes, 60)
    return '{:d}:{:02d}:{:02d}'.format(hours, minutes, seconds)

def date_to_time(d):
    if d is None:
        return "00:00:00"
    if pd.isnull(d):
        return "00:00:00"
    try:
        return d.strftime("%H:%M:%S")
    except Exception as e:
        try:
            return d.strftime("%H:%M")
        except Exception as e:
            return "00:00:00"

def dir_to_str(dir):
    dirs = ["I", "O", "B"]
    return dirs[int(dir)-1]

def calculate_nondirectional_cordon(job):
    ###
    ### we want to "pair off" appearances of a vehicle. So if there are 4 appearances of a vehicle, we pair them off as
    ### (1,2) and (3,4). Unlike directional, we dont care about whether the first is an in and the second is an out
    ###

    global df
    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")
    inMov = []
    outMov = []
    for site, details in job["sites"].items():
        for mvmtNo, mvmt in details.items():
            if not  int(mvmt["newmovement"]) in inMov:
                inMov.append(int(mvmt["newmovement"]))
            if not int(mvmt["newmovement"]) in outMov:
                outMov.append(int(mvmt["newmovement"]))
    inMov = sorted(inMov)
    outMov = sorted(outMov)
    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    start = datetime.datetime.now()
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    for index in range(0, len(times) - 1, 2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
        temp["matched"] = "N"
        grp = temp.groupby(["VRN"])
        temp["matched"][grp.cumcount() % 2 == 0] = "Y"
        temp["matched"][grp.cumcount(ascending=False) == 0] = "N"
        temp["outTime"] = temp["Date"].shift(-1)
        temp["outMovement"] = temp["newMovement"].shift(-1)
        temp["newMovement"] = temp["newMovement"].real.astype(int)
        temp["outMovement"] = temp["outMovement"].real.astype(int)
        try:
            temp.to_csv("non directional full data.csv", index=False)
        except PermissionError as e:
            pass
            #messagebox.showinfo(
                #message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")
        temp = temp[temp["matched"] == "Y"]
        dataframes.append(temp)

    temp = pd.concat(dataframes)
    temp = temp[temp["outMovement"] >= 0]
    temp["newMovement"].dropna(inplace=True)
    temp = temp[temp["matched"] == "Y"]
    del temp["Movement"]
    del temp["Duplicates"]
    del temp["Site"]
    del temp["dir"]
    del temp["timeDiff"]
    del temp["matched"]
    temp["duration"] = temp["outTime"] - temp["Date"]

    ###
    ### durations check
    ###

    dataframes = []
    if not job["durationsDictionary"] is None:
        for k, v in job["durationsDictionary"].items():
            i, o = k
            splitTime = v.split(":")
            hours = int(splitTime[0])
            mins = int(splitTime[1])
            td = datetime.timedelta(hours=hours, minutes=mins, seconds=0)
            mask = (temp["newMovement"] == i) & (temp["outMovement"] == o) & (temp["duration"] <= td)
            dataframes.append(temp[mask].copy())
        temp = pd.concat(dataframes)



    ###
    ### aggregate the results
    ###

    counts = temp.groupby(["newMovement", "outMovement"]).size()
    aggs = temp.groupby(["newMovement", "outMovement"]).agg({"duration": [pd.DataFrame.max, pd.DataFrame.min,lambda x: sum(x, pd.Timedelta(0)) / len(x) if len(x) > 0 else 0]})
    aggs["duration", "max"] = aggs["duration"]["max"].apply(format_timedelta)
    aggs["duration", "min"] = aggs["duration"]["min"].apply(format_timedelta)
    aggs["duration", "<lambda>"] = aggs["duration"]["<lambda>"].apply(format_timedelta)
    result = [list(zip(counts.index.values, counts.values.tolist()))]

    inDf = pd.DataFrame(index=inMov)
    outDf = pd.DataFrame(index=outMov)
    inTotals = pd.DataFrame(temp.groupby(["newMovement"]).size())
    inTotals = inDf.merge(inTotals, how="left", left_index=True, right_index=True).fillna(0)
    outTotals = pd.DataFrame(temp.groupby(["outMovement"]).size())
    outTotals = outDf.merge(outTotals, how="left", left_index=True, right_index=True).fillna(0)
    result.append(inTotals[0].values.tolist())
    result.append(outTotals[0].values.tolist())

    ###
    ### output the journey pairs to csv
    ###
    temp.sort_values(by=["VRN"], inplace=True, ascending=[True])
    temp["duration"] = temp["duration"].apply(format_timedelta)
    temp = temp[["VRN", "Class", "newMovement", "Date", "outMovement", "outTime", "duration"]]
    try:
        temp.to_csv(outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Cordon - in-out non-directional " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y")
             + ".csv", header=["VRN", "Class", "In Movement", "Time", "Out Movement", "Time", "Duration"], index=False)
    except PermissionError as e:
        messagebox.showinfo(
            message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")

    ###
    ### set up the results for display
    ###

    resultsDict = {}

    result = list(zip(counts.index.values, counts.values.tolist()))
    for r in result:
        resultsDict[r[0]] = [r[1]]
    result = list(zip(aggs.index.values, aggs.values.tolist()))
    for r in result:
        #print("looking at ", r)
        for val in r[1]:
            #print("adding", val)
            resultsDict[r[0]].append(val)
    print(resultsDict)


    return [resultsDict, inTotals[0].values.tolist(), outTotals[0].values.tolist()]

def calculate_cordon_in_out_only(job,checkboxes):
    ###
    ### we want to find every journey for a vehicle where it appears at an "in" movement and the next movement it appears
    ### at is an "out" movement.
    ###
    global df
    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")
    inMov = []
    outMov = []
    for site, details in job["sites"].items():
        for mvmtNo, mvmt in details.items():
            if mvmt["dir"] == 1 or mvmt["dir"] == 3:
                inMov.append(int(mvmt["newmovement"]))
            if mvmt["dir"] == 2 or mvmt["dir"] == 3:
                outMov.append(int(mvmt["newmovement"]))

    ###
    ### set up some indexes so that if any sites have 0 values, we still pick up the sites in the dataframe
    ##
    inDf = pd.DataFrame(index=inMov)
    outDf = pd.DataFrame(index=outMov)

    ###
    ### process the data according to directional in-out cordon
    ###

    fullDf = df[datetime.datetime.strftime(job["surveydate"],"%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    print("fulldf")
    print(fullDf.head())
    for index in range(0,len(times)-1,2):
        print("Processing times",times[index],times[index+1])
        temp = fullDf.between_time(times[index],times[index+1],include_end=False)
        print("temp",temp.head())
        temp.index.name= "Date"
        temp.reset_index(inplace=True)
        temp.sort_values(by=["VRN","Date"], inplace=True, ascending=[True,True])
        mask = (temp["dir"]==1) & (temp["dir"].shift(-1)==2)
        if checkboxes[1]:
            mask = mask | (temp["dir"]==1) & (temp["dir"].shift(-1)==3)
        if checkboxes[2]:
            mask = mask | (temp["dir"] == 3) & (temp["dir"].shift(-1) == 2)
        if checkboxes[3]:
            mask = mask | (temp["dir"] == 3) & (temp["dir"].shift(-1) == 3)
        mask = (mask) & (temp["VRN"] == temp["VRN"].shift(-1))
        temp["matched"] = "N"
        temp["outTime"] = temp["Date"].shift(-1)
        temp["outMovement"] = temp["newMovement"].shift(-1)
        temp["newMovement"] = temp["newMovement"].real.astype(int)
        temp["outMovement"] = temp["outMovement"].real.astype(int)
        temp.ix[mask, "matched"] = "Y"
        dataframes.append(temp)
    temp = pd.concat(dataframes)
    try:
        temp.to_csv("dumped.csv")
    except PermissionError as e:
        print(e)

    temp = temp[temp["matched"] == "Y"]

    del temp["Movement"]
    del temp["Duplicates"]
    del temp["Site"]
    del temp["dir"]
    del temp["timeDiff"]
    del temp["matched"]
    temp["duration"] = temp["outTime"] -  temp["Date"]

    ###
    ### durations check
    ###
    print("before durations check, size is ", len(temp))
    dataframes = []
    if not job["durationsDictionary"] is None:
        for k, v in job["durationsDictionary"].items():
            i, o = k
            splitTime = v.split(":")
            hours = int(splitTime[0])
            mins = int(splitTime[1])
            td = datetime.timedelta(hours=hours, minutes=mins, seconds=0)
            mask = (temp["newMovement"] == i) & (temp["outMovement"] == o) & (temp["duration"] <= td)
            print("no of selections with mask is",len(temp[mask]))
            dataframes.append(temp[mask].copy())
        temp = pd.concat(dataframes)
    temp.sort_values(by=["VRN"], inplace=True, ascending=[True])
    counts = temp.groupby(["newMovement", "outMovement"]).size()
    aggs = temp.groupby(["newMovement", "outMovement"]).agg({"duration":[pd.DataFrame.max,pd.DataFrame.min,lambda x: sum(x,pd.Timedelta(0))/len(x) if len(x) > 0 else 0]})
    aggs["duration","max"] = aggs["duration"]["max"].apply(format_timedelta)
    aggs["duration", "min"] = aggs["duration"]["min"].apply(format_timedelta)
    aggs["duration", "<lambda>"] = aggs["duration"]["<lambda>"].apply(format_timedelta)
    inTotals = pd.DataFrame(temp.groupby(["newMovement"]).size())
    inTotals = inDf.merge(inTotals, how="left", left_index=True, right_index=True).fillna(0)
    outTotals = pd.DataFrame(temp.groupby(["outMovement"]).size())
    outTotals = outDf.merge(outTotals, how="left", left_index=True, right_index=True).fillna(0)
    resultsDict = {}

    result = list(zip(counts.index.values, counts.values.tolist()))
    for r in result:
        resultsDict[r[0]] = [r[1]]
    print(resultsDict)
    print(inTotals[0].values.tolist())
    result = list(zip(aggs.index.values, aggs.values.tolist()))
    for r in result:
        for val in r[1]:
            resultsDict[r[0]].append(val)
    temp["duration"] = temp["duration"].apply(format_timedelta)
    temp = temp[["VRN","Class","newMovement","Date","outMovement","outTime","duration"]]
    temp.sort_values(by=["VRN","Date"], inplace=True, ascending=[True,True])
    try:
        temp.to_csv(outputFolder + "/" + job["jobno"] +  " " + job["jobname"]  + " Cordon - in-out directional " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".csv",header=["VRN","Class","In Movement","Time","Out Movement","Time","Duration"],index=False)
    except PermissionError as e:
        messagebox.showinfo(message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")
    return [resultsDict,inTotals[0].values.tolist(),outTotals[0].values.tolist()]

def calculate_cordon_traversal_split_by_in_out(job,filters=None):
    ###
    ### Cordon Traversal
    ###

    global df, backgroundThread
    print("filters are ",filters)
    filtered = []
    outputFolder = os.path.join(job["folder"], "output")
    try:
        os.remove(outputFolder + "/" + job["jobno"] +  " " + job["jobname"]  + " Cordon Traversal - split by in-out " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".csv")
    except PermissionError as e:
        pass
    except FileNotFoundError as e:
        pass
    result = []
    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    for index in range(0,len(times)-1,2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        if len(temp)==0:
            continue
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])

        ###
        ### if there are plates that only occur once in the data, we can remove them, as we know they cant be a match
        ###
        temp = temp[temp.duplicated(subset=["VRN"], keep=False)]


        ###
        ### make a copy of the dataframe. first remove any intermediate sites, sort by VRN and date, then find a
        ### plate where its marked as an in, and the next occurence of the same plate is marked as an out
        ### we then have 1 full journey through the cordon. We set the in as a match ( matched = "Y")
        ###
        ###


        temp.reset_index(drop=True, inplace=True)
        temp["matched"] = "N"
        matchedDf = temp.copy()
        mask1 = (matchedDf["dir"] != 3) ### exclude all intermediate sites
        matchedDf = matchedDf[mask1]
        matchedDf["shifted"] = matchedDf["dir"].shift(-1)
        mask = ((matchedDf["dir"] == 1) & (matchedDf["shifted"] == 2) & (temp["VRN"] == temp["VRN"].shift(-1)))
        matchedDf.ix[mask, "matched"] = "Y"
        matchedDf = matchedDf[["matched"]]

        ###
        ### matchedDf now has a column "matched" that has the selected journeys, we merge it back into the main
        ### dataframe, based on the index
        ###

        temp = temp.merge(matchedDf, how="left", left_index=True, right_index=True)
        del temp["matched_x"]
        print("---")
        print(temp.head())
        temp["newMovement"] = temp["newMovement"].astype(int)

        ###
        ### group the dataframe by plate, then aggregate each group into comma separated fields
        ###

        strJoin = lambda x: ",".join(x.astype(str))
        dateJoin = lambda x: ",".join(x.apply(date_to_time))
        temp = temp.groupby(["VRN", "Class"]).agg({"Date": dateJoin, "newMovement": strJoin,"matched_y":strJoin})
        temp.reset_index(inplace=True)
        temp = temp[["VRN","Class","newMovement","Date","matched_y"]]
        values=temp.values.tolist()

        print()
        temp.sort_values(by=["VRN"],inplace=True)
        print(temp)
        ###
        ### process the comma separated fields so that we split up each journey between an in and an out
        ###


        count = 0
        for v in values:
            for i in range(2, 5):
                v[i] = [item for item in v[i].split(",")]
            index = 0
            while index < len(v[4]):
                journey=[v[0],v[1]]
                selectedJourney = [] ### to keep a track of the split journey
                try:
                    index = v[4][index:].index("Y") + index
                    while v[4][index] != "N":
                        journey.append(v[2][index])
                        journey.append(v[3][index])
                        selectedJourney.append(v[2][index])
                        index+=1
                except ValueError as e:
                    break
                except IndexError as e:
                    break
                ###
                ### need to append again, as we were looking for an "N" and exiting the loop when we found it
                ###
                journey.append(v[2][index])
                journey.append(v[3][index])
                selectedJourney.append(v[2][index]) ### selectedJourney now holds the list of movements for a cordon traversal
                if not filters is None:
                    if list(map(int, selectedJourney)) in filters:
                        filtered.append(journey)
                        count += 1
                        print()
                else:
                    filtered.append(journey)



    ###
    ### durations check
    ###

    print(filtered[:20])

    result = []
    for journey in filtered:
        diffs = [(int(item),int(journey[i+2]),datetime.datetime.strptime(journey[i+3],"%H:%M:%S")- datetime.datetime.strptime(journey[i+1],"%H:%M:%S")) for i,item in enumerate(journey) if i%2==0 and i >=2 and i< len(journey)-2]
        flag = True
        for d in diffs:
            splitTime = job["durationsDictionary"][d[0],d[1]].split(":")
            hours = int(splitTime[0])
            mins = int(splitTime[1])
            td = datetime.timedelta(hours=hours, minutes=mins, seconds=0)
            if d[2] > td:
                flag=False
        if flag:
            result.append(journey)


    print(result[:20])

    try:
        with open(outputFolder + "/" + job["jobno"] +  " " + job["jobname"]  + " Cordon Traversal - split by in-out " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(result)
    except PermissionError as e:
        messagebox.showinfo(
            message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")

def calculate_route_assignment_fs_ls(job,filters=None):
    global df,backgroundThread
    ###
    ### we simply want to document the very first time a vehicle was seen, and the very last time it was seen
    ### hence - first seen/last seen
    ### In this case, we dont care about directions ( ie in, or out, or intermediate), we just care about the very
    ### first time it was seen, and the very last time
    ###

    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")
    ###
    ### set up a background thread to process and run the full routes calculation, since that is time consuming
    ### if it has already been run ( full route data.pkl exists in the folder) then we dont need to run it again
    ###
    if backgroundThread is None:
        print("thread is None")
    else:
        print(backgroundThread.is_alive())
    if not os.path.isfile(dataFolder + "/complete routes data.pkl"):
        if backgroundThread is None or not backgroundThread.is_alive():
            print("starting up thread")
            backgroundThread = threading.Thread(target=calculate_route_assignment_full_routes,args=(job,filters))
            backgroundThread.start()
    else:
        print("already exists")

    inMov = []
    outMov = []

    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    for index in range(0, len(times) - 1, 2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])

        ###
        ### if there are plates that only occur once in the data, we can remove them, as we know they cant be a match
        ###
        print("before removing singletons",len(temp))
        temp = temp[temp.duplicated(subset=["VRN"], keep=False)]
        print("after removing",len(temp))

        ###
        ### find first occurence of a plate. We dont care about direction
        ###
        grp = temp.groupby(["VRN"])
        result1 = temp[grp.cumcount() == 0]
        print(result1.head())
        ###
        ### find last occurence of a plate
        ###
        result2 = temp[grp.cumcount(ascending=False) == 0]
        print(result2.head())
        fullResult = pd.concat([result1, result2])
        print(fullResult)
        fullResult.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
        print(fullResult.head())
        fullResult["outTime"] = fullResult["Date"].shift(-1)
        fullResult["newMovement"] = fullResult["newMovement"].real.astype(int)
        fullResult["outMovement"] = fullResult["newMovement"].shift(-1)
        fullResult["outMovement"] = fullResult["outMovement"].real.astype(int)
        fullResult=fullResult.iloc[::2]
        print(fullResult.head())
        dataframes.append(fullResult)

    temp = pd.concat(dataframes)
    temp["duration"] = temp["outTime"] - temp["Date"]

    counts = temp.groupby(["newMovement", "outMovement"]).size()
    aggs = temp.groupby(["newMovement", "outMovement"]).agg({"duration": [pd.DataFrame.max, pd.DataFrame.min,lambda x: sum(x, pd.Timedelta(0)) / len(x) if len(x) > 0 else 0]})
    aggs["duration", "max"] = aggs["duration"]["max"].apply(format_timedelta)
    aggs["duration", "min"] = aggs["duration"]["min"].apply(format_timedelta)
    aggs["duration", "<lambda>"] = aggs["duration"]["<lambda>"].apply(format_timedelta)


    for site, details in job["sites"].items():
        for mvmtNo, mvmt in details.items():
            if not int(mvmt["newmovement"]) in inMov:
                inMov.append(int(mvmt["newmovement"]))
            if not int(mvmt["newmovement"]) in outMov:
                outMov.append(int(mvmt["newmovement"]))
    inMov = sorted(inMov)
    outMov = sorted(outMov)

    inDf = pd.DataFrame(index=inMov)
    outDf = pd.DataFrame(index=outMov)
    inTotals = pd.DataFrame(temp.groupby(["newMovement"]).size())
    inTotals = inDf.merge(inTotals, how="left", left_index=True, right_index=True).fillna(0)
    outTotals = pd.DataFrame(temp.groupby(["outMovement"]).size())
    outTotals = outDf.merge(outTotals, how="left", left_index=True, right_index=True).fillna(0)

    resultsDict = {}
    result = list(zip(counts.index.values, counts.values.tolist()))
    for r in result:
        resultsDict[r[0]] = [r[1]]
    result = list(zip(aggs.index.values, aggs.values.tolist()))
    for r in result:
        for val in r[1]:
            resultsDict[r[0]].append(val)


    temp["duration"] = temp["duration"].apply(format_timedelta)
    del temp["Movement"]
    del temp["Duplicates"]
    del temp["Site"]
    del temp["dir"]
    del temp["timeDiff"]
    temp = temp[["VRN", "Class", "newMovement", "Date", "outMovement", "outTime", "duration"]]
    try:
        temp.to_csv(outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Route Assignment - first seen last seen " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".csv",
                    header=["VRN", "Class", "In Movement", "Time", "Out Movement", "Time", "Duration"],index=False)
    except PermissionError as e:
        messagebox.showinfo(message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")

    return [resultsDict, inTotals[0].values.tolist(), outTotals[0].values.tolist()]

def calculate_route_assignment_journey_pairs(job):
    global df
    inMov = []
    outMov = []
    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")
    for site, details in job["sites"].items():
        for mvmtNo, mvmt in details.items():
            if mvmt["newmovement"] not in inMov:
                inMov.append(mvmt["newmovement"])
            if mvmt["newmovement"] not in outMov:
                outMov.append(mvmt["newmovement"])
    inMov = sorted(inMov)
    outMov = sorted(outMov)
    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    for index in range(0, len(times) - 1, 2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
        mask = (temp["VRN"] == temp["VRN"].shift(-1))
        temp["matched"] = "N"
        temp.ix[mask, "matched"]  = "Y"
        temp["outTime"] = temp["Date"].shift(-1)
        temp["outMovement"] = temp["newMovement"].shift(-1)
        temp["newMovement"] = temp["newMovement"].real.astype(int)
        temp["outMovement"] = temp["outMovement"].real.astype(int)

        dataframes.append(temp)
    temp = pd.concat(dataframes)

    ###
    ### we now have all journey pairs
    ###
    temp = temp[temp["matched"] == "Y"]
    temp["duration"] = temp["outTime"] - temp["Date"]



    counts = temp.groupby(["newMovement", "outMovement"]).size()
    aggs = temp.groupby(["newMovement", "outMovement"]).agg({"duration": [pd.DataFrame.max, pd.DataFrame.min,lambda x: sum(x, pd.Timedelta(0)) / len(x) if len(x) > 0 else 0]})
    aggs["duration", "max"] = aggs["duration"]["max"].apply(format_timedelta)
    aggs["duration", "min"] = aggs["duration"]["min"].apply(format_timedelta)
    aggs["duration", "<lambda>"] = aggs["duration"]["<lambda>"].apply(format_timedelta)

    inDf = pd.DataFrame(index=inMov) ### silly to name inDf similar to a name used earlier....
    outDf = pd.DataFrame(index=outMov)
    inTotals = pd.DataFrame(temp.groupby(["newMovement"]).size())
    inTotals = inDf.merge(inTotals, how="left", left_index=True, right_index=True).fillna(0)
    outTotals = pd.DataFrame(temp.groupby(["outMovement"]).size())
    outTotals = outDf.merge(outTotals, how="left", left_index=True, right_index=True).fillna(0)

    resultsDict = {}
    result = list(zip(counts.index.values, counts.values.tolist()))
    for r in result:
        resultsDict[r[0]] = [r[1]]
    result = list(zip(aggs.index.values, aggs.values.tolist()))
    for r in result:
        for val in r[1]:
            resultsDict[r[0]].append(val)
    print(resultsDict)

    temp["duration"] = temp["duration"].apply(format_timedelta)
    del temp["Movement"]
    del temp["Duplicates"]
    del temp["Site"]
    del temp["dir"]
    del temp["timeDiff"]
    temp = temp[["VRN", "Class", "newMovement", "Date", "outMovement", "outTime", "duration"]]
    try:
        temp.to_csv(outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Route Assignment - journey pairs " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".csv",
                    header=["VRN", "Class", "In Movement", "Time", "Out Movement", "Time", "Duration"], index=False)
    except PermissionError as e:
        messagebox.showinfo(
            message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")
    print("in totals are ",inTotals[0].values.tolist())
    return [resultsDict, inTotals[0].values.tolist(), outTotals[0].values.tolist()]

def calculate_route_assignment_full_routes(job,filters):
    ###
    ### each vehicle enters the cordon at a site, travels through a number of sites, and exits at a site
    ### we want to track and output the full journey taken by each vehicle, recording each movement it passed through
    ### and the time seen at that movement
    ###
    global df,backgroundThread
    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")

    ###
    ### we also want to calculate the journeys, with the full journeys split by any in-outs
    ### ie Cordon Traversal
    ###

    calculate_cordon_traversal_split_by_in_out(job, filters)

    ###
    ### now do the full routes
    ###


    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    result = []
    for index in range(0,len(times)-1,2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
        print(temp.head(20))
        ###
        ### if there are plates that only occur once in the data, we can remove them, as we know they cant be a match
        ###
        print("before removing singletons", len(temp))
        temp = temp[temp.duplicated(subset=["VRN"], keep=False)]
        print("after removing", len(temp))

        ###
        ### group by VRN, and then join each group together into 1 row in a dataframe
        ### giving us the full journey travelled by that vehicle
        ###

        strJoin = lambda x: ",".join(x.astype(str))
        dateJoin = lambda x: ",".join(x.apply(date_to_time))
        temp = temp.groupby(["VRN", "Class"]).agg({"newMovement": strJoin, "Date": dateJoin})
        temp.to_pickle(dataFolder + "/all journey pairs.pkl")
        temp.reset_index(inplace=True)

        values = temp.values.tolist()

        for v in values:
            for i in range(2, 4):
                v[i] = [item for item in v[i].split(",")]
            l = [item for sublist in list(zip(*[v[2], v[3]])) for item in sublist]
            l.insert(0, l[-1])
            l.insert(0, l[-2])
            l.insert(0, l[3])
            l.insert(0, l[3])
            l.insert(0, v[1])
            l.insert(0, v[0])
            l.insert(6, "")
            result.append(l)


        #dataframes.append(temp)
    #temp = pd.concat(dataframes)



    ###
    ### output the data to csv
    ###
    try:
        with open(outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Route Assignment - all full journeys " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".csv" , "w",newline="") as f:
            writer = csv.writer(f)
            writer.writerows(result)
    except PermissionError as e:
        messagebox.showinfo(
            message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")
    print("finished thread")
    backgroundThread = None

def produce_full_routes(job):
    ###
    ### each vehicle enters the cordon at a site, travels through a number of sites, and exits at a site
    ### we want to track and output the full journey taken by each vehicle, recording each movement it passed through
    ### and the time seen at that movement
    ###
    global df,backgroundThread
    inMov = []
    outMov = []
    outputFolder = os.path.join(job["folder"], "output")
    dataFolder = os.path.join(job["folder"], "data")
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]

    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    dataframes = []
    result = []
    for index in range(0, len(times) - 1, 2):
        print("checking times",times[index],times[index+1])
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp = temp[temp["Class"].notnull()]
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])

        ###
        ### if there are plates that only occur once in the data, we can remove them, as we know they cant be a match
        ###
        print("before removing singletons", len(temp))
        temp = temp[temp.duplicated(subset=["VRN"], keep=False)]
        print("after removing", len(temp))


        ###
        ### group by VRN, and then join each group together into 1 row in a dataframe
        ### giving us the full journey travelled by that vehicle
        ###

        strJoin = lambda x: ",".join(x.astype(str))
        dateJoin = lambda x: ",".join(x.apply(date_to_time))
        dirJoin = lambda x: ",".join(x.apply(dir_to_str))
        temp = temp.groupby(["VRN", "Class"]).agg({"newMovement": strJoin, "Date": dateJoin,"dir":dirJoin})
        temp.to_pickle(dataFolder + "/all full routes.pkl")
        temp.reset_index(inplace=True)
        temp = temp[["VRN","Class","Date","newMovement","dir"]]
        values = temp.values.tolist()
        print(values[:2])
        for v in values:
            result.append([v[0],v[1],list(zip(*[item.split(",") for item in v[2:]]))])

    with open(job["folder"] + "/data/all journeys as list.pkl", "wb") as f:
        pickle.dump(result, f)
    print(result[:2])
    print("finished thread")
    backgroundThread = None

def calculate_regex_matching(job,filters,durationCheck,durationBehaviour):
    dataFolder = os.path.join(job["folder"], "data")
    outputFolder = os.path.join(job["folder"], "output")
    journeys = pd.read_pickle(dataFolder + "/all journeys as list.pkl")

    inMov = []
    outMov = []
    bothMov = []
    for site, details in job["sites"].items():
        for mvmtNo, mvmt in details.items():
            if mvmt["dir"] ==1:
                inMov.append(mvmt["newmovement"])
            if mvmt["dir"] ==2:
                outMov.append(mvmt["newmovement"])
            if mvmt["dir"] ==3:
                bothMov.append(mvmt["newmovement"])
    inMov = sorted(inMov)
    outMov = sorted(outMov)
    bothMov = sorted(bothMov)
    if bothMov == []:
        bothMov = [10000]






    print("starting first method ",datetime.datetime.now())
    result = []
    for journey in journeys:
        data = journey[2]
        for f in filters:
            matches = anprregex.match(data,f)
            for m in matches:
                output = []
                output.append(journey[0])
                output.append(journey[1])
                temp =([(item[1],item[0]) for item in m])
                temp = [item for sublist in temp for item in sublist]
                [output.append(item) for item in temp]
                if not output in result:
                    result.append(output)

    print("finishing first method",datetime.datetime.now())

    if durationCheck:
        if not job["durationsDictionary"] is None:
            for journey in result:
                #print("checking journey",journey)
                start=2
                while start < len(journey) -2:
                    #print("start is",start,len(journey) -1)
                    duration = datetime.datetime.strptime(journey[start + 3], "%H:%M:%S") - datetime.datetime.strptime(journey[start + 1], "%H:%M:%S")
                    #print("duration is",duration,(int(journey[start]),int(journey[start+2])),job["durationsDictionary"][(int(journey[start]),int(journey[start+2]))])
                    v = job["durationsDictionary"][(int(journey[start]),int(journey[start+2]))]
                    splitTime = v.split(":")
                    hours = int(splitTime[0])
                    mins = int(splitTime[1])
                    td = datetime.timedelta(hours=hours, minutes=mins, seconds=0)
                    if duration > td:
                        if durationBehaviour ==1: ## split any journeys where a leg exceeds the duration
                            newJourney = [journey[0], journey[1]]
                            [newJourney.append(item) for item in journey[start + 2:]]
                            while len(journey) > start + 2:
                                del journey[-1]
                            if len(newJourney) >4:
                                result.append(newJourney)
                            if len(journey) < 5:
                                while len(journey) > 0:
                                    del journey[-1]
                            #print("journey is now",journey,"added journey",newJourney)
                        else: ### discard any journeys where a leg exceeds the duration
                            while len(journey) > 0:
                                del journey[-1]
                    start+=2
        result = [item for item in result if item !=[]]


    ###
    ### convert the filters into a format useable with pythons regex module
    ###

    convertedFilters = []
    for f in filters:
        tokens = f.split("-")
        for i, t in enumerate(tokens):
            tokens[i] = tokens[i].replace("B", "(" + "|".join(map(str, bothMov)) + ")")
            tokens[i] = tokens[i].replace("I", "(" + "|".join(map(str, inMov)) + ")")
            tokens[i] = tokens[i].replace("O", "(" + "|".join(map(str, outMov)) + ")")

        for i, t in enumerate(tokens):
            if "*" in t:
                print(t)
                tokens[i] = tokens[i].replace("*", "")
                tokens[i] = "(" + tokens[i] + "(?=,|\Z))*"
                #tokens[i] = "(," + tokens[i] + ")*"
                print(tokens[i])
            else:
                if i !=0:
                    tokens[i] = "," + tokens[i]
        if not "^" in f:
            tokens[0] = "(^|,)" + tokens[0]
        else:
            tokens[0] = tokens[0].replace("^", "")
            tokens[0] = "^" + tokens[0]
        if not "!" in f:
            tokens[-1] = tokens[-1] + r"(?=,|\Z)"
        else:
            tokens[-1] = tokens[-1].replace("!", "")
            tokens[-1] = tokens[-1] + "$"
        print("converted filter is", tokens)
        convertedFilters.append("".join(tokens))


    try:
        with open(outputFolder + "/" + job["jobno"] + " " + job[
            "jobname"] + " Filtered Matching - all matches " + datetime.datetime.strftime(job["surveydate"],"%d-%m-%Y") + ".csv","w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(result)
    except PermissionError as e:
        messagebox.showinfo(
            message="Couldnt write plates to csv, file is already open. Run procedure again after closing csv file")



    resultsDict ={}
    for r in result:
        resultsDict[(int(r[2]), int(r[-2]))] = resultsDict.get((int(r[2]), int(r[-2])), [])
        resultsDict[(int(r[2]), int(r[-2]))].append(datetime.datetime.strptime(r[-1], "%H:%M:%S") - datetime.datetime.strptime(r[3], "%H:%M:%S"))

    inMov = []
    outMov = []
    for key,item in resultsDict.items():
        if not key[0] in inMov:
            inMov.append(key[0])
        if not key[1] in outMov:
            outMov.append(key[1])

    ###
    ### set up some indexes so that if any sites have 0 values, we still pick up the sites in the dataframe
    ##
    inDf = pd.DataFrame(index=inMov)
    outDf = pd.DataFrame(index=outMov)

    inDf["count"] = 0
    outDf["count"] = 0
    for key,item in resultsDict.items():
        resultsDict[key] = [len(item),max(item),min(item),format_timedelta(np.mean(item))]
        inDf.loc[int(key[0])]["count"]+=len(item)
        outDf.loc[int(key[1])]["count"] += len(item)
    inDf.sort_index(inplace=True)
    outDf.sort_index(inplace=True)
    return(resultsDict,inDf["count"].tolist(),outDf["count"].tolist())

def calculate_overtaking(job,movementPair):
    global df
    mov1,mov2 = movementPair
    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")]
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    for index in range(0, len(times) - 1, 2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] >= 0]
        temp.sort_values(by=["VRN","Date"], inplace=True, ascending=[True,True])
        temp["outTime"] = temp["Date"].shift(-1)
        temp["outMovement"] = temp["newMovement"].shift(-1)
        temp["newMovement"] = temp["newMovement"].real.astype(int)
        temp["outMovement"] = temp["outMovement"].real.astype(int)

        ###
        ### do the matching, so that we have all vehicles that made journeys between mov1 and mov2
        ###

        mask = ((temp["newMovement"] == mov1) & (temp["outMovement"] == mov2) & (temp["VRN"] == temp["VRN"].shift(-1)))
        temp["matched"] = "N"
        #temp["matched"][mask] = "Y"
        temp.ix[mask, "matched"] = "Y"
        dataframes.append(temp)



    temp = pd.concat(dataframes)
    try:
        temp.to_csv("dumped.csv")
    except Exception as e:
        pass
    temp = temp[temp["matched"] == "Y"]
    try:
        temp.to_csv("matched.csv")
    except Exception as e:
        pass
    if len(temp) == 0:
        return [[],[]]
    temp["duration"] = temp["outTime"] - temp["Date"]

    ###
    ### need to bin up the durations of each journey, as we want to display the number of journeys that fall in
    ### to each bin. Each bin is 15 seconds length, and starts from the average duration of all the journeys
    ###

    avDuration = temp["duration"].mean()
    avDuration= bin_time(avDuration)
    bins = avDuration + np.arange(24) * datetime.timedelta(seconds=15)
    temp["bin"] = temp["duration"].apply(lambda x: bin_time(x))
    bins = [format_timedelta(item) for item in bins]

    binnedData = dict((k,0) for k in bins)
    for k,v in temp["bin"].value_counts().iteritems():
        key = format_timedelta(k)
        if key in bins:
            binnedData[key] = v

    ###
    ### sort the dataframe by the time seen at mov1
    ### then make the index of this sorted dataframe as the "In Order" column, which shows what order
    ### vehicles arrived at mov1
    ###

    temp.sort_values(by=["Date"], inplace=True, ascending=[True])
    temp.reset_index(inplace=True,drop=True)
    temp.index += 1
    temp.index.name = "In Order"
    temp.reset_index(inplace=True)

    ###
    ### sort the dataframe by the time seen at mov2
    ### then make the index of this sorted dataframe as the "Out Order" column, which shows what order
    ### vehicles arrived at mov2
    ###

    temp.sort_values(by=["outTime"], inplace=True, ascending=[True])
    temp.reset_index(inplace=True,drop=True)
    temp.index += 1
    temp.index.name = "Out Order"
    temp.reset_index(inplace=True)


    ###
    ### resort the dataframe by date, we now have the dataframe in order of time seen at mov1.
    ### we can now calculate the number of vehicles overtaking and overtaken by each vehicle
    ###

    temp.sort_values(by=["Date"], inplace=True, ascending=[True])
    temp.set_index(["Date"], inplace=True)
    return [temp, binnedData]

def resample_overtaking_data(job,df,time_as_string,pair):
    ###
    ### takes a previously calculated dataframe containing overtaking data for a single route
    ### and a max duration, time_as_string, eg "00:02:40"
    ### restricts the data frame by selecting all durations less than time_as_string
    ### and returns the data ready for display
    ###
    ###

    if len(df) ==0:
        return []
    outputFolder = os.path.join(job["folder"], "output")
    mov1, mov2 = pair
    d = datetime.datetime.strptime(time_as_string,"%H:%M:%S")
    dt = datetime.timedelta(seconds=d.second,minutes=d.minute,hours=d.hour)
    temp = df.copy(deep=True)
    del temp["In Order"]
    del temp["Out Order"]
    temp.reset_index(inplace=True)
    temp = temp[temp["duration"] <= dt]

    ###
    ### sort the dataframe by the time seen at mov1
    ### then make the index of this sorted dataframe as the "In Order" column, which shows what order
    ### vehicles arrived at mov1
    ###

    temp.sort_values(by=["Date"], inplace=True, ascending=[True])
    temp.reset_index(inplace=True, drop=True)
    temp.index += 1
    temp.index.name = "In Order"
    temp.reset_index(inplace=True)

    ###
    ### sort the dataframe by the time seen at mov2
    ### then make the index of this sorted dataframe as the "Out Order" column, which shows what order
    ### vehicles arrived at mov2
    ###

    temp.sort_values(by=["outTime","Date"], inplace=True, ascending=[True,True])
    temp.reset_index(inplace=True, drop=True)
    temp.index += 1
    temp.index.name = "Out Order"
    temp.reset_index(inplace=True)

    ###
    ### resort the dataframe by date, we now have the dataframe in order of time seen at mov1.
    ### we can now calculate the number of vehicles overtaking and overtaken by each vehicle
    ###

    temp.sort_values(by=["Date","outTime"], inplace=True, ascending=[True,True])
    outorder = temp["Out Order"].values.tolist()
    #print("outorder is",outorder)
    #l = []
    #for i,item in enumerate(outorder):
        #lb = i-20
        #if lb <0:
            #lb= 0
        #l.append(len([x for x in outorder[lb:i] if x > item]))
    #print("l is",l)
    temp["overtook"] = 0
    temp["overtaken by"]=0
    temp["overtook"]  = temp["In Order"].apply(lambda x: find_number_overtaking_vehicle(x - 1,outorder))
    temp["overtaken by"] = temp["In Order"].apply(lambda x: find_number_overtaken_by_vehicle(x - 1,outorder))
    temp["manouvres"] = temp["overtook"]
    try:
        temp.to_csv("dumped.csv")
    except Exception as e:
        print(e)
    outDf = temp.copy(deep=True)
    outDf = outDf[["VRN", "In Order", "Date", "Out Order", "outTime", "overtook", "overtaken by", "duration", "manouvres"]]
    outDf = outDf[["VRN", "In Order", "Date", "Out Order", "outTime", "duration"]]
    outDf["Date"] = outDf["Date"].apply(date_to_time)
    outDf["outTime"] = outDf["outTime"].apply(date_to_time)

    ###
    ### write to excel
    ###

    file = outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Overtaking " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".xlsx"

    outDf.to_excel(file, sheet_name="Movement " + str(mov1) + "-" + str(mov2),index=False,header=["VRN","In Order","Time","Out Order","Time","Duration"])


    temp.set_index(["Date"], inplace=True)
    temp = temp[["VRN","overtook", "overtaken by", "duration", "manouvres"]]
    temp = temp[["VRN","duration", "manouvres", "overtaken by","overtook"]]


    resampled = temp.resample("60T").apply({"VRN": "count", "overtook": lambda x: int((x!=0).sum()), "overtaken by": lambda x: int((x!=0).sum()),"duration": lambda x: pd.to_timedelta(x).mean(), "manouvres": np.sum})
    resampled.fillna(0, inplace=True)
    resampled["manouvres"]=resampled["manouvres"].astype(int)
    resampled["duration"] = resampled["duration"].apply(format_timedelta)


    ###
    ### set up a base dataframe spanning the whole survey period, ready to be merged with the resampled dataframe
    ### so that any periods that dont have any data, are still reported, but with 0's
    ###
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    d = datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")
    startTime = datetime.datetime.strptime(times[0], "%H:%M").replace(minute=0, second=0)
    endTime = datetime.datetime.strptime(times[-1], "%H:%M")
    if endTime.minute > 0:
        endTime.replace(hour=endTime.hour + 1, minute=0, second=0)
    rng = pd.date_range(d + " " + startTime.strftime("%H:%M"), d + " " + endTime.strftime("%H:%M"), freq="60T", closed="left")
    indexDf = pd.DataFrame(index=rng)

    ###
    ### merge the resampled dataframe with the index dataframe

    resampled = resampled.merge(indexDf, how="outer", left_index=True, right_index=True)

    print(resampled.head())
    ###
    ### reset index so that Date is a column again
    ###

    #
    resampled.index.name = "Date"
    resampled.reset_index(inplace=True)
    resampled.fillna(0)
    resampled["Date"] = resampled["Date"].apply(lambda x: datetime.datetime.strftime(x,"%H:%M"))
    resampled["speed"] = 0
    resampled = resampled[["Date","VRN","duration","speed","manouvres","overtaken by","overtook"]]
    return resampled.values.tolist()

def find_number_overtaken_by_vehicle(index,l):
    ###
    ### given a list, and an index into that list , we take a slice of size 20 from the index backwards, and
    ### count the number of values in the slice that are greater than the value at [index]
    ###

    lb = index - 20
    if lb<0:
        lb = 0
    value = l[index]
    return  len([x for x in l[lb:index] if x > value])

def find_number_overtaking_vehicle(index,l):
    ###
    ### given a list, and an index into that list , we take a slice of size 20 from the index, and
    ### count the number of values in the slice that are less than the value at [index]
    ###
    if index>=len(l):
        return 0
    ub = index + 20
    if ub > len(l):
        ub = len(l)
    value = l[index]
    return len([x for x in l[index:ub] if x < value])

def get_platoon(dt,platooningTime):
    ###
    ### simple function to return the platoon number
    ### this is used by apply in the dataframe
    ###
    global platoon
    if dt.total_seconds() <=platooningTime:
        return platoon
    else:
        platoon+=1
        return platoon

def platooning_resample_method(s):
    ###
    ### s is a pandas Series object
    ###

    if len(s) > 0:
        c = s.value_counts()
        return list(zip(c.axes[0], c.values))
    else:
        return []

def calculate_platooning(job,movement,platooningTime):
    global df,platoon
    result = []
    platoon = 0
    outputFolder = os.path.join(job["folder"], "output")
    fullDf = df[datetime.datetime.strftime(job["surveydate"], "%Y-%m-%d")].copy()
    fullDf = fullDf[fullDf["Class"].notnull()]
    times = [x for x in job["timeperiod1"].split("-") + job["timeperiod2"].split("-") + job["timeperiod3"].split("-")
             + job["timeperiod4"].split("-") if x != ""]
    dataframes = []
    summary = []
    for index in range(0, len(times) - 1, 2):
        temp = fullDf.between_time(times[index], times[index + 1], include_end=False)
        temp.index.name = "Date"
        temp.reset_index(inplace=True)
        temp = temp[temp["newMovement"] == movement]
        temp.sort_values(by=["Date"], inplace=True, ascending=[True])
        temp.reset_index(inplace=True,drop=True)
        temp["timeDiff"] = temp["Date"].shift(-1) - temp["Date"]
        temp["timeDiff"] = temp["timeDiff"].shift(1)
        #temp["timeDiff"].iloc[0] = datetime.timedelta(seconds=10000)
        dataframes.append(temp)

        temp["platoon"] = 0
        temp["platoon"] = temp["timeDiff"].apply(lambda x: get_platoon(x,platooningTime))
        strJoin = lambda x: ",".join(x.astype(str))
        dateJoin = lambda x: ",".join(x.apply(date_to_time))
        temp = temp.groupby("platoon").agg({"Date": dateJoin,"Class":strJoin,"VRN":strJoin})
        mask = temp["Class"].apply(lambda x: len(x.split(",")) > 1)
        temp = temp[mask]
        temp["length"] = 0
        temp["length"] = temp["VRN"].apply(lambda x: len(x.split(",")))
        temp["time"] = temp["Date"].apply(lambda x: pd.to_datetime(x.split(",")[0]))
        temp = temp[["VRN","Date","Class","length","time"]]
        values = temp.values.tolist()
        for v in values:
            output = []
            plates = v[0].split(",")
            t = v[1].split(",")
            classes = v[2].split(",")
            output.append(plates[0])
            output.append(classes[0])
            output.append(t[0])
            for c in classes[1:]:
                output.append(c)
            result.append(output)

        ###
        ### format the results into a summary table for display in the app
        ###

        startTime = datetime.datetime.strptime(times[index], "%H:%M").replace(minute=0,second=0)
        endTime = datetime.datetime.strptime(times[index + 1], "%H:%M")
        if endTime.minute >0:
            endTime.replace(hour=endTime.hour+1,minute=0,second=0)

        ###
        ### set up the summary list for the platooning info
        ### the list has the time, and then 0's for each entry
        ###

        output = []
        while startTime < endTime:
            l = [date_to_time(startTime)]
            for i in range(1, 11):
                l.append(0)
            startTime = startTime + datetime.timedelta(hours=1)
            summary.append(l)

        ###
        ### resample the data into 1 hour bins
        ###

        temp.set_index(["time"], inplace=True)


        ### new method
        temp["length"]= temp["VRN"].apply(lambda x:len(x.split(",")))
        resampled = temp.resample("1H").apply({"length": lambda x: platooning_resample_method(x)})
        resampled.reset_index(inplace=True)
        print(resampled.info())
        resampled = resampled.values.tolist()


        ###
        ###  fill the data into our output list
        ###
        for i, row in enumerate(resampled):
            print(row,row[0],row[1])
            t, data = row[0], row[1]
            print(t, data)
            for o in summary:
                print(o[0],date_to_time(t))
                if o[0] == date_to_time(t):
                    print("match!")
                    for item in data:
                        if item[0] < 11:
                            o[item[0] - 1] = item[1]
                        else:
                            o[10] += item[1]


    print("Summary is",summary)
    #temp = pd.concat(dataframes)



    ###
    ### write the results to excel
    ###
    filename = outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Platooning " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".xlsx"

    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError as e:
        wb = openpyxl.Workbook(optimized_write=True)
    try:
        sht = wb.get_sheet_by_name("Movement " + str(movement))
        tempSheet = wb.create_sheet(title="temp")
        wb.remove_sheet(sht)
        tempSheet.title = "Movement " + str(movement)
        sht = tempSheet
    except KeyError as e:
        sht=wb.create_sheet(title="Movement " + str(movement))
    for r in result:
        sht.append(r)
    try:
        wb.save(filename=outputFolder + "/" + job["jobno"] + " " + job["jobname"] + " Platooning " + datetime.datetime.strftime(job["surveydate"], "%d-%m-%Y") + ".xlsx" )
    except PermissionError as e:
        messagebox.showinfo(message="Platooning Excel file already open, cant save,please try again")


    total = [sum(i) for i in list(zip(*summary))[1:]]
    total.insert(0,"Total")
    summary.append(total)
    return summary



file = "C:/Users/NWatson/Desktop/ANPR data/3279-Lon, Oxford_Unclassed_Plates_9d3f3926-36ba-47b2-bda9-70516b735a874085337599898377328 (4).xlsx"
#file = "C:/Users/NWatson/Desktop/ANPR data/test.xlsx" ## cut down version of above file


df = None
overviewDf = None
backgroundThread = None


def test():
    global df,overviewDf





    myDB.set_file("C:/Users/NWatson/PycharmProjects/ANPR/blah.sqlite")
    job = myDB.load_job("3105-IRE","Coldcut","07/10/16")
    load_job(job)
    #produce_full_routes(job)
    calculate_regex_matching(job,["I-(B-O)"])
    exit()
    print("lenth of df BEFORE plate restrictions", len(df))
    job["platerestrictionpercentages"] = []

    if len(df) != 0:
        job["platerestrictionpercentages"].append(100)
        print(len(df[(df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 7)]))
        job["platerestrictionpercentages"].append(len(df[(df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 7)])*100 / len(df))
        job["platerestrictionpercentages"].append(len(df[(df["VRN"].str.len() >= 5) & (df["VRN"].str.len() <= 7)])*100 / len(df))
        job["platerestrictionpercentages"].append(len(df[(df["VRN"].str.len() >= 4) & (df["VRN"].str.len() <= 8)])*100 / len(df))

    print(job["platerestrictionpercentages"])
    if job["platerestriction"] == 1:
        return True
    if job["platerestriction"] == 2:
        mask = (df["VRN"].len() >= 4) & (df["VRN"].len() <= 7)
    if job["platerestriction"] == 3:
        mask = (df["VRN"].len() >= 5) & (df["VRN"].len() <= 7)
    if job["platerestriction"] == 4:
        mask = (df["VRN"].len() >= 4) & (df["VRN"].len() <= 8)
    df = df[mask]
    print("lenth of df after plate restrictions", len(df))

    exit()


    #calculate_route_assignment_full_routes(job)
    df.index.name = "Date"
    df.reset_index(inplace=True)
    df.sort_values(by=["VRN","Date"], inplace=True, ascending=[True,True])
    df.reset_index(drop=True,inplace=True)
    df["matched"] = "N"

    temp = df.copy()
    mask1 = (temp["dir"] != 3)
    temp = temp[mask1]
    temp["shifted"] = temp["dir"].shift(-1)
    mask = ((temp["dir"] == 1) & (temp["shifted"] == 2))  # & (temp["VRN"] == temp["VRN"].shift(-1)))

    #print(temp[mask].head())
    temp.ix[mask,"matched"] ="Y"
    temp = temp[["matched"]]
    df = df.merge(temp,how="left",left_index=True,right_index=True)
    print(df.head(50))
    grps = df.groupby(["VRN"])##.apply(lambda g: g[g['matched_y'] == "Y"])
    print(grps.get_group("00C15057"))


    #calculate_cordon_in_out_only(job)


    exit()

    ### below is the code I used to extract WEM images for graham
    ###
    ###
    file = "S:/SCOTLAND DRIVE 2/JOB FOLDERS/3 - Tadcaster/ANPR Jobs/xCompleted 2016/3092-Mid, Wem_images/"
    #print(job["sites"])

    #path = filedialog.askdirectory()
    df = pd.read_excel(file + "/Movement 9.xlsx")

    df.columns = ["movement","plate","time"]
    print("before",len(df))
    print("no of unique plates",len(df["plate"].unique()))
    df = df.drop_duplicates(subset="plate")
    print("after", len(df))
    movement = "Movement " + str(df.iloc[0]["movement"])
    print("movement is",movement)
    print(df.head())
    print(df.info())
    fileList = os.listdir(file + "Movement 9 & 10/")
    #print(fileList)
    for row in df.itertuples():
        #print(row)
        #print("looking for plate", row[2])
        plate = row[2]
        for fname in fileList:
            s = fname.split("_")
            if len(s) > 1:
                #print(s)
                #print(s[1])
                if plate == s[1] and ".jpeg" in fname and not "patch" in fname:
                    #print("found",plate,fname)
                    shutil.copy(file + "Movement 9 & 10/" + fname ,file + "/Selected/")
                    break
        else:
            print("didnt find",plate)


        #shutil.move(path + "/" + file , "C:/Users/NWatson/Desktop/ANPR data/" + file)
    exit()


#test()

win = mainwindow.mainWindow()
win.setCallbackFunction("load unclassed",load_unclassed_plates)
win.setCallbackFunction("load job",load_job)
win.setCallbackFunction("load overview count",load_completed_count)
win.setCallbackFunction("load classed",load_classes)
win.setCallbackFunction("get unclassed comparison",get_comparison_data)
#win.setCallbackFunction("reprocess data",reprocess_data)
win.setCallbackFunction("set duplicates",set_new_duplicates_value)
win.setCallbackFunction("get cordon in out only data",calculate_cordon_in_out_only)
win.setCallbackFunction("get cordon non directional data",calculate_nondirectional_cordon)
win.setCallbackFunction("get journey pairs",calculate_route_assignment_journey_pairs)
win.setCallbackFunction("get fs-ls data",calculate_route_assignment_fs_ls)
win.setCallbackFunction("get overtaking data",calculate_overtaking)
win.setCallbackFunction("resample overtaking data",resample_overtaking_data)
win.setCallbackFunction("update data after job save",update_sites_and_movements)
win.setCallbackFunction("recalculate platooning",calculate_platooning)
win.setCallbackFunction("filtered matching",calculate_regex_matching)

win.mainloop()
