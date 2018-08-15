import myDB
import os
import pickle
import datetime
import pandas as pd
from pandas.api.types import CategoricalDtype
import numpy as np
import anprregex
import pyexcelerate
import openpyxl
from tkinter import messagebox
import copy
import unicodedata
import string

class ANPRproject():

    def __init__(self):
        self.project = None
        self.durations = None
        self.folder = None
        self.data = None
        self.inMovproj  = None
        self.outMov = None
        self.bothMov = None
        self.times = None
        self.allJourneys = None


    def load_project(self,iid):
        self.projectId = iid
        self.jobDetails = myDB.get_project_details(iid)
        self.inMov,self.outMov,self.allMov = myDB.get_movements(iid)
        self.folder = myDB.get_folder(iid)

        if not os.path.exists(os.path.join(self.folder,"output")):
            os.mkdir(os.path.join(self.folder,"output"))
        if not os.path.exists(os.path.join(self.folder,"data")):
            os.mkdir(os.path.join(self.folder,"data"))
        self.load_durations()
        self.times = self.set_up_times()
        print("times are", self.times)
        #self.load_plates()


    def load_durations(self,val="12:00"):
        if not os.path.exists(os.path.join(self.folder,"data","durations.pkl")):
            self.durations = {}
            for m in self.allMov:
                for m2 in self.allMov:
                    self.durations[m,m2] = val
            self.save_durations()
        else:
            with open(os.path.join(self.folder,"data","durations.pkl"),"rb") as f:
                self.durations = pickle.load(f)
        print("no of durations loaded", len(self.durations))


    def clear_data_folder(self):
        if os.path.exists(os.path.join(self.folder, "data", "data.pkl")):
            os.remove(os.path.join(self.folder, "data", "data.pkl"))
        if os.path.exists(os.path.join(self.folder, "data", "all journeys as list.pkl")):
            os.remove(os.path.join(self.folder, "data", "all journeys as list.pkl"))
        if os.path.exists(os.path.join(self.folder, "data", "durations.pkl")):
            os.remove(os.path.join(self.folder, "data", "durations.pkl"))
        if os.path.exists(os.path.join(self.folder, "data", "last run.pkl")):
            os.remove(os.path.join(self.folder, "data", "last run.pkl"))


    def get_durations(self):
        return self.durations


    def save_durations(self):
        with open(os.path.join(self.folder,"data","durations.pkl"), "wb") as f:
            pickle.dump(self.durations, f)


    def get_direction(self,mov):
        if mov in self.inMov:
            return "In"
        if mov in self.outMov:
            return "Out"
        return "Both"


    def set_up_times(self):
        result = []
        times = myDB.get_times(self.projectId)
        for i in range(0,len(times),5):
            start = times[i]
            end = times[i+1]
            fromTime = times[i+2]
            toTime = times[i+3]
            splt = times[i+4]
            if not start is None and start != "" and not end is None and end != "":
                try:
                    fromTime = datetime.datetime.strptime(fromTime,"%H:%M").time()
                    toTime = datetime.datetime.strptime(toTime, "%H:%M").time()
                    if splt:
                        while start <= end:
                            if toTime<fromTime:
                                result.append([datetime.datetime.combine(start, fromTime), datetime.datetime.combine(start + datetime.timedelta(days=1), toTime)])
                            else:
                                result.append([datetime.datetime.combine(start, fromTime), datetime.datetime.combine(start, toTime)])
                            start+=datetime.timedelta(days=1)
                    else:
                        result.append([datetime.datetime.combine(start,fromTime),datetime.datetime.combine(end,toTime)])
                except Exception as e:
                    print(e)


        return result


    def load_plates(self,callbackFunction,changeProgressMessageFunction):
        self.data = None
        print(".looking for", os.path.join(self.folder,"data", "data.pkl"))
        if not os.path.exists(os.path.join(self.folder,"data", "data.pkl")):
            print("couldnt find data file")
            uploadedFile = myDB.get_uploaded_file(self.projectId)
            if not uploadedFile is None and os.path.exists(uploadedFile):
                print("converting plates")
                #self.project.beingProcessed = True
                #self.project.save()
                if not self.convert_excel_plates(changeProgressMessageFunction):
                    callbackFunction()
                    return False
                #self.project.beingProcessed = False
                #self.project.save()
                # self.create_full_journeys()
            else:
                print("no plates loaded")
                callbackFunction()
                return False

        try:
            changeProgressMessageFunction("Reading plates from data file")
            self.data = pd.read_pickle(os.path.join(self.folder,"data", "data.pkl"))
            self.data.reset_index(inplace=True)
            print(self.data["VRN"].head())
            #printable = set(string.printable)
            #self.data["VRN"] = self.data["VRN"].apply(lambda x:"".join([item for item in x if item in printable]))
            #print(self.data["VRN"].head())
            dups = len(self.data[self.data.duplicated(subset = ["VRN","Date","movement"],keep=False)])
            print("dups")
            #print(self.data[self.data.duplicated()])
            if dups > 0:
                answer = messagebox.askyesno(
                    message="There are " + str(dups) + " duplicate plates , do you want to remove them before continuing?")
                if  answer:
                    self.data = self.data.drop_duplicates(subset = ["VRN","Date","movement"])
                    dups = len(self.data[self.data.duplicated(subset=["VRN", "Date", "movement"], keep=False)])
                    print("there are now",dups,"duplicates")


            self.data.set_index("Date", inplace=True)
            print("no of plates loaded,", len(self.data))
        except Exception as e:
            print(e)
            callbackFunction()
            return False
        callbackFunction()
        return True


    def convert_excel_plates(self,changeProgressMessageFunction):
        changeProgressMessageFunction("Reading plates from excel file")
        uploadedFile = myDB.get_uploaded_file(self.projectId)
        if not os.path.exists(uploadedFile):
            return False
        ext = uploadedFile[uploadedFile.rfind("."):]
        if ext not in (".xlsx", ".csv", ".xlsm", ".xls"):
            return False
        try:
            if ".csv" in uploadedFile:
                df = pd.read_csv(uploadedFile)
            else:
                print("reading file with date converters", datetime.datetime.now())
                df = pd.read_excel(uploadedFile,
                                   converters={"VRN": str, "Direction": str, "Movement": int,"Class":str})
                df["Date"] = df.apply(lambda x: pd.datetime.combine(x["Date"], x["Time"]), 1)
            df = df[["Date", "Time", "Movement", "VRN", "Class"]]
            df.columns = ["Date", "Time", "oldMovement", "VRN", "Class"]
            changeProgressMessageFunction("Setting up movements")
            movements = myDB.get_project_movements(self.projectId)
            print("starting new method")
            for mov in movements:
                if not mov[3] is None and mov[3] != "":
                    print("converting", len(df[df["oldMovement"] == mov[3]]), "with oldmov ", mov[2], "to",
                          mov[4])
                    df.loc[df["oldMovement"] == mov[2], "direction"] = mov[4]
                    df.loc[df["oldMovement"] == mov[2], "movement"] = mov[3]
            print("finished new method")
            changeProgressMessageFunction("Converting Plates")
            df.drop(["Time"], inplace=True, axis=1)
            df["Duplicates"] = "N"
            df.sort_values(by=["VRN", "movement"], inplace=True, ascending=[True, True])
            df["timeDiff"] = df["Date"].diff()
            df.set_index(["Date"], inplace=True)
            mask = (df["VRN"] != df["VRN"].shift(-1)) | (df["movement"] != df["movement"].shift(-1))
            df.loc[mask, "timeDiff"] = np.nan
            print("finished", datetime.datetime.now())
            print(df[["oldMovement", "movement"]].head())
            self.data = df
            self.data["oldMovement"] = self.data["oldMovement"].astype("uint8")
            self.data["movement"].fillna(0, inplace=True)
            self.data["movement"] = self.data["movement"].astype(int)
            self.data["movement"] = self.data["movement"].astype("uint8")
            self.data["Class"] = self.data["Class"].astype("str")
            changeProgressMessageFunction("Removing unicode characters")
            self.data["Class"] = self.data["Class"].apply(lambda x:bytes(x, 'utf-8').decode('ascii', 'ignore'))
            self.data["VRN"] = self.data["VRN"].apply(lambda x: bytes(x, 'utf-8').decode('ascii', 'ignore') if type(x)==str else "")

            self.data["Class"] = self.data["Class"].astype("category")

            cat_type = CategoricalDtype(categories=["I", "O", "B"], ordered=True)
            self.data["direction"] = self.data["direction"].astype(cat_type)

            cat_type = CategoricalDtype(categories=["Y", "N"], ordered=True)
            self.data["Duplicates"] = self.data["Duplicates"].astype(cat_type)
            changeProgressMessageFunction("Creating Journeys")
            self.create_full_journeys()
            print("finished creating journeys",datetime.datetime.now())
            #self.project.uploadedDate = datetime.datetime.now()
            #self.project.save()
        except ValueError as e:
            ###
            ### TODO do something when the reading of plates failed
            print(e)
            self.data = None
            return False
        changeProgressMessageFunction("Saving Plates")
        self.save_plates()
        return True


    def save_plates(self):
        self.data.to_pickle(os.path.join(self.folder,"data", "data.pkl"))


    def create_full_journeys(self):
        if self.data is None:
            return
        result = []
        df = self.data#[self.project.projectDate.strftime("%Y-%m-%d")]
        print(df.info())
        df = df[df["Class"].notnull()]
        print("*"*100)
        print("creating full hjourneys",datetime.datetime.now())

        #print(df[df["VRN"] == "111MJT"])
        for t in self.times:#
            print(t[0],t[1])
            mask = (df.index >= t[0]) & (df.index <= t[1])
            temp = df[mask]
            if len(temp) > 0:
                temp.index.name = "Date"
                temp.reset_index(inplace=True)
                temp = temp[temp["movement"] > 0]
                temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
                strJoin = lambda x: ",".join(x.astype(str))
                dateJoin = lambda x: ",".join(x.apply(date_to_time))
                temp = temp.groupby(["VRN", "Class"]).agg({"movement": strJoin, "Date": dateJoin,"direction":strJoin})
                temp.reset_index(inplace=True)
                temp = temp[["VRN", "Class", "Date", "movement", "direction"]]
                #temp.to_csv("wibble.csv")
                values = temp.values.tolist()
                f = open("test list.txt","w")
                for v in values:
                    #print("v is",v)
                    r = [v[0], v[1], list(zip(*[item.split(",") for item in v[2:]]))]
                    #print("r is",r)
                    r[2] = [list(item) for item in r[2]]
                    for item in r[2]:
                        #print("converting item",item)
                        for i in range(0,len(item),3):
                            item[i] = datetime.datetime.strptime(item[i],"%d/%m/%Y %H:%M:%S")
                    f.write(",".join(v) + "\n")
                    result.append(r)
                f.close()
        with open(os.path.join(self.folder,"data","all journeys as list.pkl"), "wb") as f:
            for r in result:
                pickle.dump(r, f)
            print("finished creating full hjourneys", datetime.datetime.now())
            print("*" * 100)


    def calculate_regex_matching(self, filters, durationCheck, durationBehaviour,durationMaxValue,filterBehaviour,callback):
        if False:
            print("method 1",datetime.datetime.now())
            strJoin = lambda x: ",".join(x.astype(str))
            dateJoin = lambda x: ",".join(x.apply(date_to_time))
            result = []
            for t in self.times:#
                print(t[0],t[1])
                mask = (self.data.index >= t[0]) & (self.data.index <= t[1])
                temp = self.data[mask]
                if len(temp) > 0:
                    temp.reset_index(inplace=True)
                    temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
                    grps = temp.groupby(["VRN"])
                    for name, grp in grps:
                        journey = grp[["VRN","Class"]].values.tolist()[0]
                        data = []
                        for index, row in grp.iterrows():
                            data.append([row["Date"],row["movement"],row["direction"]])
                        journey.append(data)
                        #print("journey  is",journey)
                        journeyList = [journey]
                        remainders = []
                        for f in filters:
                            remainders = []
                            for journey in journeyList:
                                #print("looking at ",journey)
                                data = list(journey[2])
                                matches, rem = anprregex.match2(data, f)
                                rem = [[journey[0], journey[1], r] for r in rem]
                                remainders += rem
                                for m in matches:
                                    output = []
                                    output.append(journey[0])
                                    output.append(journey[1])
                                    temp = ([(item[1], item[0]) for item in m])
                                    temp = [item for sublist in temp for item in sublist]
                                    [output.append(item) for item in temp]
                                    if not output in result:
                                        result.append(output)
                            journeyList = remainders


        print("method2", datetime.datetime.now())
        if os.path.exists(os.path.join(self.folder,"data", "last run.pkl")):
            os.remove(os.path.join(self.folder,"data", "last run.pkl"))
        if not os.path.join(self.folder,"data","all journeys as list.pkl"):
            return self.get_matched_counts([], [])
        allPlates = []# [[journey[0], journey[1],journey[2][i][1],journey[2][i][0]] for journey in self.allJourneys for i in
                         #range(len(journey[2]))]

       # print("number of plates",len(allPlates))
        result = []
        #print("filters",filters)
        if filters == []:
            return self.get_matched_counts([],[])
        for journey in read_from_pickle(os.path.join(self.folder,"data","all journeys as list.pkl")):
            allPlates+=[[journey[0], journey[1],journey[2][i][1],journey[2][i][0]]  for i in
                         range(len(journey[2]))]
            journeyList = [journey]
            for f in filters:
                #print("journey list is",journeyList)
                remainders = []
                for journey in journeyList:
                    data = list(journey[2])
                    matches,rem = anprregex.match2(data, f)
                    rem = [[journey[0], journey[1], r] for r in rem]
                    remainders+=rem
                    for m in matches:
                        output = []
                        output.append(journey[0])
                        output.append(journey[1])
                        temp = ([(item[1], item[0]) for item in m])
                        temp = [item for sublist in temp for item in sublist]
                        [output.append(item) for item in temp]
                        if not output in result:
                            result.append(output)
                if not filterBehaviour:
                    ###
                    ### we can either apply each filter to the full journey,
                    ### or we can apply a filter, and the apply the next filter to the unmatched parts of the journey
                    ### if filterBehaviour is false, we want to apply filters to the unmatched parts of the journey.
                    journeyList = remainders


        print("finsihed", datetime.datetime.now())
        allPlates = pd.DataFrame(allPlates)
        print("allplates is",allPlates)
        if len(allPlates) > 0:
            allPlates.columns = ["VRN", "Class", "movement", "Date"]
            allPlates["movement"] = allPlates["movement"].astype(int, errors="ignore")
        result = self.check_durations(result, durationCheck, durationBehaviour,durationMaxValue)
        #print("after durations check, first journey is", result[0])
        with open(os.path.join(self.folder,"data","last run.pkl"), "wb") as f:
            pickle.dump({"result":result,"allPlates":allPlates,"name":"Filtered Matching"}, f)
        #self.save_matched_data(result,allPlates,"Filtered Matching",timeType)
        #print("after saving check, first journey is", result[0])
        callback( self.get_matched_counts(result,allPlates))


    def calculate_nondirectional_cordon(self, durationCheck, durationBehaviour,durationMaxValue,callback):
        ###
        ### we want to "pair off" appearances of a vehicle. So if there are 4 appearances of a vehicle, we pair them off as
        ### (1,2) and (3,4). Unlike directional, we dont care about whether the first is an in and the second is an out
        ###
        if os.path.exists(os.path.join(self.folder,"data", "last run.pkl")):
            os.remove(os.path.join(self.folder,"data", "last run.pkl"))
        if self.data is None:
            return self.get_matched_counts([], [])
        journeys = []
        allPlates = []
        df = self.data#[self.project.projectDate.strftime("%Y-%m-%d")]
        print("before removing null classes",len(df))
        df = df[df["Class"].notnull()]
        print("after removing null classes", len(df))
        for t in self.times:
            mask = (df.index >= t[0]) & (df.index <= t[1])
            temp = df[mask]
            print("no of selected vrns",len(temp))
            temp.index.name = "Date"
            temp.reset_index(inplace=True)
            temp = temp[temp["movement"] > 0]
            temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
            allPlates.append(temp.copy())
            ###
            ### new method
            ###
            if False:
                temp = temp[["VRN", "Class", "movement", "Date"]]
                temp["matched"] = None
                temp = temp[temp.duplicated(subset=["VRN"], keep=False)]
                grp = temp.groupby(["VRN"])
                temp["matched"] = grp.cumcount() // 2 + 1
                grp = temp.groupby(["VRN", "matched"])
                temp = grp.filter(lambda x: len(x) == 2)
                temp.loc[grp.cumcount() == 0, "pos"] = "S"
                temp.loc[grp.cumcount(ascending=False) == 0, "pos"] = "E"
                temp = temp[temp["pos"].isin(["S", "E"])]
                temp["outTime"] = temp["Date"].shift(-1)
                temp["outMovement"] = temp["movement"].shift(-1)
                temp = temp[temp["pos"] == "S"]
            if True:
                ### old method
                temp["matched"] = "N"
                grp = temp.groupby(["VRN"])
                temp.loc[grp.cumcount() % 2 == 0,"matched"] = "Y"#temp["Date"][grp.cumcount() % 2 == 0]
                temp.loc[grp.cumcount(ascending=False) == 0,"matched"] = "N"
                #print("temp")
                #print(temp[["VRN","Date","matched"]].tail(10))
                temp["outTime"] = temp["Date"].shift(-1)
                #temp["Date"] = temp["Date"].apply(date_to_time)
                #temp["outTime"] = temp["outTime"].apply(date_to_time)
                temp["outMovement"] = temp["movement"].shift(-1)
                temp["movement"] = temp["movement"].real.astype(int)
                temp["outMovement"] = temp["outMovement"].real.astype(int)

                temp.to_csv("dumped.csv")
            ###
            ### get the matches
            ###
            temp = temp[temp["matched"] == "Y"]
            temp["movement"].dropna(inplace=True)
            print("at end of process, temp is",len(temp))
            journeys.append(temp)
        journeys = pd.concat(journeys)
        print("after concatenation, journets is",len(journeys))
        #print(journeys.head())
        #journeys["Date"] = journeys["Date"].apply(date_to_time)
        #journeys["outTime"] = journeys["outTime"].apply(date_to_time)
        #print("journeys")
        #print(journeys.head())
        journeys = journeys[["VRN", "Class", "movement", "Date", "outMovement", "outTime"]].values.tolist()
        print("len of journeys is",len(journeys))
        allPlates = pd.concat(allPlates)
        print("size of allplates is",len(allPlates))
        #allPlates["Date"] = allPlates["Date"].apply(date_to_time)
        journeys = self.check_durations(journeys, durationCheck, durationBehaviour, durationMaxValue)
        with open(os.path.join(self.folder,"data","last run.pkl"), "wb") as f:
            pickle.dump({"result":journeys,"allPlates":allPlates,"name":"Non Directional"}, f)
        print("journeys", journeys[:5])
        callback(self.get_matched_counts(journeys,allPlates))


    def calculate_directional_cordon(self, durationCheck, durationBehaviour,durationMaxValue,callback):
        ###
        ### we want to "pair off" appearances of a vehicle. So if there are 4 appearances of a vehicle, we pair them off as
        ### (1,2) and (3,4). Unlike directional, we dont care about whether the first is an in and the second is an out
        ###
        if os.path.exists(os.path.join(self.folder,"data", "last run.pkl")):
            os.remove(os.path.join(self.folder,"data", "last run.pkl"))
        if self.data is None:
            return self.get_matched_counts([], [])
        journeys = []
        allPlates = []
        df = self.data#[self.project.projectDate.strftime("%Y-%m-%d")]
        print("before removing null classes",len(df))
        df = df[df["Class"].notnull()]
        print("after removing null classes", len(df))
        for t in self.times:
            mask = (df.index >= t[0]) & (df.index <= t[1])
            temp = df[mask]
            print("no of selected vrns",len(temp))
            temp.index.name = "Date"
            temp.reset_index(inplace=True)
            temp = temp[temp["movement"] > 0]
            temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
            allPlates.append(temp.copy())

            ### old method
            mask = (temp["direction"] == "I") & (temp["direction"].shift(-1) == "O")
            mask = (mask) & (temp["VRN"] == temp["VRN"].shift(-1))
            temp["matched"] = "N"
            temp.ix[mask, "matched"] = "Y"
            temp["outTime"] = temp["Date"].shift(-1)
            temp["outMovement"] = temp["movement"].shift(-1)
            temp["movement"] = temp["movement"].real.astype(int)
            temp["outMovement"] = temp["outMovement"].real.astype(int)
            ###
            ### get the matches
            ###
            temp = temp[temp["matched"] == "Y"]
            temp["movement"].dropna(inplace=True)
            print("at end of process, temp is",len(temp))
            journeys.append(temp)
        journeys = pd.concat(journeys)
        print("after concatenation, journets is",len(journeys))
        print(journeys.head())
        #journeys["Date"] = journeys["Date"].apply(date_to_time)
        #journeys["outTime"] = journeys["outTime"].apply(date_to_time)
        print("journeys")
        print(journeys.head())
        journeys = journeys[["VRN", "Class", "movement", "Date", "outMovement", "outTime"]].values.tolist()
        allPlates = pd.concat(allPlates)
        print("size of allplates is",len(allPlates))
        #allPlates["Date"] = allPlates["Date"].apply(date_to_time)
        journeys = self.check_durations(journeys, durationCheck, durationBehaviour, durationMaxValue)
        with open(os.path.join(self.folder,"data","last run.pkl"), "wb") as f:
            pickle.dump({"result":journeys,"allPlates":allPlates,"name":"In Out Matching"}, f)
        print("journeys", journeys[:5])
        callback( self.get_matched_counts(journeys,allPlates))



    def calculate_pairs(self,durationCheck, durationBehaviour,durationMaxValue,callback):
        if os.path.exists(os.path.join(self.folder,"data", "last run.pkl")):
            os.remove(os.path.join(self.folder,"data", "last run.pkl"))
        if self.data is None:
            return self.get_matched_counts([], [])
        journeys = []
        allPlates = []
        df = self.data#[self.project.projectDate.strftime("%Y-%m-%d")]
        print("before removing classes, df is",len(df))
        print("after removing classes, df is", len(df))
        df = df[df["Class"].notnull()]
        for t in self.times:
            mask = (df.index >= t[0]) & (df.index <= t[1])
            temp = df[mask]
            temp.index.name = "Date"
            temp.reset_index(inplace=True)
            temp = temp[temp["movement"] > 0]
            print("no of entries in df",len(temp))
            allPlates.append(temp.copy())
            temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
            mask = (temp["VRN"] == temp["VRN"].shift(-1))
            temp["matched"] = "N"
            temp.ix[mask, "matched"] = "Y"
            #temp["Date"] = temp["Date"].apply(date_to_time)
            temp["outTime"] = temp["Date"].shift(-1)
            temp["outMovement"] = temp["movement"].shift(-1)
            temp["movement"] = temp["movement"].real.astype(int)
            temp["outMovement"] = temp["outMovement"].real.astype(int)

            selected = temp[temp["matched"] == "Y"]
            selected["movement"].dropna(inplace=True)
            journeys += selected[["VRN", "Class", "movement", "Date", "outMovement", "outTime"]].values.tolist()
        allPlates = pd.concat(allPlates)
        print("all opklates")
        print(allPlates.info())
        journeys = self.check_durations(journeys, durationCheck, durationBehaviour, durationMaxValue)
        with open(os.path.join(self.folder,"data","last run.pkl"), "wb") as f:
            pickle.dump({"result":journeys,"allPlates":allPlates,"name":"Pairs"}, f)
        callback(self.get_matched_counts(journeys,allPlates))


    def calculate_fs_ls(self,durationCheck,durationBehaviour,durationMaxValue,callback):
        if os.path.exists(os.path.join(self.folder,"data", "last run.pkl")):
            os.remove(os.path.join(self.folder,"data", "last run.pkl"))
        if self.data is None:
            return self.get_matched_counts([], [])
        journeys = []
        allPlates = []
        df = self.data#[self.project.projectDate.strftime("%Y-%m-%d")]
        df = df[df["Class"].notnull()]
        for t in self.times:
            mask = (df.index >= t[0]) & (df.index <= t[1])
            temp = df[mask]
            temp.index.name = "Date"
            temp.reset_index(inplace=True)
            temp = temp[temp["movement"] > 0]
            #temp["Date"] = temp["Date"].apply(date_to_time)
            print("no of entries in df", len(temp))
            temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])

            ###
            ### append a copy of the selected dataframe to allPlates for later use, before we remove singletons
            ###

            allPlates.append(temp.copy())

            ###
            ### remove singletons
            ###
            temp = temp[temp.duplicated(subset=["VRN"], keep=False)]

            ###
            ### find first occurence of a plate. We dont care about direction
            ###
            grp = temp.groupby(["VRN"])
            result1 = temp[grp.cumcount() == 0]
            print(result1.head())
            ###
            ### find last occurence of a plate
            ###
            result2 = temp[grp.cumcount(ascending=False) == 0]
            print(result2.head())
            fullResult = pd.concat([result1, result2])
            print(fullResult)
            fullResult.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
            print(fullResult.head())
            fullResult["outTime"] = fullResult["Date"].shift(-1)
            fullResult["movement"] = fullResult["movement"].real.astype(int)
            fullResult["outMovement"] = fullResult["movement"].shift(-1)
            fullResult["outMovement"] = fullResult["outMovement"].real.astype(int)
            fullResult = fullResult.iloc[::2]
            journeys += fullResult[["VRN", "Class", "movement", "Date", "outMovement", "outTime"]].values.tolist()

            print(fullResult.head())
        allPlates = pd.concat(allPlates)
        journeys = self.check_durations(journeys, durationCheck, durationBehaviour, durationMaxValue)
        with open(os.path.join(self.folder,"data","last run.pkl"), "wb") as f:
            pickle.dump({"result":journeys,"allPlates":allPlates,"name":"First Seen Last Seen"}, f)
        callback(self.get_matched_counts(journeys,allPlates))


    def calculate_full_journeys(self,durationCheck,durationBehaviour,durationMaxValue,callback):
        if not os.path.exists(os.path.join(self.folder,"data","all journeys as list.pkl")):
            return self.get_matched_counts([], [])
        journeys = []
        allPlates = []
        #print("in calculate, values are",durationCheck,durationBehaviour,durationMaxValue)
        df = self.data#[self.project.projectDate.strftime("%Y-%m-%d")]
        df = df[df["Class"].notnull()]
        for t in self.times:
            mask = (df.index >= t[0]) & (df.index <= t[1])
            print("no of entries in df", len(df))
            temp = df[mask]
            temp.index.name = "Date"
            temp.reset_index(inplace=True)
            print("no of entries in df", len(temp))
            temp = temp[temp["movement"] > 0]
            #temp["Date"] = temp["Date"].apply(date_to_time)
            print("no of entries in df", len(temp))
            print(temp.head())
            temp.sort_values(by=["VRN", "Date"], inplace=True, ascending=[True, True])
            print("temp is",temp)
            ###
            ### append a copy of the selected dataframe to allPlates for later use, before we remove singletons
            ###

            allPlates.append(temp.copy())

            ###
            ### remove singletons
            ###
            print("before removing singletons, size is", len(temp))
            temp = temp[temp.duplicated(subset=["VRN"], keep=False)]
            print("after removing singletons, size is",len(temp))
            if False:
                ###
                ### new method
                ###

                temp = temp[["VRN", "Class", "movement", "Date"]]
                temp["matched"] = None
                temp = temp[temp.duplicated(subset=["VRN"], keep=False)]
                grp = temp.groupby(["VRN"])
                temp["matched"] = grp.cumcount() // 2 + 1
                grp = temp.groupby(["VRN", "matched"])
                temp = grp.filter(lambda x: len(x) >= 2)
                temp.loc[grp.cumcount() == 0, "pos"] = "S"
                temp.loc[grp.cumcount(ascending=False) == 0, "pos"] = "E"
                temp.loc[(grp.cumcount() != 0) & (grp.cumcount(ascending=False) != 0), "pos"] = "M"
                print("temp is")
                print(temp.head())
                temp = temp[temp["pos"].isin(["S", "E"])]
                temp["outTime"] = temp["Date"].shift(-1)
                temp["outMovement"] = temp["movement"].shift(-1)
                temp = temp[temp["pos"] == "S"]
            if True:
                strJoin = lambda x: ",".join(x.astype(str))
                dateJoin = lambda x: ",".join(x.apply(date_to_time))
                if len(temp) > 0:
                    temp = temp.groupby(["VRN", "Class"]).agg({"movement": strJoin, "Date": strJoin})
                    temp.reset_index(inplace=True)
                    print(temp.head())
                    values = temp[["VRN", "Class", "movement", "Date"]].values.tolist()
                    for v in values:
                        #print("v is",v)
                        for i in range(2, 4):
                            v[i] = [item for item in v[i].split(",")]
                        l = [item for sublist in list(zip(*[v[2], v[3]])) for item in sublist]
                        l.insert(0, v[1])
                        l.insert(0, v[0])

                        for i in range(3, len(l), 2):
                            l[i] = datetime.datetime.strptime(l[i], "%Y-%m-%d %H:%M:%S")
                        #print("l is",l)
                        journeys.append(l)
                        #print("appending",l)
        print("journeys", journeys[:5])
        allPlates = pd.concat(allPlates)
        journeys = self.check_durations(journeys, durationCheck, durationBehaviour,durationMaxValue)
        with open(os.path.join(self.folder,"data","last run.pkl"), "wb") as f:
            pickle.dump({"result":journeys,"allPlates":allPlates,"name":"Full Journeys"}, f)
        callback(self.get_matched_counts(journeys,allPlates))


    def get_matched_counts(self,journeys,allPlates):
        resultsDict = {} ### holds the counts for the first seen and last seen plate, to display in the matrix
        totalCounts = {int(k):0 for k in self.allMov} ### holds the counts for every plate seen at a movement
        inTotals = {int(k): 0 for k in self.allMov}
        outTotals = {int(k): 0 for k in self.allMov}
        print("totalcounts are",totalCounts)
        for journey in journeys:
            #print("journey",journey)
            resultsDict[(int(journey[2]), int(journey[-2]))] = resultsDict.get((int(journey[2]), int(journey[-2])), [])
            resultsDict[(int(journey[2]), int(journey[-2]))].append(journey[-1] - journey[3])
            inTotals[int(journey[2])] = inTotals[int(journey[2])] + 1
            outTotals[int(journey[-2])] = outTotals[int(journey[-2])] + 1

            for movement in journey[2::2]:
                #print("adding 1 to ",movement,type(movement))
                count  = totalCounts.get(int(movement),0)
                count+=1
                totalCounts[int(movement)] = count

        ###
        ### set up some indexes so that if any sites have 0 values, we still pick up the sites in the dataframe
        ###
        allDf = pd.DataFrame(index=self.allMov)
        allDf["count"] = 0
        for key, item in resultsDict.items():
            resultsDict[key] = [len(item), format_timedelta(max(item)), format_timedelta(min(item)),
                                format_timedelta(np.mean(item))]
            allDf.loc[int(key[0])]["count"] += len(item)
            allDf.loc[int(key[1])]["count"] += len(item)
        allDf.sort_index(inplace=True)
        counts = self.get_plate_count_by_movement(allPlates)
        total = 0
        for key,item in inTotals.items():
            resultsDict[(key,"Total")] = item
            total +=item
        for key,item in outTotals.items():
            resultsDict[("Total",key)] = item
            total += item
        resultsDict[("Total","Total")] = int(total/2)

        return (resultsDict, allDf["count"].to_dict(), counts,totalCounts,inTotals,outTotals)


    def check_durations(self,journeys,durationCheck, durationBehaviour,durationValue = None):
        print("starting durations check",durationCheck, durationBehaviour,durationValue)
        count = 0
        if durationCheck is None or durationBehaviour is None:
            return journeys
        if not self.durations is None:
            if durationCheck == "max":
                try:
                    val = datetime.datetime.strptime(durationValue, "%H:%M")
                except ValueError as e:
                    try:
                        val = int(durationValue)
                    except ValueError as e:
                        ###
                        ### entered value is not a time or an int
                        ### just return the unaltered journeys
                        ###
                        return journeys
                for journey in journeys:
                    start = 2
                    if type(val) == int:
                        ###
                        ### split by max journey length
                        ###
                        journeyLength = (len(journey) - 2)/2
                        if val < journeyLength:
                            if durationBehaviour == "split":
                                ###
                                ### split the journey, make a new journey out of the end of the current journey
                                ###
                                newJourney = [journey[0], journey[1]]
                                newJourney += journey[(val*2)+2:]
                                del journey[(val*2)+2:]
                                journeys.append(newJourney)
                            else:
                                ###
                                ### discard full journey
                                ###
                                del journey[:]
                        else:
                            if journeyLength < 2:
                                #print("found journey of less than 2 length")
                                del journey[:]

                    else:
                        ###
                        ### split by time
                        ###
                        hour,min = val.hour,val.minute
                        td = datetime.timedelta(hours=hour,minutes=min)
                        index = 5
                        if len(journey) < 5:
                            del journey[:]
                        while index < len(journey):
                            diff = journey[index]- journey[3]
                            if journey[0] == "SEP0002-000803":
                                print("looking at ",journey,journey[index])
                                print("td is",td)
                                print("diff is",diff)
                                print("diff > td?",diff>td)
                            if diff > td:
                                if durationBehaviour == "split":
                                    ###
                                    ### is the split at the first movement? If so, we dont need to keep that segment, its too small
                                    ###
                                    newJourney = [journey[0], journey[1]]
                                    newJourney += journey[index-1:]
                                    journeys.append(newJourney)
                                    if index > 5:
                                        del journey[index-1:]
                                    else:
                                        del journey[:]
                                else:
                                    del journey[:]
                            index+=2






            else:
                for journey in journeys:
                    #print("checking journey", journey)
                    start = 2
                    while start < len(journey) - 2:
                        # print("start is",start,len(journey) -1)
                        duration = journey[start + 3] - journey[start + 1]
                        v = self.durations[(int(journey[start]), int(journey[start + 2]))]
                        splitTime = v.split(":")
                        hours = int(splitTime[0])
                        mins = int(splitTime[1])
                        td = datetime.timedelta(hours=hours, minutes=mins, seconds=0)
                        # print("duration is", duration, (int(journey[start]), int(journey[start + 2])))
                        if duration > td:
                            if durationBehaviour == "split" and len(
                                    journey) > 6:  ## split any journeys where a leg exceeds the duration
                                newJourney = [journey[0], journey[1]]
                                [newJourney.append(item) for item in journey[start + 2:]]
                                while len(journey) > start + 2:
                                    del journey[-1]
                                if len(newJourney) > 4:
                                    journeys.append(newJourney)
                                if len(journey) < 5:
                                    while len(journey) > 0:
                                        del journey[-1]
                                        # print("journey is now",journey,"added journey",newJourney)
                            else:  ### discard any journeys where a leg exceeds the duration
                                count+=1
                                #print("deleting journey",journey)
                                #print("appending",[[journey[0],journey[1]] + journey[i:i+2] for i in range(2, len(journey), 2)])
                                while len(journey) > 0:
                                    del journey[-1]
                        start += 2
        print("no of deleted journeys",count)
        result = [item for item in journeys if item != []]
        return result


    def get_plate_count_by_movement(self,df):
        ###
        ### count all vehicles seen at each movement for the specified project time period(s)
        ###
        counts = {m:0 for m in self.allMov}
        #print("in get plate count, counts are",counts)
        if self.data is None:
            return counts
        if len(df) >0:
            grps=df.groupby("movement").size()
            for k,v in counts.items():
                #print("looking for key",k,type(k))
                try:
                    print(counts[k],grps[k])
                    counts[k] = grps[k]
                except KeyError as e:
                    print("error",e)
                    counts[k] = 0
        print("counts are",counts)
        return counts


    def save_matched_data(self,timeType,daysInSeparateSheets):
        ###
        ### This function changes the format of each journey, from the full journey, to the first seen and last seen
        ### movements, and then the full journey
        ###
        timeType = 2 # because we only want to output time at the moment, this might change if we go back to different time outputs
        #print("timetype is",timeType,type(timeType))
        if not os.path.exists(os.path.join(self.folder,"data", "last run.pkl")):
            return
        with open(os.path.join(self.folder,"data", "last run.pkl"), "rb") as f:
            data = pickle.load(f)
        journeys = data["result"]
        allPlates = data["allPlates"]
        fileTitle = data["name"]
        ###
        ### Get the unmatched plates
        ###
        unmatchedData = []
        matchedPlates = []

        print("*"*100)
        print("journeys")
        #print(journeys[:10])
        #print("allplatyes")
        if len(allPlates) == 0:
            return
        allPlates = allPlates[["VRN", "Class", "movement", "Date"]]
        #print(allPlates.info())
        uniqueMatches = [[journey[0], journey[1]] + journey[i:i + 2] for journey in journeys for i in
                         range(2, len(journey), 2)]
        print("after first loading, unique is",len(uniqueMatches))
        #uniqueMatches = [list(x) for x in set(tuple(x) for x in uniqueMatches)]
        #print("fater processing, unique is", len(uniqueMatches))
        journeysDf = pd.DataFrame(uniqueMatches)
        journeysDf.columns = ["VRN", "Class", "movement", "Date"]
        #print(journeysDf.head())
        #print(journeysDf.info())
        print("num duplicates",len(journeysDf[journeysDf.duplicated(subset = ["VRN","Date","movement"],keep=False)]))
        journeysDf[journeysDf.duplicated(subset = ["VRN","Date","movement"],keep=False)].to_csv("duplicates.csv")
        #print(journeysDf[journeysDf.duplicated()])
        print("non duplicates",len(journeysDf[~journeysDf.duplicated(subset = ["VRN","Date","movement"])]))
        journeysDf = journeysDf.drop_duplicates(subset = ["VRN","Date","movement"])
        print("after dropping duplicates, ",len(journeysDf))
        journeysDf.sort_values(by=["VRN","Date"],inplace=True)
        uniqueDates = journeysDf["Date"].map(pd.Timestamp.date).unique()

        print("unique dates are",uniqueDates)
        del uniqueMatches
        #print("journeysDF is",len(journeysDf))
        #print(journeysDf.head())
        if len(journeysDf) > 0:
            journeysDf["movement"] = journeysDf["movement"].astype(int)
            #print("journeysDf2)")
            #print(journeysDf.head())
            #print(journeysDf.info())
            df_all = allPlates.merge(journeysDf, on=["VRN", "Date", "movement"],
                                     how='left', indicator=True)
            print("df all")
            print(df_all.info())
            #print(df_all["_merge"].values.unique())
            print("len of df all",len(df_all))
            print("both",len(df_all[df_all["_merge"] == "both"]),"left",len(df_all[df_all["_merge"] == "left_only"]))


            df_all.to_csv("dumped.csv")
            unmatchedData = df_all[df_all["_merge"] == "left_only"][["VRN","Class_x","movement", "Date"]].sort_values(by=["VRN","Date"])
            unmatchedData.columns  = ["VRN","Class","movement", "Date"]
            del df_all
            #print(unmatchedData.info())
            print("all plates", len(allPlates), "matched plates", len(journeysDf),"unmacthed plates",len(unmatchedData))
            if daysInSeparateSheets:
                selectedJourneys = []
                for uniqueDate in uniqueDates:
                    selectedJourneys = [copy.deepcopy(j) for j in journeys if j[3].date() == uniqueDate]

                    matchedPlates = journeysDf[journeysDf["Date"].dt.date == uniqueDate].sort_values(by=["VRN","Date"])
                    unmatchedPlates = unmatchedData[unmatchedData["Date"].dt.date == uniqueDate].sort_values(by=["VRN","Date"])
                    name = self.jobDetails[0] + " " + self.jobDetails[1] + " " + fileTitle + " for " + uniqueDate.strftime("%d %b %Y")  +\
                           " processed on " + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx"
                    self.output_to_excel(selectedJourneys,unmatchedPlates,matchedPlates,name)
            else:
                name = self.jobDetails[0] + " " + self.jobDetails[1] + " " + fileTitle  + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx"
                self.output_to_excel(journeys, unmatchedData, journeysDf, name)


    def output_to_excel(self,journeys,unmatchedPlates,matchedPlates,fileName):
        for index,row in enumerate(journeys):
            for i in range(3,len(row),2):
                row[i] = row[i].time()
            row[0] = str(row[0])
            journeys[index] = row[:4] + row[-2:] + [""] + row[2:]
            #print("row is now",row)
        matchedPlates["Date"] = matchedPlates["Date"].dt.time
        matchedPlates = matchedPlates.values.tolist()
        unmatchedPlates["Date"] = unmatchedPlates["Date"].dt.time
        unmatchedPlates = unmatchedPlates.values.tolist()
        try:
            #print("starting")
            wb = pyexcelerate.Workbook()
            wb.new_sheet("Matches", data=journeys)
            wb.new_sheet("Unmatched Plates", data=unmatchedPlates)
            wb.new_sheet("Matched Plates as list", data=matchedPlates)
            wb.save(os.path.join(self.folder, "output", fileName.replace(",", "")))
        except PermissionError as e:
            pass


    def export_OVTemplate(self):
        wb = openpyxl.load_workbook("OV Template.xlsm", keep_vba=True)
        try:
            sheet = wb.get_sheet_by_name("All Sites - 12Hr")
        except Exception as e:
            print("phoo")
            return
        # classes = self.currentJob["classification"].split(",")
        # classes = [x for i, x in enumerate(classes) if i % 2 == 0]
        # movements = project.movement_set.all()
        # print("movements is",type(movements))
        classes = myDB.get_classes(self.projectId)
        col = 13
        for mov in self.allMov:
            sheet.cell(row=4, column=col - 1).value = mov
            sheet.cell(row=6, column=col).value = "Overview"
            sheet.cell(row=6, column=col + len(classes) + 1).value = "ANPR"
            sheet.cell(row=6, column=col + (2 * (len(classes) + 1))).value = "Comparison"
            for i, c in enumerate(classes):
                print(c)
                sheet.cell(row=7, column=col + i).value = c
                sheet.cell(row=7, column=col + len(classes) + 1 + i).value = c
                sheet.cell(row=7, column=col + (2 * (len(classes) + 1)) + i).value = c
                sheet.cell(row=7, column=col + len(classes)).value = "O Total"
                sheet.cell(row=7, column=col + (2 * len(classes)) + 1).value = "A Total"
                sheet.cell(row=7, column=col + (3 * len(classes)) + 2).value = "C Total"

            row = 9

            for s, e in self.times:
                print(s)
                while s < e:
                    sheet.cell(row=row, column=col - 1).value = s.strftime("%H:%M")
                    s += datetime.timedelta(minutes=int(self.jobDetails[4]))
                    row += 1
                    if s.minute == 0 and sheet.cell(row=row-1,column=col-1).value != "1Hr":
                        sheet.cell(row=row, column=col - 1).value = "1 Hr"
                        row += 1
            col += (3 * (len(classes) + 1)) + 3
        wb.save(os.path.join(self.folder,"output",self.jobDetails[0] + " " + self.jobDetails[1] + " OV Output.xlsm"))


def format_timedelta(td):
    if pd.isnull(td):
        return 0
    minutes, seconds = divmod(td.seconds + td.days * 86400, 60)
    hours, minutes = divmod(minutes, 60)
    return '{:d}:{:02d}:{:02d}'.format(hours, minutes, seconds)


def read_from_pickle(path):
    with open(path, 'rb') as file:
        try:
            while True:
                yield pickle.load(file)
        except EOFError:
            pass


def date_to_time(d):
    if d is None:
        return "00:00:00"
    if pd.isnull(d):
        return "00:00:00"
    try:
        return d.strftime("%d/%m/%Y %H:%M:%S")
    except Exception as e:
        try:
            return d.strftime("%d/%m/%Y %H:%M")
        except Exception as e:
            return "00:00:00"

#myDB.set_file("C:/Users/NWatson/PycharmProjects/ANPR/blah.sqlite")
#proj = ANPRproject()
#proj.load_project("3206-MID", "Dudley 3", "2016-11-26")
#job = myDB.load_job("3206-MID", "Dudley 3", "2016-11-26")
#print(job)