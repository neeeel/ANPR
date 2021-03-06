
import tkinter
import tkinter.font as font

import tkinter.ttk as ttk
from tkinter import messagebox
import datetime
import openpyxl
import win32com.client
from PIL import Image,ImageDraw,ImageTk,JpegImagePlugin,Jpeg2KImagePlugin
from tkinter import filedialog
import threading
import os
import myDB
import pickle
import subprocess
import time
import re
import csv
import matrix
import copy
import pandas as pd


class mainWindow(tkinter.Tk):

    def __init__(self):
        super(mainWindow, self).__init__()

        self.colourLabels = []
        self.entryValues = []
        self.revertButton = None
        self.tracsisBlue = "#%02x%02x%02x" % (20, 27, 77)
        self.tracsisGrey = "#%02x%02x%02x" % (99, 102, 106)
        ttk.Style().configure(".", bg="white",fg="red")
        self.processOvertakingThread = None
        self.siteLabel = None
        self.box1Value = 0
        self.box2Value = 0 ### to keep track of the combo boxes on the comparison display sheet
        self.user = ""
        self.oldJobData = None
        self.recalcuatePlatooningfunction = None
        self.filteredMatchingfunction = None
        self.platooningTime = 5
        self.getJourneyPairsFunction = None
        self.getOvertakingDataFunction = None
        self.getRouteAssignmentFsLsFunction = None
        self.loadUnclassedFunction = None
        self.loadClassedFunction = None
        self.loadJobFunction = None
        self.getNonDirectionalCordonFunction = None
        self.reprocessDataFunction = None
        self.setDuplicatesFunction = None
        self.getCordonFunction = None
        self.getRouteAssignmentFsLsFunction = None
        self.updateDataFunction = None
        self.displayWin = None
        self.currentSelected = [0,0]
        self.loadOVCountsFunction = None
        self.getUnclassedComparisonFunction = None
        self.resampleOvertakingDataFunction =None
        self.displayStatus = "edited" ## stores whether to display base comparison data, or edited comparison data
        self.geometry("600x500")
        self.movementTabs = None
        self.displayedDataIndex = 0
        self.comparisonWindow = None
        self.summaryTree = None
        self.state("zoomed")
        self.numCams =0 ## this is so we can delete entryvalues from the list if the no of cameras gets updated on the survey set up window
        self.movementsFrame = None ### this frame goes in the parameters window, to display the site/movement data
        self.dataFrame = tkinter.Frame(self,width = 50,height  = 600,bg = "green")
        self.comparisonDataStructure = []
        self.configure(bg = "white")
        self.jobListBox = None
        self.matrixData = None
        self.scrollFrame = None
        self.currentJob = None
        self.durationsDictionary = None
        self.selectedDuplicates = None
        self.tempEditedDataStore  = []
        self.overtakingPairsDict = {}
        self.comparisonDataStructure = []
        self.dataList=[]

        self.menubar = tkinter.Menu(self)
        menu = tkinter.Menu(self.menubar, tearoff=0)
        menu.add_command(label="Change Settings", command=self.spawn_settings_window)
        self.menubar.add_cascade(label="Settings", menu=menu)
        self.config(menu=self.menubar)
        self.Dbfile = ""
        self.user = ""
        self.user,self.Dbfile=self.load_settings()
        if self.user == "" or self.Dbfile == "":
            self.spawn_settings_window()
        else:
            self.spawn_survey_setup_screen()

    def get_cordon_data(self):
        data = self.getCordonFunction(self.currentJob)
        self.draw_cordon_matrix(self.matrixCanvas,data)

    def spawn_overtaking_setup_screen(self):
        for child in self.winfo_children():
            child.destroy()
        self.overtakingPairsDict = {}
        width = self.winfo_screenwidth() - 320
        height = self.winfo_screenheight() - 200
        f = tkinter.font.Font(family="helvetica", size=10)
        vcmd = (self.register(self.validate_time_cell_input), "%d", "%s", "%S")
        frame = tkinter.Frame(self,bg="white")
        frame.grid(row=0,column=0,padx=20,pady=15)
        tkinter.Label(frame, text=" Mvmt 1",bg="white").grid(row=0, column=0)
        tkinter.Label(frame, text="  ",bg="white").grid(row=0, column=1)
        tkinter.Label(frame, text=" Mvmt 2",bg="white").grid(row=0, column=2)
        e =tkinter.Entry(frame,width=4,bg="white",validate="key",validatecommand=vcmd)
        e.grid(row=1,column = 0)
        e.focus()
        tkinter.Label(frame,text = " - " ,bg="white").grid(row =1,column = 1)
        e =tkinter.Entry(frame,width=4,bg="white",validate="key",validatecommand=vcmd)
        e.bind("<Return>",self.add_overtaking_pair)
        e.bind("<Tab>", self.add_overtaking_pair)
        e.grid(row=1, column=2)
        lbox =tkinter.Listbox(frame,bg= "white",font = f)
        lbox.grid(row = 2,column=0,columnspan=4)
        lbox.configure(exportselection=False)
        #lbox.bind("<Double-Button-1>",self.overtaking_pair_selected)
        lbox.bind("<<ListboxSelect>>", self.display_overtaking_data)
        tkinter.Label(frame, text=" Overtaking", bg="white").grid(row=0, column=4)
        cols = ["Time","No. of Vehicles","Average Duration","Average Speed","No. of Maneouvres","No. Overtaking","No. Overtaken"]
        self.overtakingTree = ttk.Treeview(frame,columns=cols,height=12,show="headings")
        self.overtakingTree.grid(row = 1,rowspan = 4,column = 4,padx=20,pady=20,sticky="n")
        self.overtakingTree.heading(0, text="WERW")
        for i,c in enumerate(cols):
            self.overtakingTree.heading(i, text=c)
            self.overtakingTree.column(i, width=100, anchor=tkinter.CENTER)
        durationFrame = tkinter.Frame(frame,bg="white",relief=tkinter.GROOVE,borderwidth=2)
        f1 = tkinter.font.Font(family="helvetica", size=10)
        tkinter.Label(durationFrame,text = "Time Bin",font = f1,bg="white").grid(row = 0,column = 0)
        tkinter.Label(durationFrame, text="No of Vehicles", font=f1,bg="white").grid(row=0, column=1)
        for i in range(1,25):
            l = tkinter.Label(durationFrame, text="", font=f1,bg="white")
            l.grid(row=i, column=0)
            l.bind("<Double-Button-1>",self.select_time_bin)
            tkinter.Label(durationFrame, text="", font=f1,bg="white",width=10).grid(row=i, column=1)
        durationFrame.grid(row = 3,column=0,columnspan=3,pady=20)
        tkinter.Button(frame, text="Back", command=self.spawn_home_window, font=f,width=10).grid(row=4, column=0,columnspan=3, padx=20,pady=20)


        cols = ["Time", "2", "3", "4", "5","6", "7", "8", "9", "10", "11+"]

        platooningFrame = tkinter.Frame(self,bg="white")
        self.platoonTree1 = ttk.Treeview(platooningFrame,columns=cols,height=14,show="headings")
        self.platoonTree2 = ttk.Treeview(platooningFrame, columns=cols, height=14, show="headings")
        tkinter.Label(platooningFrame,text= "Platooning",bg="white").grid(row=0,column=0,columnspan=2,pady = 10)
        tkinter.Label(platooningFrame, text="Time = ", bg="white").grid(row=0, column=1, pady=10)
        e =tkinter.Entry(platooningFrame, text="5",width=4, bg="white", validate="key", validatecommand=vcmd)
        e.grid(row=0,column=2,sticky="w")
        e.delete(0,tkinter.END)
        e.insert(tkinter.END,"5")
        e.bind("<Return>",self.recalculate_platooning_data)
        e.bind("<Tab>", self.recalculate_platooning_data)
        e.bind("<FocusOut>", self.recalculate_platooning_data)
        for i, c in enumerate(cols):
            self.platoonTree1.heading(c, text=c)
            self.platoonTree1.column(c, width=50, anchor=tkinter.CENTER)
            self.platoonTree2.heading(c, text=c)
            self.platoonTree2.column(c, width=50, anchor=tkinter.CENTER)
        self.platoonTree1.column("Time", width=100, anchor=tkinter.CENTER)
        self.platoonTree2.column("Time", width=100, anchor=tkinter.CENTER)
        self.platoonTree1.grid(row=2,column=0,columnspan=3,padx=20,pady=20)
        self.platoonTree2.grid(row=3, column=0,columnspan=3,padx=20,pady=20)
        platooningFrame.grid(row =0,column=6,sticky="n")

    def recalculate_platooning_data(self,event):
        ###
        ### get the data relating to the pair parameter, and display it in the treeview
        ###

        ###
        ### find the selected movment pair
        ###
        frame = self.nametowidget(self.winfo_children()[0])
        children = frame.winfo_children()
        listBox = self.nametowidget(children[6])
        print("curselection is", listBox.curselection())
        pair = listBox.get(listBox.curselection()[0])
        pair = tuple([int(p) for p in pair.split("-")])
        print("pair is", pair)

        ###
        ### find the entry box for the platooning time
        ###
        frame = self.nametowidget(self.winfo_children()[1])
        e = self.nametowidget(frame.winfo_children()[4])
        try:
            platooningTime = int(e.get())
        except ValueError as e:
            platooningTime = 5
        try:
            if platooningTime != self.overtakingPairsDict[pair]["platooningTime"]:
                self.overtakingPairsDict[pair]["platooningTime"] = platooningTime
                self.overtakingPairsDict[pair]["data"][3] = self.recalcuatePlatooningfunction(self.currentJob, pair[0], platooningTime)
                self.overtakingPairsDict[pair]["data"][4] = self.recalcuatePlatooningfunction(self.currentJob, pair[1], platooningTime)
            self.display_overtaking_data(None)
        except KeyError as e:
            print("no pair",pair)

    def display_overtaking_data(self,event):
        ###
        ### display the previously retrieved data for a movement pair, selected by clicking in the list box
        ###

        ###
        ### find the movement pair
        ###

        if event is None:
            frame = self.nametowidget(self.winfo_children()[0])
            children = frame.winfo_children()
            listBox = self.nametowidget(children[6])
            print("curselection is", listBox.curselection())
            pair = listBox.get(listBox.curselection()[0])
            pair = tuple([int(p) for p in pair.split("-")])
            print("pair is", pair)
        else:
            print("curselection is",event.widget.curselection())
            pair = event.widget.get(event.widget.curselection()[0])
            pair = tuple([int(p) for p in pair.split("-")])
            print("pair is", pair)


        ###
        ### find the entry box for the platooning time
        ###
        frame = self.nametowidget(self.winfo_children()[1])
        e = self.nametowidget(frame.winfo_children()[4])


        self.overtakingTree.tag_configure("tree", font="courier 8")
        self.platoonTree1.tag_configure("tree", font="courier 8")
        self.platoonTree2.tag_configure("tree", font="courier 8")
        self.platoonTree1.tag_configure("total", font="helvetica 8",foreground="red")
        self.platoonTree2.tag_configure("total", font="helvetica 8", foreground="red")
        f = tkinter.font.Font(family="courier", size=8)
        try:
            data = self.overtakingPairsDict[pair]["data"] # data is [dataframe,binnedData,overtakingData,platoon1,platoon2]
            e.delete(0,tkinter.END)
            e.insert(0,self.overtakingPairsDict[pair]["platooningTime"])
            self.overtakingTree.delete(*self.overtakingTree.get_children())
            for item in data[2]:
                self.overtakingTree.insert("","end",values =item,tags=("tree",))
            frame = self.nametowidget(self.winfo_children()[0])
            frame = self.nametowidget(frame.winfo_children()[9])
            labels = frame.winfo_children()
            labelIndex = 2 ### because we want to start at the 3rd label
            for k,v in sorted(data[1].items()):
                if k == self.overtakingPairsDict[pair]["selected"]:
                    self.nametowidget(labels[labelIndex]).configure(text=k, font=f, bg="red",relief=tkinter.RAISED)
                else:
                    self.nametowidget(labels[labelIndex]).configure(text=k,font = f,bg="white",relief=tkinter.GROOVE)
                self.nametowidget(labels[labelIndex + 1]).configure(text=v,font = f,bg="white",relief=tkinter.GROOVE)
                labelIndex+=2
            self.platoonTree1.delete(*self.platoonTree1.get_children())
            self.platoonTree2.delete(*self.platoonTree2.get_children())
            for item in data[3][:-1]:
                self.platoonTree1.insert("", "end", values=item, tags=("tree",))
            self.platoonTree1.insert("", "end", values=data[3][-1], tags=("total",))
            for item in data[4][:-1]:
                self.platoonTree2.insert("", "end", values=item, tags=("tree",))
            self.platoonTree2.insert("", "end", values=data[4][-1], tags=("total",))
        except KeyError as e:
            print("no key found",pair)

    def select_time_bin(self,event):
        ###
        ### user has selected the time parameter to use to exclude any runs with time greater
        ### than the selected time. We set the "selected" entry of the dictionary to the selected time
        ### and then run display_overtaking_data to display the updated data.
        ###

        frame = self.nametowidget(self.winfo_children()[0])
        children = frame.winfo_children()
        listBox = self.nametowidget(children[6])
        if len(listBox.curselection()) == 0:
            messagebox.showinfo(message="Please select a movement pair from the list box")
            return
        pair = listBox.get(listBox.curselection()[0])
        pair = tuple([int(p) for p in pair.split("-")])
        try:
            self.overtakingPairsDict[pair]["selected"] = event.widget.cget("text")
            self.overtakingPairsDict[pair]["data"][2] = self.resampleOvertakingDataFunction(self.currentJob, self.overtakingPairsDict[pair]["data"][0],self.overtakingPairsDict[pair]["selected"], pair)

            parent = self.nametowidget(event.widget.winfo_parent())
            for child in parent.winfo_children():
                self.nametowidget(child).configure(bg="white")
            event.widget.configure(bg="red")
            self.display_overtaking_data(None)
        except KeyError as e:
            print(pair,"not found")

    def process_overtaking_pairs(self,listBox):
        ###
        ### this function runs in a separate thread from the main thread. This allows the user to enter
        ### new journey pairs while processing any previously entered pairs.
        ###


        ###
        ### find the entry box for the platooning time
        ###
        frame = self.nametowidget(self.winfo_children()[1])
        e = self.nametowidget(frame.winfo_children()[4])
        try:
            platooningTime = int(e.get())
        except ValueError as e:
            platooningTime = 5


        print("starting processing")
        while True:
            flag = False
            for i in range(listBox.size()):
                if listBox.itemcget(i, "bg") == "red":
                    pair = listBox.get(i)
                    pair = tuple([int(p) for p in pair.split("-")])
                    flag = True
                    self.overtakingPairsDict[pair] = {}
                    self.overtakingPairsDict[pair]["selected"] = "23:59:59"
                    self.overtakingPairsDict[pair]["platooningTime"] = platooningTime
                    result = self.getOvertakingDataFunction(self.currentJob,pair)
                    platoon1=self.recalcuatePlatooningfunction(self.currentJob,pair[0],platooningTime)
                    platoon2 = self.recalcuatePlatooningfunction(self.currentJob, pair[1], platooningTime)
                    overtaking = self.resampleOvertakingDataFunction(self.currentJob,result[0],self.overtakingPairsDict[pair]["selected"],pair)
                    result.append(overtaking)
                    result.append(platoon1)
                    result.append(platoon2)
                    self.overtakingPairsDict[pair]["data"] = result
                    listBox.itemconfig(i, bg="light green")
            if not flag:
                self.processOvertakingThread = None
                print("closing thread")
                return

    def add_overtaking_pair(self,event):
        ###
        ### the user has entered a new journey pair into the entry boxes, and added it to the list box for processing
        ### we start up a new thread to process the entered journey pair
        ###

        parent= event.widget.winfo_parent()
        parent = self.nametowidget(parent)
        children = parent.winfo_children()
        mvmt1 = self.nametowidget(children[3]).get()
        mvmt2 = self.nametowidget(children[5]) .get()
        self.nametowidget(children[3]).delete(0, tkinter.END)
        self.nametowidget(children[5]).delete(0, tkinter.END)
        w = self.nametowidget(children[3])
        w.focus()
        if mvmt1 == "" or mvmt2=="":
            return "break"
        listBox = self.nametowidget(children[6])

        listBox.insert(tkinter.END,mvmt1 + " - " + mvmt2)
        listBox.itemconfig(listBox.size()-1,bg="red")


        if self.processOvertakingThread is None:
            self.processOvertakingThread = threading.Thread(target=self.process_overtaking_pairs,args=(listBox,))
            self.processOvertakingThread.start()
        return "break"

    def display_option_changed(self,event):
        widget = event.widget
        print(widget.get(),widget.current())
        self.display_data(widget.current())

    def matrix_label_clicked(self,selectedMovement):
        ###
        ### when a user clicks on a movement number on the matrix
        ### the type of movement ( in, out, or both) is displayed in radio buttons
        ### in the extrasFrame of the matching results screen
        print("in matrix label clicked")
        if selectedMovement is None:
            print("no selected movement")
            return
        if self.radioVar is None:
            print("no radiovar")
            return
        if type(self.nametowidget(self.winfo_children()[0])) == tkinter.Toplevel: ### another window is open
            label = self.nametowidget(self.winfo_children()[2])
            innerframe = self.nametowidget(self.winfo_children()[1])
        else:
            frame = self.nametowidget(self.winfo_children()[1])
            innerframe = self.nametowidget(frame.winfo_children()[0])

        extrasFrame = self.nametowidget(innerframe.winfo_children()[2])
        for child in extrasFrame.winfo_children():
            widget = self.nametowidget(child)
            if type(widget) == tkinter.Radiobutton:
                widget.configure(state="normal")
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                #print("direction of movement",mvmtNo,"is",mvmt["dir"])
                if selectedMovement == mvmtNo:
                    #print("woohoo")
                    #lbl = self.nametowidget(extrasFrame.winfo_children()[1])
                    #lbl.configure(text="Movement "+ str(selectedMovement))
                    self.radioVar.trace_vdelete("w",self.radioVar.trace_id)
                    self.radioVar.set(mvmt["dir"])
                    self.radioVar.trace_id = self.radioVar.trace("w", self.direction_of_movement_changed)
                    self.display_movement_percentages(selectedMovement)

    def display_movement_percentages(self,selectedMovement):
        ###
        ### on each match screen, when user clicks on a movement label in the matrix, we want to display the
        ### percentage match for that movement.
        ###

        if type(self.nametowidget(self.winfo_children()[0])) == tkinter.Toplevel: ### another window is open
            label = self.nametowidget(self.winfo_children()[2])
            innerframe = self.nametowidget(self.winfo_children()[1])
        else:
            frame = self.nametowidget(self.winfo_children()[1])
            innerframe = self.nametowidget(frame.winfo_children()[0])
        extrasFrame = self.nametowidget(innerframe.winfo_children()[2])
        frame = self.nametowidget(extrasFrame.winfo_children()[0])
        children = frame.winfo_children()
        self.nametowidget(children[5]).configure(text=str(selectedMovement))
        self.nametowidget(children[6]).configure(text=self.currentJob["movementCounts"][selectedMovement][0])
        self.nametowidget(children[7]).configure(text=self.currentJob["movementCounts"][selectedMovement][1])
        self.nametowidget(children[8]).configure(text=self.currentJob["movementCounts"][selectedMovement][2])

    def direction_of_movement_changed(self,*args):
        ###
        ### user can change the direction of a movement (in, out, both) on the Non directional
        ### matching display. This function deals with that

        print(self.radioVar.get())
        if type(self.nametowidget(self.winfo_children()[0])) == tkinter.Toplevel: ### another window is open
            label = self.nametowidget(self.winfo_children()[2])
            innerframe = self.nametowidget(self.winfo_children()[1])
        else:
            frame = self.nametowidget(self.winfo_children()[1])
            innerframe = self.nametowidget(frame.winfo_children()[0])
        extrasFrame = self.nametowidget(innerframe.winfo_children()[2])
        frame = self.nametowidget(extrasFrame.winfo_children()[0])
        children = frame.winfo_children()
        selectedMovement =  self.nametowidget(children[5]).cget("text")
        print("selectedmovement is",selectedMovement)
        if selectedMovement == "":
            return
        print("movement number from label is",selectedMovement)
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                #print("direction of movement", mvmtNo, "is", mvmt["dir"])
                if int(selectedMovement) == mvmtNo:
                    print("site",site,"movement",mvmtNo,"old dir is",mvmt["dir"],"new dir is",self.radioVar.get())
                    mvmt["dir"]= self.radioVar.get()
                    print("site", site, "movement", mvmtNo, "dir is now", mvmt["dir"])
        self.display_data(0)

    def save_changes_to_movement_directions(self):
        data = {}
        data["job"] = self.currentJob
        sites = []
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                #print("site no",site,"mvmtno",mvmtNo,"original movments",mvmt["originalmovements"],"newmovement",mvmt["newmovement"])
                for m in mvmt["originalmovements"]:
                    sites.append([site,mvmtNo,m,mvmt["dir"],""])
        data["sites"] = sites
        myDB.save_Job(data,self.user)
        #self.updateDataFunction(self.currentJob)
        self.oldJobData = copy.deepcopy(self.currentJob)

    ###
    ### setting up, getting, and displaying the matching results
    ###


    def spawn_matching_results_screen(self,matchingType):

        if self.oldJobData is not None:
            self.currentJob = copy.deepcopy(self.oldJobData)
            self.oldJobData = None

        self.radioVar = None ### we only want a radioVar to exist for non directional matching.
        self.oldJobData = None ### we only want this to exist for non directional matching
        for child in self.winfo_children():
            if type(self.nametowidget(child)) != tkinter.Toplevel:
                child.destroy()
        print("screen settings")
        print("actual screen dimensions",self.winfo_screenwidth(),self.winfo_screenheight())
        print("window dimensions",self.winfo_width(),self.winfo_height())
        if self.winfo_screenheight() >900:
            width = 1000
            height = 800
        else:
            width = self.winfo_screenwidth() - 300
            height = self.winfo_screenheight() - 200
        print("set width and height to ",width,height)
        f = tkinter.font.Font(family="Helvetica", size=18, weight=tkinter.font.BOLD)
        tkinter.Label(self, bg="white", text=matchingType + " Matching", font=f, fg=self.tracsisBlue).grid(row=0,column=0,columnspan=3,pady = 20)

        outerFrame = tkinter.Frame(self,bg="white",width = 240,height = height, relief=tkinter.GROOVE, borderwidth=3)
        outerFrame.grid(row=1,column=0,padx=10,sticky="ns")
        outerFrame.grid_propagate(False)
        controlPanel = tkinter.Frame(outerFrame,bg="white")
        box = ttk.Combobox(controlPanel, width=10)
        box["values"] = ("Count", "Max", "Min", "Avg")
        box.bind("<<ComboboxSelected>>", self.display_option_changed)
        box.grid(row=0,column=0,pady=10)
        box.current(0)
        tkinter.Button(controlPanel, text="Match", bg="white", height=2, width=12,command=lambda: self.get_matches(matchingType)).grid(row=1,column=0,pady=10)
        extrasFrame = tkinter.Frame(controlPanel, bg="white")
        f = tkinter.font.Font(family="helvetica", size=8)
        frame = tkinter.Frame(extrasFrame, bg="white", relief=tkinter.RAISED, borderwidth=1)
        tkinter.Label(frame, text="Matched %ages", font=f, bg="Light blue").grid(row=0, column=0, columnspan=10,sticky="ew")
        tkinter.Label(frame, text="Mvmt", font=f, bg="Light blue").grid(row=1, column=0, sticky="ew")
        tkinter.Label(frame, text="Plates", font=f, bg="Light blue").grid(row=1, column=1, sticky="ew")
        tkinter.Label(frame, text="Matches", font=f, bg="Light blue").grid(row=1, column=2, sticky="ew")
        tkinter.Label(frame, text="Matched %", font=f, bg="Light blue").grid(row=1, column=3, sticky="ew")
        tkinter.Label(frame, text="", bg="white", font=f).grid(row=2, column=0)
        tkinter.Label(frame, text="0", bg="white", font=f).grid(row=2, column=1)
        tkinter.Label(frame, text="0", bg="white", font=f).grid(row=2, column=2)
        tkinter.Label(frame, text="0%", bg="white", font=f).grid(row=2, column=3)
        frame.grid(row=0, column=0, pady=(0, 10),columnspan=2)
        matrixFrame = tkinter.Frame(self, relief=tkinter.GROOVE, borderwidth=3, bg="white", width=width, height=height)
        self.matrix = matrix.MatrixDisplay(matrixFrame, width, height)
        self.matrix.enable_click()
        self.matrix.set_matrix_clicked_callback_function(self.display_movement_percentages)
        if matchingType == "Directional":
            f = tkinter.font.Font(family="courier", size=8)
            var = tkinter.IntVar()
            ch = tkinter.Checkbutton(extrasFrame, text="In-Out   ", font=f, bg="white", variable=var)
            ch.grid(row=1, column=0, columnspan=4)
            ch.select()
            ch.var = var
            var = tkinter.IntVar()
            ch = tkinter.Checkbutton(extrasFrame, text="In-Both  ", font=f, bg="white", variable=var)
            ch.grid(row=2, column=0, columnspan=4)
            ch.var = var
            var = tkinter.IntVar()
            ch = tkinter.Checkbutton(extrasFrame, text="Both-Out ", font=f, bg="white", variable=var)
            ch.grid(row=3, column=0, columnspan=4)
            ch.var = var
            var = tkinter.IntVar()
            ch = tkinter.Checkbutton(extrasFrame, text="Both-Both", font=f, bg="white", variable=var)
            ch.grid(row=4, column=0, columnspan=4)
            ch.var = var
            f = tkinter.font.Font(family='Arial', size=8)
            l = tkinter.Label(controlPanel,text="Switch to Non Directional",font = f,bg="white")
            l.grid(row=3,column=0,padx =10)
            l.bind("<Enter>", self.on_label_entry)
            l.bind("<Leave>", self.on_label_exit)
            l.bind("<Button-1>", lambda event:self.spawn_matching_results_screen("Non Directional"))


        if matchingType == "Non Directional":
            f = tkinter.font.Font(family="courier", size=8)
            self.radioVar = tkinter.IntVar()
            self.matrix.set_matrix_clicked_callback_function(self.matrix_label_clicked)
            #tkinter.Label(extrasFrame,text="Movement ",bg="white",font = f).grid(row=1,column=0)
            self.radioVar.trace_id = self.radioVar.trace("w",self.direction_of_movement_changed)
            tkinter.Radiobutton(extrasFrame, text="In  ", bg="white", font=f, value=1,variable=self.radioVar,state="disabled").grid(row=2, column=0)
            tkinter.Radiobutton(extrasFrame, text="Out ", bg="white", font=f, value=2,variable=self.radioVar,state="disabled").grid(row=3, column=0)
            tkinter.Radiobutton(extrasFrame, text="Both", bg="white", font=f, value=3,variable=self.radioVar,state="disabled").grid(row=4, column=0)
            tkinter.Button(extrasFrame,text="Save",font=f,command=self.save_changes_to_movement_directions).grid(row=5,column=0)
            f = tkinter.font.Font(family='Arial', size=8)
            l = tkinter.Label(controlPanel, text="Switch to Directional", font=f, bg="white")
            l.grid(row=3, column=0, padx=10)
            l.bind("<Enter>", self.on_label_entry)
            l.bind("<Leave>", self.on_label_exit)
            l.bind("<Button-1>", lambda event: self.spawn_matching_results_screen("Directional"))


        if matchingType == "First Seen/Last Seen":
            f = tkinter.font.Font(family="courier", size=8)
            f = tkinter.font.Font(family='Arial', size=8)
            l = tkinter.Label(controlPanel, text="Switch to Journey Pairs", font=f, bg="white")
            l.grid(row=3, column=0, padx=10)
            l.bind("<Enter>", self.on_label_entry)
            l.bind("<Leave>", self.on_label_exit)
            l.bind("<Button-1>", lambda event: self.spawn_matching_results_screen("All Journey Pairs"))
            var = tkinter.IntVar()
            tkinter.Label(extrasFrame,text="(applies to full journeys only)", font=f).grid(row=6,column=0,columnspan=3)
            check = tkinter.Checkbutton(extrasFrame, text="Duration Check", bg="white", font=f,
                                        command=lambda: self.duration_check_selected(var), variable=var)
            check.grid(row=7, column=0, columnspan=3)
            check.var = var
            self.durationVar = tkinter.IntVar()
            tkinter.Radiobutton(extrasFrame, text="Split   ", bg="white", font=f, variable=self.durationVar, value=1,
                                state="disabled").grid(row=8, column=0, columnspan=3)
            tkinter.Radiobutton(extrasFrame, text="Discard", bg="white", font=f, variable=self.durationVar, value=2,
                                state="disabled").grid(row=9, column=0, columnspan=3)
            self.durationVar.set(1)

        if matchingType == "All Journey Pairs":
            f = tkinter.font.Font(family="courier", size=8)
            f = tkinter.font.Font(family='Arial', size=8)
            l = tkinter.Label(controlPanel, text="Switch to First Seen/Last Seen", font=f, bg="white")
            l.grid(row=3, column=0, padx=10)
            l.bind("<Enter>", self.on_label_entry)
            l.bind("<Leave>", self.on_label_exit)
            l.bind("<Button-1>", lambda event: self.spawn_matching_results_screen("First Seen/Last Seen"))

        if matchingType == "Filtered":


            tkinter.Label(extrasFrame,text="Enter Filter String",bg="white",font = f).grid(row=1,column=0, columnspan=3)
            e = tkinter.Entry(extrasFrame, width=15)
            e.grid(row=2, column=0, padx=10, pady=10, columnspan=3)
            tkinter.Label(extrasFrame, text="Current Filters", bg="white", font=f).grid(row=3, column=0, columnspan=3)
            lb = tkinter.Listbox(extrasFrame, bg="white", font=f)
            lb.bind("<Double-Button-1>", self.remove_filter)
            lb.grid(row=4, column=0, columnspan=3)
            e.bind("<Return>", lambda event: self.add_filter(event, lb))
            tkinter.Button(extrasFrame, text="Clear", bg="white", height=1, width=5,
                           command=lambda: lb.delete(0, tkinter.END)).grid(row=5, column=0, pady=10)
            tkinter.Button(extrasFrame, text="Run", bg="white", height=1, width=5,
                           command=lambda: self.get_matches("Filtered")).grid(row=5, column=1, pady=10)
            var = tkinter.IntVar()
            check = tkinter.Checkbutton(extrasFrame, text="Duration Check", bg="white", font=f,
                                        command=lambda: self.duration_check_selected(var), variable=var)
            check.grid(row=6, column=0, columnspan=3)
            check.var = var
            self.durationVar = tkinter.IntVar()
            tkinter.Radiobutton(extrasFrame, text="Split   ", bg="white", font=f, variable=self.durationVar, value=1,
                                state="disabled").grid(row=7, column=0, columnspan=3)
            tkinter.Radiobutton(extrasFrame, text="Discard", bg="white", font=f, variable=self.durationVar, value=2,
                                state="disabled").grid(row=8, column=0, columnspan=3)
            self.durationVar.set(1)


        controlPanel.grid(row=1,column=0,sticky="ns")
        self.update()
        print("control panel width",controlPanel.winfo_width(),controlPanel.winfo_reqwidth())
        controlPanel.grid_configure(padx = (198-controlPanel.winfo_width())/2)
        if extrasFrame is not None:
            extrasFrame.grid(row=2,column=0,pady=10)
        matrixFrame.grid(row=1, column=1)
        matrixFrame.grid_propagate(False)
        tkinter.Button(self, text="Back", bg="white",command=self.spawn_home_window,width=12).grid(row=4, column=0, padx=10, pady=10,sticky = "s")

    def get_matches(self,matchType):
        if type(self.nametowidget(self.winfo_children()[0])) == tkinter.Toplevel: ### another window is open
            print("duration matrix open")
            innerframe = self.nametowidget(self.winfo_children()[2])
        else:
            innerframe = self.nametowidget(self.winfo_children()[1])
        controlPanel = self.nametowidget(innerframe.winfo_children()[0])
        extrasFrame = self.nametowidget(controlPanel.winfo_children()[2])
        if matchType == "Directional":
            checkboxes = []
            for child in extrasFrame.winfo_children():
                if type(child) == tkinter.Checkbutton:
                    checkboxes.append(self.nametowidget(child).var.get())
            self.matrixData = self.getCordonFunction(self.currentJob,checkboxes)
        if matchType == "Non Directional":
            self.oldJobData = copy.deepcopy(self.currentJob)
            self.matrixData = self.getNonDirectionalCordonFunction(self.currentJob)
        if matchType == "All Journey Pairs":
            self.matrixData = self.getJourneyPairsFunction(self.currentJob)
        if matchType == "First Seen/Last Seen":
            for child in extrasFrame.winfo_children():
                widget = self.nametowidget(child)
                if type(self.nametowidget(child)) == tkinter.Checkbutton:
                    durationCheck = widget.var.get()
            self.matrixData = self.getRouteAssignmentFsLsFunction(self.currentJob,durationCheck,self.durationVar.get(),None)
        if matchType == "Filtered":
            filters = []
            durationCheck = 0
            lb = self.nametowidget(extrasFrame.winfo_children()[4])
            if lb.get(0) == "ALL":
                filters = ["I-B*-O", "(I-B-B*)-I", "O-(B-B*-O)", "I-B-B*!", "^B-B*-O", "^B-B-B*!"]
            else:
                for row in lb.get(0, tkinter.END):
                    try:
                        filters.append(row)
                    except Exception as e:
                        pass
            print("filters are",filters)
            for child in extrasFrame.winfo_children():
                widget = self.nametowidget(child)
                if type(self.nametowidget(child)) == tkinter.Checkbutton:
                    durationCheck = widget.var.get()
            self.matrixData = self.filteredMatchingfunction(self.currentJob, filters, durationCheck,self.durationVar.get())

        self.display_data(0)

    def display_data(self,index):
        if self.matrixData is None:
            return
        inMov = []
        outMov=[]
        #for site, details in self.currentJob["sites"].items():
            #for mvmtNo, mvmt in details.items():
                #if not int(mvmt["newmovement"]) in inMov:
                    #inMov.append(int(mvmt["newmovement"]))
                #if not int(mvmt["newmovement"]) in outMov:
                    #outMov.append(int(mvmt["newmovement"]))
        #inMov = sorted(inMov)
        #outMov = sorted(outMov)
        data = {}
        for k, v in self.matrixData[0].items():
            if k[0] not in inMov:
                inMov.append(k[0])
            if k[1] not in outMov:
                outMov.append(k[1])
            data[k] = v[index]
        inMov = sorted(inMov)
        outMov = sorted(outMov)
        if index == 0:
            for i, mov in enumerate(inMov):
                data[(mov, "total")] = int(self.matrixData[1][i])
            for i, mov in enumerate(outMov):
                data[("total", mov)] = int(self.matrixData[2][i])
            inMov.append("total")
            outMov.append("total")
            data[("total", "total")] = int(sum(self.matrixData[1]))
            self.matrix.draw(inMov, outMov, data,self.currentJob)
        else:
            self.matrix.draw(inMov, outMov, data,self.currentJob, fontsize=6)

    def add_filter(self,event,lb):
        ###
        ### this fuction allows the user to add routes to a listbox, eg, 1,2,3,4
        ### which is then passed to the cordon traversal function to filter on specified routes
        ###

        text = event.widget.get()
        if text == "":
            return
        try:
            i = lb.get(0,tkinter.END).index(text)
            ###its already in the list, we dont want to add it again
        except Exception as e:
            ### its not in the list, so we add it to the end
            if self.validate_filter(text):
                lb.insert(tkinter.END,text)
        event.widget.delete(0,tkinter.END)

    def validate_filter(self,filter):
        if filter == "ALL":
            return True
        if "(" in filter:
            if (")" in filter and filter.index("(") > filter.index(")")) or not ")" in filter:
                messagebox.showinfo(message="Incorrect Brackets")
                return False
        if ")" in filter:
            if ("(" in filter and filter.index("(") > filter.index(")")) or not "(" in filter:
                messagebox.showinfo(message="Incorrect Brackets")
                return False
        tokens = filter.split("-")
        if len(tokens) < 2:
            messagebox.showinfo(message="You need to have 2 or more tokens in a filter")
            return False
        if "^"  in filter and not "^" in tokens[0]:
            messagebox.showinfo(message="If you are using ^, it must always be the first character in the filter")
            return False
        if "!"  in filter and not "!" in tokens[-1]:
            messagebox.showinfo(message="If you are using !, it must always be the last character in the filter")
            return False
        if "" in tokens:
            messagebox.showinfo(message="Blank tokens not allowed")
            return False
        for i,t in enumerate(tokens[:-1]):
            if "*" in t:
                temp = t.replace("*","")
                print("temp is",temp,"t is",t)
                if temp == tokens[i+1]:
                    temp = t
                    tokens[i] = tokens[i+1]
                    tokens[i+1] = temp
        for t in tokens:
            t = t.replace("*","")
            t = t.replace("^", "")
            t = t.replace("!", "")
            t = t.replace("¬", "")
            t = t.replace("(", "")
            t = t.replace(")", "")
            if t == "":
                messagebox.showinfo(message="A token cannot only contain a special char(!,^,*,¬)")
                return False
            if t not in ["I","B","O"]:
                try:  #### is it numeric?
                    temp = int(t)
                except ValueError as e:
                    messagebox.showinfo(message="A token must contain I,B,O or a number")
                    return False
                return True
            else:
                return True

    def load_filter_from_csv(self,lb):
        lb.delete(0,tkinter.END)
        file = filedialog.askopenfilename()
        if file == "":
            return
        filters = []
        with open(file, 'rt') as file:
            spamreader = csv.reader(file, delimiter=',')
            for row in spamreader:
                try:
                    journey = [int(item) for item in row if item != ""]
                    filters.append(journey)
                except Exception as e:
                    pass
        print(filters)
        for journey in filters:
            lb.insert(tkinter.END,",".join(map(str,journey)))

    def spawn_filter_matching_screen(self):
        for child in self.winfo_children():
            if type(self.nametowidget(child)) != tkinter.Toplevel:
                child.destroy()
        width = self.winfo_screenwidth() - 120
        height = self.winfo_screenheight() - 50
        frame = tkinter.Frame(self, bg="white")
        box = ttk.Combobox(frame, width=10)
        box["values"] = ("Count", "Max", "Min", "Avg")
        box.bind("<<ComboboxSelected>>", self.filtered_display_option_changed)
        box.grid(row=0, column=0,columnspan=3,sticky="n")
        box.current(0)

        e = tkinter.Entry(frame, width=15)
        e.grid(row=1, column=0, padx=10, pady=10,columnspan = 3)
        f = tkinter.font.Font(family="helvetica", size=8)
        lb = tkinter.Listbox(frame, bg="white", font=f)
        lb.bind("<Double-Button-1>",self.remove_filter)
        lb.grid(row=2, column=0,columnspan =3)
        e.bind("<Return>", lambda event: self.add_filter(event, lb))
        tkinter.Button(frame, text="Clear", bg="white", height=1, width=5,command=lambda: lb.delete(0,tkinter.END)).grid(row=3, column=0, pady=10)
        tkinter.Button(frame, text="Run", bg="white", height=1, width=5,command=lambda: self.calculate_filtered_matching(lb,0)).grid(row=3, column=1, pady=10)
        var = tkinter.IntVar()
        check = tkinter.Checkbutton(frame,text = "Duration Check",bg="white",font=f,command=lambda:self.duration_check_selected(var),variable=var)
        check.grid(row=4,column=0,columnspan=3)
        check.var = var
        self.durationVar = tkinter.IntVar()
        tkinter.Radiobutton(frame,text = "Split   ",bg="white",font=f,variable=self.durationVar,value =1,state="disabled").grid(row=5,column=0,columnspan=3)
        tkinter.Radiobutton(frame, text="Discard", bg="white", font=f, variable=self.durationVar, value=2,state="disabled").grid(row=6,column=0,columnspan=3)
        self.durationVar.set(1)
        tkinter.Button(frame, text="Back", bg="white", height=1, width=5,command=self.spawn_home_window).grid(row=7, column=0, padx=10, pady=10,columnspan =3)
        frame.grid(row=0, column=0, padx=20, pady=10, sticky="w", rowspan=5)
        tkinter.Label(self, bg="white", text="Filtered Matching").grid(row=0, column=1)
        frame = tkinter.Frame(self, bg="white", relief=tkinter.GROOVE, borderwidth=3, width=800, height=800)
        frame.grid(row=1, column=1,sticky="n")
        frame.grid_propagate(False)
        self.matrix = matrix.MatrixDisplay(frame, width, height)

        frame = tkinter.Frame(bg="white", relief=tkinter.SUNKEN, borderwidth=2, height=700)
        frame.grid(row=1, column=2, sticky="ne", padx=(50, 0))
        inMov = []
        numRows = 40
        maxRows = 25
        i = 2
        while numRows / i > maxRows: i += 1
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                if mvmt["newmovement"] not in inMov:
                    inMov.append(mvmt["newmovement"])
        inMov = sorted(inMov)
        f = tkinter.font.Font(family="helvetica", size=8)

        tkinter.Label(frame, text="Matched %ages", font=f, bg="Light blue").grid(row=0, column=0,columnspan=10, sticky="ew")
        tkinter.Label(frame, text="Mvmt", font=f, bg="Light blue").grid(row=1, column=0,  sticky="ew")
        tkinter.Label(frame, text="Plates", font=f, bg="Light blue").grid(row=1, column=1,  sticky="ew")
        tkinter.Label(frame, text="Matches", font=f, bg="Light blue").grid(row=1, column=2, sticky="ew")
        tkinter.Label(frame, text="Matched %", font=f, bg="Light blue").grid(row=1, column=3,  sticky="ew")


        for count, i in enumerate(inMov):
            tkinter.Label(frame, text="Mvmt " + str(i), bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2))
            tkinter.Label(frame,text="0", bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2)+1)
            tkinter.Label(frame, text="0", bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2) + 2)
            tkinter.Label(frame, text="0%", bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2) + 3)

    def duration_check_selected(self,var):
        value = var.get()
        if type(self.nametowidget(self.winfo_children()[0])) == tkinter.Toplevel: ### another window is open
            print("duration matrix open")
            innerframe = self.nametowidget(self.winfo_children()[2])
        else:
            innerframe = self.nametowidget(self.winfo_children()[1])
        controlPanel = self.nametowidget(innerframe.winfo_children()[0])
        extrasFrame = self.nametowidget(controlPanel.winfo_children()[2])

        for child in extrasFrame.winfo_children():
            widget = self.nametowidget(child)
            if type(self.nametowidget(child)) == tkinter.Radiobutton:
                if value == 1:
                    widget.configure(state="normal")
                else:
                    widget.configure(state="disabled")

    def remove_filter(self,event):
        lb = event.widget
        if len(lb.curselection())==0:
            return
        lb.delete(lb.curselection()[0])

    def calculate_filtered_matching(self,lb,index,get_data=True):
        ###
        ### user can specify a list of filters to apply for matching
        ###
        filters = []
        durationCheck = 0
        print("first row of list is",lb.get(0))
        if lb.get(0) == "ALL":
            filters = ["I-B*-O","(I-B-B*)-I","O-(B-B*-O)","I-B-B*!","^B-B*-O","^B-B-B*!"]
        else:
            for row in lb.get(0, tkinter.END):
                try:
                    filters.append(row)
                except Exception as e:
                    pass
        print("filters are",filters)
        if type(self.nametowidget(self.winfo_children()[0])) == tkinter.Toplevel:
            frame = self.nametowidget(self.winfo_children()[1])
        else:
            frame = self.nametowidget(self.winfo_children()[0])
        for child in frame.winfo_children():
            widget = self.nametowidget(child)
            if type(self.nametowidget(child)) == tkinter.Checkbutton:
                durationCheck= widget.var.get()

        if get_data:
            self.matrixData = self.filteredMatchingfunction(self.currentJob,filters,durationCheck,self.durationVar.get())
        f = tkinter.font.Font(family="helvetica", size=8)
        numRows = 30
        count = 0
        parent = self.nametowidget(frame.winfo_parent())
        percentagesFrame = self.nametowidget(parent.winfo_children()[3])
        for child in percentagesFrame.winfo_children()[5:]:
            child.destroy()
        for key,value in self.currentJob["movementCounts"].items():
            tkinter.Label(percentagesFrame, text="Mvmt " + str(key), bg="white", font=f).grid(row=(count % numRows) + 2, column=(int(count / numRows) * 2),padx=5)
            tkinter.Label(percentagesFrame, text=str(value[0]), bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2) + 1,padx=5)
            tkinter.Label(percentagesFrame, text=str(value[1]), bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2) + 2,padx=5)
            tkinter.Label(percentagesFrame, text=str(value[2]) + "%", bg="white", font=f).grid(row=(count % numRows) + 2,column=(int(count / numRows) * 2) + 3,padx=5)

            count+=1
        inMov = []
        outMov = []
        for key, item in self.matrixData[0].items():
            if not int(key[0]) in inMov:
                inMov.append(int(key[0]))
            if not int(key[1]) in outMov:
                outMov.append(int(key[1]))
        inMov = sorted(inMov)
        outMov = sorted(outMov)
        box = self.nametowidget(frame.winfo_children()[0])
        box.current(index)
        print(inMov,outMov)
        data = {}
        for k, v in self.matrixData[0].items():
            data[k] = v[index]
        if index == 0:
            for i, mov in enumerate(inMov):
                data[(mov, "total")] = int(self.matrixData[1][i])
            for i, mov in enumerate(outMov):
                data[("total", mov)] = int(self.matrixData[2][i])
            inMov.append("total")
            outMov.append("total")
            data[("total", "total")] = int(sum(self.matrixData[1]))
            self.matrix.draw(inMov, outMov, data,self.currentJob)
        else:
            self.matrix.draw(inMov, outMov, data,self.currentJob, fontsize=6)

    def spawn_duration_matrix_screen(self):

        win =tkinter.Toplevel()
        win.state("zoomed")
        width = win.winfo_screenwidth() - 120
        height = win.winfo_screenheight() - 200

        frame=tkinter.Frame(win)
        frame.grid(row=0, column=0, columnspan=3)
        tkinter.Label(frame,text = "Base duration").grid(row=0,column=0)
        vcmd = (self.register(self.validate_time_cell_input), "%d", "%s", "%S")
        e =tkinter.Entry(frame, validate="key", validatecommand=vcmd)
        e.grid(row=0,column=1)
        tkinter.Button(frame, text="Fill", command=lambda: self.fill_duration_matrix(e)).grid(row=0, column=2)
        frame = tkinter.Frame(win, bg="white", relief=tkinter.GROOVE, borderwidth=3, width=800, height=800)
        frame.grid(row = 1,column=0,columnspan = 3,padx=10,pady=10)
        self.durationMatrix = matrix.MatrixDisplay(frame, width, height)
        self.durationMatrix.set_matrix_clicked_callback_function(self.duration_matrix_clicked,True)
        self.draw_duration_matrix_screen()

    def fill_duration_matrix(self,e):
        text = e.get()
        if text.strip() == "":
            return
        text = self.validate_text_to_hhmm(text)
        if text != "":
            e.delete(0, tkinter.END)
            for k,v in self.durationsDictionary.items():
                    self.durationsDictionary[k] = text
            with open(self.currentJob["folder"] + "/data/durations.pkl", "wb") as f:
                pickle.dump(self.durationsDictionary, f)
            self.draw_duration_matrix_screen()
        else:
            messagebox.showinfo(message=e.get() + " is not a valid time")
            e.delete(0, tkinter.END)

    def validate_text_to_hhmm(self,text):
        if len(text) > 5:
            return ""
        elif len(text) == 1:
            if not text.isdigit():
                return ""
            text = "0" + text + ":00"
        elif len(text) == 2:
            if not text.isdigit():
                return ""
            text = text + ":00"
        elif len(text) == 3 and text[-1] == ":":
            if not text[:-1].isdigit():
                return ""
            text = text + "00"
        elif len(text) == 4:
            text = text + "0"
        hours = text.split(":")[0]
        mins = text.split(":")[1]
        if int(hours) > 23 or int(mins) > 60:
            return ""
        return text

    def draw_duration_matrix_screen(self):
        inMov = []
        outMov = []
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                if mvmt["newmovement"] not in inMov:
                    inMov.append(mvmt["newmovement"])
                if mvmt["newmovement"] not in outMov:
                    outMov.append(mvmt["newmovement"])
        inMov = sorted(inMov)
        outMov = sorted(outMov)
        if self.durationsDictionary is None:
            self.durationsDictionary = {}
            try:
                with open(self.currentJob["folder"] + "/data/durations.pkl","rb") as f:
                    self.durationsDictionary=pickle.load(f)
            except IOError as e:
                print(e)
                messagebox.showinfo(
                    message="Durations havent been previously set up, filling out matrix with base value of 30 minutes")
                for i in inMov:
                    for o in outMov:
                        self.durationsDictionary[i, o] = "00:30:00"
        self.currentJob["durationsDictionary"]=self.durationsDictionary
        with open(self.currentJob["folder"] + "/data/durations.pkl","wb") as f:
            pickle.dump(self.durationsDictionary,f)
        self.durationMatrix.draw(inMov,outMov,self.durationsDictionary,self.currentJob)

    def duration_matrix_clicked(self,row,column):

        win=tkinter.Toplevel()
        win.wm_title("in " + str(row) + " - out " + str(column))
        frame = tkinter.Frame(win)
        tkinter.Label(frame,text="Enter Duration").grid(row=0,column=0)
        vcmd = (self.register(self.validate_time_cell_input), "%d", "%s", "%S")
        e =tkinter.Entry(frame,validate="key",validatecommand=vcmd)
        e.grid(row=0,column=1)
        tkinter.Button(frame,text="Save",command=lambda :self.save_duration_matrix(e,win)).grid(row=0,column=2)
        frame.grid(row=0,column=0)
        e.insert(0,self.durationsDictionary[row,column])

    def save_duration_matrix(self,e,win):
        ###
        ### save the time entered in the pop up box into the durations dictionary, if it is a correctly formatted time
        ###
        text = e.get()
        if text.strip() == "":
            return
        text = self.validate_text_to_hhmm(text)
        if text != "":
            temp = win.title().split("-")
            i,o= temp[0],temp[1]
            i = int(i.replace("in","").strip())
            o = int(o.replace("out", "").strip())
            self.durationsDictionary[i,o]=text
            print("setting value for",self.durationsDictionary[i,o],"to",text)
            with open(self.currentJob["folder"] + "/data/durations.pkl", "wb") as f:
                pickle.dump(self.durationsDictionary, f)
            self.draw_duration_matrix_screen()
            win.destroy()
        else:
            messagebox.showinfo(message=e.get() + " is not a valid time")
            e.delete(0,tkinter.END)

    def scroll_matrix_screen(self,event):
        print(event)
        print(event.widget.cget("orient"), event.x, event.y)

        if event.widget.cget("orient") == "vertical":
            top, bottom = (event.widget.get())
            thumbsize = bottom - top
            f = event.widget.fraction(event.x, event.y)
            if f < top:
                f = f - (thumbsize / 2)
            self.matrixCanvasList[0].yview_moveto(f)
            self.matrixCanvasList[1].yview_moveto(f)
            return "break"
        else:
            left, right = (event.widget.get())
            thumbsize = right - left
            f = event.widget.fraction(event.x, event.y)
            if f < left:
                f = f - (thumbsize / 2)
            self.matrixCanvasList[0].xview_moveto(f)
            self.matrixCanvasList[2].xview_moveto(f)

    def spawn_survey_setup_screen(self):
        ###
        ### The screen that displays a list of current projects
        ### and allows you to add, edit or delete projects
        ###

        try:
            self.joblist = myDB.get_jobs()
        except :
            messagebox.showinfo(message="Couldnt open database, please select a database file")
            self.spawn_settings_window()
            return


        self.entryValues = []

        for child in self.winfo_children():
            child.destroy()
        self.wm_title("Project Setup")

        self.menubar = tkinter.Menu(self)
        menu = tkinter.Menu(self.menubar, tearoff=0)
        menu.add_command(label="Change Settings", command=self.spawn_settings_window)
        self.menubar.add_cascade(label="Settings", menu=menu)
        self.config(menu=self.menubar)

        frame = tkinter.Frame(self,  bg="white")
        print("screenheight is ",self.winfo_screenheight())
        treefontsize = 8
        fontsize = 12
        headingWidth = 120
        if self.winfo_screenheight() < 1000:
            fontsize = 10
            treefontsize = 7
            headingWidth = 100
        if self.winfo_screenheight() < 900:
            fontsize = 8
            treefontsize = 8
            headingWidth = 85
        f = tkinter.font.nametofont("TkDefaultFont").configure(size=fontsize)
        tkinter.Button(frame, text="Create new ANPR \nProject", bg="white",fg=self.tracsisBlue, height=3,command=self.spawn_parameters_window).grid(row=0, column=0, padx=20, pady=20)
        tkinter.Button(frame, text="Edit ANPR \nProject", width=17, height=3, bg="white",fg=self.tracsisBlue,command=self.edit_job).grid(row=0, column=1, padx=20,pady=20)
        tkinter.Button(frame, text="Delete ANPR \nProject", width=17, height=3, bg="white",fg=self.tracsisBlue,command=self.delete_job).grid(row=0, column=2, padx=20,pady=20)

        frame.grid(row=0, column=0,pady=(20,0),padx=(120,0))
        frame = tkinter.Frame(self, bg="white")

        ### set up the treeview to display the ANPR projects

        cols = ("Job No","Job Name","Survey Date","Survey Times","OV Template","OV Counts","Unclassed VRN","Classed VRN","Comparison","Completed","Created By","Created Date")

        self.tree = ttk.Treeview(frame,columns=tuple(range(len(cols))),show="headings",height = 30)
        style = ttk.Style()
        style.configure(".", font=('Helvetica', 6), foreground="white")
        style.configure("Treeview.Heading", foreground=self.tracsisBlue,background="black")
        self.tree.bind("<Double-Button-1>", self.load_job)
        self.tree.heading(0,text="WERW")
        self.tree.tag_configure("odd", background="white",foreground=self.tracsisBlue)
        self.tree.tag_configure("even", background="azure2", foreground=self.tracsisBlue)
        self.tree.tag_configure("grn",foreground="dark blue")
        for i,col in enumerate(cols):
            self.tree.heading(i,text=col)
            self.tree.column(i,width  = headingWidth,anchor=tkinter.CENTER)
        for i in range(2):
            self.tree.column(i,width=headingWidth+30)
            self.tree.grid(row=0,column=0)
            self.tree.tag_configure("tree",font="courier " + str(treefontsize))
        for i,job in enumerate(self.joblist):
            if i%2 == 0:
                self.tree.insert("","end",values =job,tags=("tree","even"))
            else:
                self.tree.insert("", "end", values=job, tags=("tree", "odd"))
        #self.tree.configure(height =len(self.joblist))
        frame.grid(row=1, column=0,padx=(120,0))

    def spawn_parameters_window(self):

        for child in self.winfo_children():
            child.destroy()
        win = tkinter.Frame(self,width= 1500,height = 900,bg = "white")
        #win.grid_propagate(False)
        win.grid(row=0,column=0)
        outerFrame = tkinter.Frame(win,bg = "white")

        ###
        ### set up the top left frame
        ###

        f = tkinter.font.Font(family="helvetica", size=9)

        frame = tkinter.Frame(outerFrame,width=330,height = 150,bg = "white",relief=tkinter.GROOVE,borderwidth=2)
        #frame.grid_propagate(False)
        tkinter.Label(frame, text="Job No", bg="white",font = f).grid(row=0, column=0,sticky="e")
        tkinter.Label(frame,text = "Job Name",bg = "white",font = f).grid(row=1,column = 0,sticky="e")
        tkinter.Label(frame, text="Date",bg = "white",font = f).grid(row=2, column=0,sticky="e")
        self.entryValues.append(tkinter.StringVar())
        tkinter.Entry(frame, width=20, textvariable=self.entryValues[-1],bg = "white").grid(row=0, column=1, pady=10, padx=10)
        self.entryValues.append(tkinter.StringVar())
        tkinter.Entry(frame, width=20, textvariable=self.entryValues[-1],bg = "white").grid(row=1, column=1, pady=10, padx=10,
                                                                               sticky="w")

        vcmd = (self.register(self.validate_date),"%d","%s","%S")
        self.entryValues.append(tkinter.StringVar())
        e = tkinter.Entry(frame, width=20, textvariable=self.entryValues[-1], bg="white",validate="key",validatecommand=vcmd)
        e.grid(row=2, column=1, pady=10,padx=10, sticky="w")
        e.bind("<FocusOut>", self.validate_date_on_focus_out)



        frame.grid(row=0, column=0)

        self.update()
        width = frame.winfo_reqwidth()




        ###
        ### set up mid left frame
        ###
        vcmd = (self.register(self.validate_time_cell_input),"%d", "%s","%S")
        frame = tkinter.Frame(outerFrame,width=width,height=260,bg = "white",relief=tkinter.GROOVE,borderwidth=2)
        frame.grid_propagate(False)
        tkinter.Label(frame, text="From", bg="white",font = f).grid(row=0, column=1, pady=10, padx=5)
        tkinter.Label(frame, text="To", bg="white",font = f).grid(row=0, column=2, pady=10, padx=5)
        tkinter.Label(frame, text="Time 1",bg = "white",font = f).grid(row=1, column=0,pady = 10,padx = 5,sticky="e")
        tkinter.Label(frame, text="Time 2",bg = "white",font = f).grid(row=2, column=0,pady = 10,padx = 5,sticky="e")
        tkinter.Label(frame, text="Time 3",bg = "white",font = f).grid(row=3, column=0,pady = 10,padx = 5,sticky="e")
        tkinter.Label(frame, text="Time 4",bg = "white",font = f).grid(row=4, column=0,pady = 10,padx = 5,sticky="e")
        for i in range(1, 5):
            self.entryValues.append(tkinter.StringVar())
            e = tkinter.Entry(frame, width=7, textvariable=self.entryValues[-1], bg="white",validate="key",validatecommand=vcmd)
            e.grid(row=i, column=1, pady=10,padx=5)
            e.bind("<FocusOut>",self.validate_hhmm)
            self.entryValues.append(tkinter.StringVar())
            e = tkinter.Entry(frame, width=7, textvariable=self.entryValues[-1], bg="white",validate="key",validatecommand=vcmd)
            e.grid(row=i, column=2, pady=10,padx=5)
            e.bind("<FocusOut>", self.validate_hhmm)
        tkinter.Label(frame, text="Interval", bg="white",font = f).grid(row=5, column=0, pady=10, padx=10, sticky="e")
        self.entryValues.append(tkinter.StringVar())
        box = ttk.Combobox(frame, textvariable=self.entryValues[-1], width=15)
        box["values"] = ("5", "15", "30", "60")
        box.grid(row=5, column=1,columnspan = 3)
        frame.grid(row=1, column=0,pady=10)

        ###
        ### set up lower left frame
        ###


        frame = tkinter.Frame(outerFrame, width=width,height =50, bg="white", relief=tkinter.GROOVE, borderwidth=2)
        frame.grid_propagate(False)
        vcmd = (self.register(self.validate_is_numeric_only), "%d", "%s", "%S")
        tkinter.Label(frame, text="# Cameras", bg="white",font = f).grid(row=0, column=0, pady=10, padx=10, sticky="e")
        self.entryValues.append(tkinter.StringVar())
        tkinter.Entry(frame, width=7, textvariable=self.entryValues[-1], bg="white",validate="key",validatecommand=vcmd).grid(row=0, column=1, pady=10,
                                                                                           padx=0, sticky="w")
        tkinter.Button(frame,text = "Update",height =1,command = self.update_movement_window,font = f).grid(row=0,column=2)
        frame.grid(row=2,column=0)
        outerFrame.grid(row=0, column=0,padx = 10,rowspan=2)

        ###
        ### set up classification frame
        ###
        classificationFrame = tkinter.Frame(outerFrame, bg="white", width = 300,height = 480)
        frame = tkinter.Frame(classificationFrame,bg = "white", relief=tkinter.GROOVE, borderwidth=2)
        #frame.grid_propagate(False)
        tkinter.Label(frame, text="Classification",bg = "white",font = f).grid(row=0, column=0,columnspan = 3)
        tkinter.Label(frame, text="Overview",bg = "white",font = f).grid(row=1, column=0, pady=10, padx=10)
        tkinter.Label(frame, text="ANPR classes",bg = "white",font = f).grid(row=1, column=1, pady=10, padx=10)

        for i in range(10):
            self.entryValues.append(tkinter.StringVar())
            tkinter.Entry(frame, width=10, textvariable=self.entryValues[-1],bg = "white").grid(row=2 + i, column=0, pady=10, padx=10)
            self.entryValues.append(tkinter.StringVar())
            tkinter.Entry(frame, width=10, textvariable=self.entryValues[-1],bg = "white").grid(row=2+i, column=1, pady=10, padx=10)
        frame.grid(row=0,column =1)


        ###
        ### frame with buttons
        ###

        frame = tkinter.Frame(outerFrame, bg="white", width=300, height=70)
        tkinter.Button(frame, text="Back", bg="white",command=self.spawn_survey_setup_screen).grid(row=0, column=0, padx=10,sticky="w")
        tkinter.Button(frame, text="Import", bg="white",command=self.import_movement_details_from_excel).grid(row=0, column=1, padx=10,sticky="w")
        tkinter.Button(frame,text="Save",bg = "white",command=self.save_job).grid(row=0,column=2,padx = 10,sticky = "e")

        frame.grid(row=3,column=1,padx = 10,pady=10)
        classificationFrame.grid(row=0, column=1,rowspan  =3)

        print("width of outer frame ( on left) is",outerFrame.winfo_reqwidth())
        ###
        ### movements frame
        ###
        width = self.winfo_screenwidth()- outerFrame.winfo_width()
        height = self.winfo_screenheight()
        movementsFrame = tkinter.Frame(win, bg="white", width=width, height=height-100 )
        #movementsFrame.grid_propagate(False)
        labelFrame = tkinter.Frame(movementsFrame, bg="white")
        tkinter.Label(labelFrame, text="On Site Movements", bg="white").grid(row=0, column=0,columnspan=3)
        tkinter.Label(labelFrame, text="ANPR Movements", bg="white").grid(row=0, column=3,columnspan=3,padx = (200,0))
        tkinter.Label(labelFrame, text="Site", bg="white").grid(row=1, column=0,padx = (10,0))
        tkinter.Label(labelFrame, text="Cam", bg="white").grid(row=1, column=1,padx = (60,0))
        tkinter.Label(labelFrame, text="Mvmt", bg="white").grid(row=1, column=2,padx = (60,0))
        tkinter.Label(labelFrame, text="Movement", bg="white").grid(row=1, column=3,padx = (150,90))
        tkinter.Label(labelFrame, text="Dir", bg="white").grid(row=1, column=4,padx = 0)
        tkinter.Label(labelFrame, text="", bg="white").grid(row=1, column=5, padx=0)
        labelFrame.grid(row=0, column=0)
        self.update()
        print("label frame, height is",labelFrame.winfo_height())
        self.movementsFrame = tkinter.Frame(movementsFrame,bg="white", width=width, height=height - 100, relief=tkinter.GROOVE, borderwidth=3)
        #self.movementsFrame.grid_propagate(False)
        self.movementsFrame.grid(row=1,column  = 0,columnspan = 6,padx =0,pady=0)
        movementsFrame.grid(row=0, column=1, pady=(10, 0),padx=10)
        self.update()

    def import_movement_details_from_excel(self):
        file = filedialog.askopenfilename()
        if file == "" or (not ".xls" in file and not ".xlsx" in file and not ".xlsm" in file):
            messagebox.showinfo(message="You need to select an excel sheet")
            return
        try:
            df = pd.read_excel(file,converters={0:int,1:str,2:int,3:int,4:str})
            df["Original Movement"] = df["Original Movement"].fillna(0)
            df["Direction"] = df["Direction"].fillna("0")
            df["Direction"] = df["Direction"].apply(str.lower)
            df["Direction"][df["Direction"] == "in"] = 1
            df["Direction"][df["Direction"] == "out"] = 2
            df["Direction"][df["Direction"] == "both"] = 3
            df["Direction"] = df["Direction"].apply(int)
            df = df.fillna("")
            if df.columns.tolist() == ['Site', 'Cam', 'Original Movement', 'New Movement', 'Direction']:
                numMovements = df["Original Movement"].max()
                if numMovements & 1:
                    numMovements+=1
                self.entryValues[12].set(int(numMovements/2))
                self.update_movement_window()
                if len(self.entryValues) > 33:
                    for i in range(0, len(self.entryValues[33:]), 4):
                        for index, item in df.iterrows():
                            if int(i/4) + 1 == int(item[2]) and item[3] != "":
                                self.entryValues[33 + i].set(item[0])
                                self.entryValues[33 + i + 2].set(item[3])
                                self.entryValues[33 + i + 3].set(item[4])
            else:
                messagebox.showinfo(message="Incorrect format, expected headers of \n" + ",".join(
                    ['Site', 'Cam', 'Original Movement', 'New Movement', 'Direction']) + " but got \n" + ",".join(
                    df.columns.tolist()))
                return
        except ValueError as e:
            messagebox.showinfo(message="Couldnt import file, incorrect file")
            print(e)
            return

    def edit_job(self):
        if self.tree.selection() == "":
            return
        jobname = self.tree.item(self.tree.selection()[0])
        print("selected job", jobname["values"])
        self.currentJob = myDB.load_job(jobname["values"][0], jobname["values"][1], datetime.datetime.strptime(jobname["values"][2],"%d/%m/%y").date())
        self.spawn_parameters_window()
        job = self.currentJob
        self.entryValues[0].set(job["jobno"])
        self.entryValues[1].set(job["jobname"])
        self.entryValues[2].set(datetime.datetime.strftime(job["surveydate"], "%d/%m/%y"))

        t =job["timeperiod1"]
        self.entryValues[3].set(t.split("-")[0])
        self.entryValues[4].set(t.split("-")[1])
        t = job["timeperiod2"]
        self.entryValues[5].set(t.split("-")[0])
        self.entryValues[6].set(t.split("-")[1])
        t = job["timeperiod3"]
        self.entryValues[7].set(t.split("-")[0])
        self.entryValues[8].set(t.split("-")[1])
        t = job["timeperiod4"]
        self.entryValues[9].set(t.split("-")[0])
        self.entryValues[10].set(t.split("-")[1])
        self.entryValues[12].set(job["noOfCameras"])
        self.entryValues[11].set(job["interval"])
        classes = job["classification"].split(",")
        for index,e in enumerate(self.entryValues[13:33]):
            if index < len(classes):
                e.set(classes[index])

        print("no of sites is",job["sites"])
        self.update_movement_window()
        sites = job["sites"]
        print("sites are ",sites)
        print("no of entry values is",len(self.entryValues))
        if len(self.entryValues) > 33:
            print("no of entry values is", len(self.entryValues))
            count = 1
            for i in range(0, len(self.entryValues[33:]), 4):
                for siteNo,site in sites.items():
                    #print("site no",siteNo,"site",site)
                    for newmvNo,mvmt in site.items():
                        #print("newmvno",newmvNo,"oldmovements",mvmt)
                        for oldMv in mvmt["originalmovements"]:
                            if oldMv == count:
                                self.entryValues[33 + i].set(siteNo)
                                self.entryValues[33 + i + 2].set(newmvNo)
                                self.entryValues[33 + i + 3].set(mvmt["dir"])
                count += 1

    def spawn_duplicates_window(self):
        ###
        ### various settings for the ANPR project
        ###


        ###
        ### set the time value for excluding duplicates
        ###

        try:
            e = self.currentJob["duplicateValues"]
        except KeyError as e:
            messagebox.showinfo(message="No plates loaded, cant display duplicates")
            return
        for child in self.winfo_children():
            child.destroy()
        win = tkinter.Frame(self, bg="white",relief=tkinter.SUNKEN,borderwidth=2)
        win.grid(row=0, column=0,padx=(50,0))
        f = tkinter.font.Font(family="helvetica", size=10)
        tkinter.Label(win, text="Duplicate Removal", font=f, bg="Light blue").grid(row=0, column=0,columnspan=4,sticky="ew",pady=(0,5))
        tkinter.Label(win,text = "Duration",font = f,bg = "Lavender").grid(row=1,column=0)
        tkinter.Label(win, text="VRN Count",font = f,bg = "Lavender").grid(row=1, column=1)
        tkinter.Label(win, text="Duration",font = f,bg = "Lavender").grid(row=1, column=2)
        tkinter.Label(win, text="VRN Count",font = f,bg = "Lavender").grid(row=1, column=3)
        for i in range(31):
            colour = "white"
            rel = tkinter.GROOVE
            if self.currentJob["selectedduplicates"] != -1:
                if self.currentJob["selectedduplicates"] == i:
                    colour = "red"
                    rel = tkinter.RAISED
            l = tkinter.Label(win,text = datetime.timedelta(seconds = i),font = f,bg = colour,relief=rel)
            l.grid(row=i+2,column = 0)
            l.bind("<Double-Button-1>",self.select_duplicate)
            tkinter.Label(win,text = self.currentJob["duplicateValues"][i],font = f,bg = "white",relief=tkinter.GROOVE,width = 5).grid(row=i+2,column = 1,padx=(0,5))
            colour = "white"
            rel=tkinter.GROOVE
            if self.currentJob["selectedduplicates"] != -1:
                if self.currentJob["selectedduplicates"] == i+31:
                    colour = "red"
                    rel = tkinter.RAISED
            l = tkinter.Label(win, text=datetime.timedelta(seconds=i*15),font = f,bg= colour,relief=rel)
            l.grid(row=i + 2, column=2)
            l.bind("<Double-Button-1>", self.select_duplicate)
            tkinter.Label(win, text=self.currentJob["duplicateValues"][i+31],font = f,bg = "white",relief=tkinter.GROOVE,width = 5).grid(row=i+2, column=3)


        ###
        ### set the time offset for each movement, in case a cameras internal clock was slightly out
        ###
        frame = tkinter.Frame(bg="white",relief=tkinter.SUNKEN,borderwidth=2,height = 700)
        frame.grid(row=0,column=1,sticky="n",padx=(50,0))
        inMov = []
        numRows = 30
        maxRows = 25
        i = 2
        while numRows / i > maxRows: i += 1
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                if mvmt["newmovement"] not in inMov:
                    inMov.append(mvmt["newmovement"])
        inMov = sorted(inMov)
        f = tkinter.font.Font(family="helvetica", size=10)
        tkinter.Label(frame, text="Time Offset Adjustment", font=f, bg="Light blue").grid(row=0, column=0,columnspan=10,sticky="ew",pady=(0,5))
        #tkinter.Label(frame, text="", font=f, bg="white").grid(row=1, column=5, pady=(10, 10),columnspan=5)

        for count, i in enumerate(inMov):
            e = tkinter.Entry(frame, width=8)
            e.bind("<FocusOut>", self.validate_time_entry)
            td = datetime.timedelta(seconds=abs(self.currentJob["timeAdjustmentsDictionary"][i]))
            if self.currentJob["timeAdjustmentsDictionary"][i] < 0:
                s = "-" + str(td)
            else:
                s = str(td)
            e.insert(0, s)
            e.grid(row=(count % numRows)+1, column=(int(count / numRows) * 2) + 1,padx=10)
            tkinter.Label(frame, text="Mvmt " + str(i), bg="white", font=f).grid(row=(count % numRows) + 1,column=(int(count / numRows) * 2) )

        ###
        ### Restrict plates based on length of VRN
        ###

        v = tkinter.IntVar()
        v.set(self.currentJob["platerestriction"])
        frame = tkinter.Frame(bg="white", relief=tkinter.SUNKEN, borderwidth=2, height=700)
        frame.grid(row=0, column=2, sticky="n",padx=(50,0))
        tkinter.Label(frame, text="Plate Length Restriction", font=f, bg="Light blue").grid(row=0,column=0,sticky="ew")
        tkinter.Radiobutton(frame,text="All plates",variable=v,value=1, bg="white", font=f).grid(row=1,column=0,sticky="w")
        tkinter.Radiobutton(frame, text="4-7 : " + str(self.currentJob["platerestrictionpercentages"][1]) + "%", variable=v,value=2, bg="white", font=f).grid(row=2,column=0,sticky="w")
        tkinter.Radiobutton(frame, text="5-7 : " + str(self.currentJob["platerestrictionpercentages"][2])+ "%", variable=v,value=3, bg="white", font=f).grid(row=3,column=0,sticky="w")
        tkinter.Radiobutton(frame, text="4-8 : " + str(self.currentJob["platerestrictionpercentages"][3])+ "%", variable=v,value=4, bg="white", font=f).grid(row=4,column=0,sticky="w")

        ###
        ### some buttons
        ###

        frame = tkinter.Frame(bg="white")
        frame.grid(row=1, column=0,pady=20)
        tkinter.Button(frame, text="Save", command=lambda:self.close_duplicates_window(v)).grid(row=0, column=0)
        tkinter.Button(frame, text="Back", command=self.quit_duplicates_window).grid(row=0, column=1)

    def select_duplicate(self,event):
        index = 0
        print("widget is",event.widget,event.widget.cget("text"),event.widget.grid_info())
        text = event.widget.cget("text")
        info = event.widget.grid_info()
        if info["column"] == 2:
            index = 31
        index = index + info["row"] -2
        print("selected index is",index)
        parent=self.nametowidget(event.widget.winfo_parent())
        children = parent.winfo_children()
        for child in children[5:]:
            self.nametowidget(child).configure(bg="white",relief=tkinter.GROOVE)
        event.widget.configure(bg="red",relief=tkinter.RAISED)
        self.currentJob["selectedduplicates"] = index
        #self.spawn_duplicates_window()

    def validate_comparison_entry(self,action,text,char):
        if action == "0":  ### action 0 is delete. We dont mind what they delete
            return True
        return char.isdigit()

    def validate_date(self,action,text,char):
        if action == "0":
            return True
        print("text is",text,"char is ",char)
        if len(text)== 8:
            return False
        if len(text) == 2 or len(text) == 5:
            return(char =="/")
        else:
            return char.isdigit()
        return False

    def validate_date_on_focus_out(self,event):
        ###
        ### validate that a correct date , in a correct format, has been entered
        text = event.widget.get()
        try:
            d = datetime.datetime.strptime(text,"%d/%m/%y")
        except Exception as e:
            messagebox.showinfo(message=text + " isnt a valid date, must be in format dd/mm/yy")
            event.widget.delete(0,tkinter.END)

    def validate_is_numeric_only(self,action,text,char):
        if action == "0":  ### action 0 is delete. We dont mind what they delete
            return True
        return char.isdigit()

    def validate_time_cell_input(self,action,text,char):
        ###
        ### validate that only numbers, or a colon, or a correct time can be entered in the cells for the project times on the project setup screen
        ###

        pattern  = re.compile("[0-9][0-9]:[0-9][0-9]")
        if len(text) == 0 and pattern.match(char):
            return True
        if action == "0": ### action 0 is delete. We dont mind what they delete
            return True
        if len(text) ==5:### dont allow any string greater than length 5
            return False
        if len(text) == 2: ### if 2 numbers have been entered, the next char must be a :
            return (char == ":" or char == ":00")
        if len(text)==1:
            if char == ":":
                pass
        return(char.isdigit() or char == ":00")

    def validate_hhmm(self,event):
        ###
        ### validate the user input for the project times on the project setup screen, when the input widget loses focus.
        ###

        text = event.widget.get()
        if text.strip() == "":
            return
        text = self.validate_text_to_hhmm(text)
        print("text is - ",text)
        event.widget.delete(0,tkinter.END)
        event.widget.insert(0,text)
        return "break"

    def save_job(self):
        ###
        ### save the job details entered in the form, to the main job database
        ### will also save an edited job
        ### checking of data is done here before saving
        ###

        ###
        ### set up a dictionary containing all the details entered on the form
        job = {}

        job["jobno"] = self.entryValues[0].get()
        job["jobname"]=self.entryValues[1].get()
        try:
            job["surveydate"]= datetime.datetime.strptime(self.entryValues[2].get() ,"%d/%m/%y").date()
        except Exception as e:
            messagebox.showinfo(message="Incorrect date format, project not saved")
            return

        for i in range(3,11,2):
            try:
                d = datetime.datetime.strptime(self.entryValues[i].get(),"%H:%M")
                d1 = datetime.datetime.strptime(self.entryValues[i+1].get(), "%H:%M")
                if d >= d1:
                    messagebox.showinfo(message="You have entered a pair of times where\n the start time is equal to or after the end time.\n no data saved")
                    return False
                ### TODO: verify end time is after start time, verify that if one is filled, the other is filled
            except Exception as e:
                self.entryValues[i].set("")
                self.entryValues[i+1].set("")

        if self.entryValues[3].get() == "" or self.entryValues[4].get() == "":
            messagebox.showinfo(message="You must enter at least one time period")
            return

        job["timeperiod1"] = self.entryValues[3].get() + "-" + self.entryValues[4].get()
        job["timeperiod2"] = self.entryValues[5].get() + "-" + self.entryValues[6].get()
        job["timeperiod3"] = self.entryValues[7].get() + "-" + self.entryValues[8].get()
        job["timeperiod4"] = self.entryValues[9].get() + "-" + self.entryValues[10].get()
        job["noOfCameras"] =self.entryValues[12].get()
        job["interval"] = self.entryValues[11].get()
        cl = [str(x.get()) for x in self.entryValues[13:33] if x.get() != ""]
        print("classification is",cl)
        job["classification"] = ",".join(cl)

        ###
        ### do some basic checking to make sure that a value has been entered for each key in the dictionary
        ### prompt the user if they havent
        ###
        for key,value in job.items():
            print(key,value)
            if value =="":
                messagebox.showinfo(message="You must enter " + key)
                return

        ###
        ### set up the site data
        ###
        data = {}
        data["job"] = job
        data["sites"] = []
        if len(self.entryValues) >33: ### the site details start at entryValues[33]
            count = 1
            for i in range(0,len(self.entryValues[33:]),4):
                if self.entryValues[33 +i+2].get() != "":
                    if self.entryValues[33 + i+3].get()==0:
                        messagebox.showinfo(message="For site " + self.entryValues[33+i].get() + " movement " + str(count) + " you havent selected a direction")
                        return
                    l =[self.entryValues[33+i].get(),self.entryValues[33 +i+2].get(),count,self.entryValues[33 + i+3].get(),self.entryValues[33 +i+1].get()] # [Site no, new movement no, old movement no, dir,cam]
                    data["sites"].append(l)
                count+=1



        ###
        ### prompt the user to select the location for storing any files and data produced by the software
        ###


        result = myDB.save_Job(data,self.user)
        print("in win, job folder is",result["folder"])
        if result == False:
            print("failed to save to db")
            return
        job = myDB.load_job(job["jobno"],job["jobname"],job["surveydate"])
        print("in win 2nd time, job folder is", result["folder"])
        job["timeAdjustmentsDictionary"] = {}
        self.updateDataFunction(job)
        self.spawn_survey_setup_screen()

    def update_movement_window(self):
        ###
        ### when setting up a project, we need to enter the relations between site numbers, old movement numbers,
        ### new movement numbers, etc. The number of movements is 2 * the number of cameras in the project
        ### this function fills out the window with the required number of widgets
        ###

        ###
        ### delete any previous stored entryValues
        ###

        print("no of existing entryvalues is",len(self.entryValues))
        if self.numCams >0 and len(self.entryValues)>33:
            for i in range(self.numCams  * 8): ## 4 variables, 2 rows, for each cam
                del self.entryValues[-1]

        for child in self.movementsFrame.winfo_children():
            child.destroy()
        self.numCams = int(self.entryValues[12].get())
        count = int(self.entryValues[12].get()) * 2

        ###
        ### set up the label frame
        ###

        self.scrollFrame = VerticalScrolledFrame(self.movementsFrame,bg="beige")
        vcmd = (self.register(self.validate_is_numeric_only), "%d", "%s", "%S")
        for i in range(count):
            self.entryValues.append(tkinter.StringVar())
            tkinter.Entry(self.scrollFrame.interior,textvariable=self.entryValues[-1],width = 5,validate="key",validatecommand=vcmd).grid(row=i,column = 0, padx=(25,0))
            self.entryValues.append(tkinter.StringVar())
            tkinter.Entry(self.scrollFrame.interior, textvariable=self.entryValues[-1],width = 8).grid(row=i, column=1, padx=(60,0))
            self.entryValues.append(tkinter.StringVar())
            tkinter.Entry(self.scrollFrame.interior, textvariable=self.entryValues[-1], width=5,validate="key",validatecommand=vcmd).grid(row=i, column=3,padx=(150,10))
            tkinter.Label(self.scrollFrame.interior, text=str(i + 1),bg="white").grid(row=i, column=2, padx=(70,10))
            self.entryValues.append(tkinter.IntVar())
            tkinter.Radiobutton(self.scrollFrame.interior,text = "In",variable=self.entryValues[-1],value=1,bg="white").grid(row=i,column = 4,padx=(50,0))
            tkinter.Radiobutton(self.scrollFrame.interior, text="Out", variable=self.entryValues[-1],value=2,bg="white").grid(row=i, column=5)
            tkinter.Radiobutton(self.scrollFrame.interior, text="both", variable=self.entryValues[-1],value=3,bg="white").grid(row=i, column=6,padx =(0,30))

            self.scrollFrame.grid(row=1,column = 0,padx = 0,pady=0)
        self.update()
        print("size of scrollframe is", self.scrollFrame.winfo_height())

    def spawn_summary_window(self):
        ###
        ### the summary window displays a summary of all sites in the project, in a pop up window.
        ###

        win = tkinter.Toplevel(self)
        win.protocol("WM_DELETE_WINDOW",lambda: self.summary_window_closed(win))
        f = tkinter.font.Font(family="helvetica",size=18)
        win.state("zoomed")
        frame = tkinter.Frame(win)
        frame.grid(row=0,column=0)
        tkinter.Label(frame,text = "Summary",font=f).grid(row=0,column=0)
        cols = ("Movement", "Site", "OVCount", "VRN Count", "Av % Capture", "Min % Capture", "Max % Capture",
                "Time < 85%","Comments")
        self.summaryTree = ttk.Treeview(frame, columns=tuple(range(len(cols))), show="headings", height=40)
        for i, col in enumerate(cols):
            self.summaryTree.heading(i, text=col)
            self.summaryTree.column(i, width=20, anchor=tkinter.CENTER,stretch=tkinter.NO)
        self.summaryTree.column(i,width=500)
        self.summaryTree.grid(row=1,column=0,padx = 100,pady=30)
        self.summaryTree.bind("<Double-Button-1>",self.comment_clicked)
        self.summaryTree.bind("<Button-3>", self.movement_selected_via_summary)
        self.update_summary_screen()

    def movement_selected_via_summary(self,event):
        curItem = event.widget.identify_row(event.y)
        movement = self.summaryTree.item(curItem)["values"][0]
        movementList = self.movementBox.cget("values")
        self.movementBox.current(movementList.index("Movement " + str(movement)))
        self.movementBox.event_generate("<<ComboboxSelected>>", when="tail")

    def summary_window_closed(self,win):
        self.summaryTree = None
        win.destroy()

    def spawn_home_window(self):
        ###
        ### this window shows once the user has loaded a job. It allows the user to do various
        ### tasks related to the ANPR project
        ###

        if self.oldJobData is not None:
            self.currentJob = copy.deepcopy(self.oldJobData)
            self.oldJobData = None

        for child in self.winfo_children():
            child.destroy()
        self.colourLabels = []
        self.summaryTree = None


        f = tkinter.font.nametofont("TkDefaultFont").configure(size=12)
        offset = self.winfo_screenwidth() - 820  ### the amount to offset the mainframe in the window so its centred
        offset /= 2
        if self.winfo_screenheight() >900:
            width = 1000
            height = 800
            fontsize = 14
            f = tkinter.font.nametofont("TkDefaultFont").configure(size=14)
        else:
            width = 800
            height = 600
            fontsize=9
            f = tkinter.font.nametofont("TkDefaultFont").configure(size=9)
        mainframe = tkinter.Frame(self, width=820,bg="white")
        mainframe.grid(row=0, column=0, pady=(100, 0), padx=(offset, 0))

        frame = tkinter.Frame(mainframe, bg="light grey", width=820,relief=tkinter.GROOVE,borderwidth=1)
        f = tkinter.font.Font(family='Helvetica', size=fontsize, weight=tkinter.font.BOLD)
        d = datetime.datetime.strftime(self.currentJob["surveydate"],"%d/%m/%y")
        tkinter.Label(frame, text=self.currentJob["jobno"] + " " + self.currentJob["jobname"]+ " " + d , bg="light grey",fg=self.tracsisBlue,font=f).grid(row=0, column=0,padx=(220,10),ipady=10)

        f = tkinter.font.Font(family='Arial', size=fontsize)
        try:
            self.img = ImageTk.PhotoImage(Image.open("folder-icon.jpg").resize((30,30),Image.ANTIALIAS))
            tkinter.Button(frame, image=self.img, command=self.open_project_folder).grid(row=0, column=1)
        except Exception as e:
            tkinter.Button(frame,text = "Open\nFolder", command=self.open_project_folder,font=f).grid(row=0, column=1)

        frame.grid(row=0, column=0,sticky="ew")
        f = tkinter.font.Font(family='Arial', size=fontsize)
        frame = tkinter.Frame(mainframe, bg="white",relief=tkinter.GROOVE,borderwidth=1,width=820)
        tkinter.Label(frame,text = "Overviews", bg="white",fg=self.tracsisBlue).grid(row=1,column = 0,padx = (200,0),columnspan = 2,sticky="ew")
        tkinter.Button(frame, text="Create Overview \nCount Template", width=17, bg="white",fg=self.tracsisBlue, height=3,
                       command=self.export_OVTemplate).grid(row=2, column=0, padx=(220,20), pady=10)
        tkinter.Button(frame, text="Load Overview \nCount Results", width=17, height=3, bg="white",fg=self.tracsisBlue,command=self.load_OV_counts).grid(row=2, column=1, padx=20,pady=10)
        frame.grid(row=1,column=0,pady=5,sticky="ew")

        frame = tkinter.Frame(mainframe, bg="white",relief=tkinter.GROOVE,borderwidth=1,width=820)
        tkinter.Label(frame, text="VRNs", bg="white",fg=self.tracsisBlue).grid(row=3, column=0,ipadx =30,columnspan = 10)
        tkinter.Button(frame, text="Load Unclassed\n VRNs", width=17, height=3, bg="white",fg=self.tracsisBlue,command = self.load_unclassed_plates).grid(row=4, column=0,padx=20,pady=10)
        tkinter.Button(frame, text="Load Classed\n VRNs", width=17, height=3, bg="white",fg=self.tracsisBlue,command=self.load_classes).grid(row=4, column=1,
                                                                                                 padx=20, pady=10)
        tkinter.Button(frame, text="Duplicate Removal/ \nTime Adjustments", width=17, height=3, bg="white",fg=self.tracsisBlue,command=self.spawn_duplicates_window).grid(row=4, column=2,padx=20,
                                                                                                        pady=10)
        tkinter.Button(frame, text="Duration \n Limiter", width=17, height=3, bg="white",fg=self.tracsisBlue,command=self.spawn_duration_matrix_screen).grid(row=4, column=3, padx=20, pady=10, sticky="e")
        frame.grid(row=2, column=0,pady=5,sticky="ew")
        self.update()
        print("wibble",frame.winfo_width(),self.winfo_reqwidth())

        frame = tkinter.Frame(mainframe, bg="white",relief=tkinter.GROOVE,borderwidth=1,width=820)
        tkinter.Label(frame, text="Comparison", bg="white").grid(row=5, column=0,padx = (200,0),columnspan = 10)
        tkinter.Button(frame, text="View Comparison", width=17, height=3, bg="white",fg=self.tracsisBlue,command= self.get_comparison_data).grid(row=6, column=0, padx=(220,20), pady=10)
        tkinter.Button(frame, text="Create Client\nComparison", width=17, height=3, bg="white",fg=self.tracsisBlue).grid(row=6, column=1,padx=20, pady=10)
        frame.grid(row=3, column=0,pady=5,sticky="ew")

        frame = tkinter.Frame(mainframe, bg="white",relief=tkinter.GROOVE,borderwidth=1,width=820)
        tkinter.Label(frame, text="Matching", bg="white").grid(row=7, column=0,ipadx =30,columnspan = 10)
        tkinter.Button(frame, text="Open/Closed\nCordon", width=17, height=3, bg="white",fg=self.tracsisBlue,command=lambda:self.spawn_matching_results_screen("Directional")).grid(row=8, column=0,padx=20, pady=10)
        tkinter.Button(frame, text="Route\nAssignment", width=17, height=3, bg="white",fg=self.tracsisBlue,command=lambda:self.spawn_matching_results_screen("First Seen/Last Seen")).grid(row=8, column=1,padx=20, pady=10)
        tkinter.Button(frame, text="Overtaking/\nPlatooning", width=17, height=3, bg="white",fg=self.tracsisBlue,command = self.spawn_overtaking_setup_screen).grid(row=8, column=3, padx=20,pady=10)
        tkinter.Button(frame, text="Filtered\nMatching", width=17, height=3, bg="white",fg=self.tracsisBlue,command=lambda:self.spawn_matching_results_screen("Filtered")).grid(row=8, column=2, padx=20,pady=10)
        frame.grid(row=4, column=0,pady=5,sticky="ew")
        self.update()
        print("size of mainframe is",mainframe.winfo_height(),mainframe.winfo_reqheight())
        offsetForBackButton = (self.winfo_height()- (mainframe.winfo_reqheight()))/2
        offsetForBackButton-=60
        print("offset is",offsetForBackButton)
        #frame = tkinter.Frame(self, bg="white", width=820)
        tkinter.Button(self, text="Back", width=12, height=1, bg="white",fg=self.tracsisBlue,command = self.spawn_survey_setup_screen).grid(row=5, column=0, padx=60,
                                                                                      pady=offsetForBackButton,sticky="sw")

    def quit_duplicates_window(self):
        self.currentJob["selectedduplicates"] = myDB.get_value_of_field(self.currentJob["id"],"selectedDuplicates")
        print("setting value of duplicates to ",self.currentJob["selectedduplicates"])
        self.spawn_home_window()

    def close_duplicates_window(self,var):
        with open(self.currentJob["folder"] + "/data/timeAdjustments.pkl", "wb") as f:
            pickle.dump(self.currentJob["timeAdjustmentsDictionary"], f)
        myDB.update_value_of_field(self.currentJob["id"],"plateRestriction",var.get())
        if self.currentJob["selectedduplicates"] != myDB.get_value_of_field(self.currentJob["id"],"selectedDuplicates"):
            self.setDuplicatesFunction(self.currentJob["selectedduplicates"], self.currentJob)
            myDB.update_duplicates(self.currentJob["id"], self.currentJob["selectedduplicates"])
        self.loadJobFunction(self.currentJob)
        self.spawn_home_window()
        var=None

    def validate_time_entry(self,event):
        ###
        ### this function validates a value entered on the time adjustments screen, and puts it into the correct
        ### movement in the data dictionary.
        ###

        w = event.widget
        info = w.grid_info()
        row = info["row"]
        col = info["column"]
        print("row, col",row,col)
        frame = self.nametowidget(self.winfo_children()[1])
        for label in frame.children.values():
            print(label.cget("text"))
            info = label.grid_info()
            print("info is",info)
            if info["row"] == row  and info["column"] == col - 1:
                movement = label.cget("text")
                print("movement is",movement)
                movement = int(movement.replace("Mvmt ",""))
        prevValue = self.currentJob["timeAdjustmentsDictionary"][movement]
        text = w.get()
        negative = False
        pattern = re.compile("-{0,1}[0-9]:[0-9][0-9]:[0-9][0-9]")
        if pattern.match(text):
            if text[0]=="-":
                negative  = True
                text = text[1:]
            t = text.split(":")
            h = int(t[0])
            m = int(t[1])
            s = int(t[2])
            td = datetime.timedelta(hours=h,minutes=m,seconds=s)
            if negative:
                self.currentJob["timeAdjustmentsDictionary"][movement] = 0 - td.total_seconds()
            else:
                self.currentJob["timeAdjustmentsDictionary"][movement] = td.total_seconds()
        else:
            w.delete(0,tkinter.END)
            w.insert(0,datetime.timedelta(seconds=prevValue))

    def open_project_folder(self):
        if os.path.isdir(self.currentJob["folder"]):
            p = os.path.normpath(self.currentJob["folder"])
            subprocess.Popen('explorer "{0}"'.format(p))
        else:
            messagebox.showinfo(message="Project folder doesnt exist")

    def comment_clicked(self,event):
        curItem = event.widget.identify_row(event.y)
        print(event.widget.identify_column(event.y))
        values = self.summaryTree.item(curItem)["values"]
        self.spawn_comment_window(values[-1])

    def spawn_comment_window(self,text):
        win = tkinter.Toplevel(self)
        win.protocol("WM_DELETE_WINDOW", lambda:self.destroy__window(win))
        txt = tkinter.Text(win)
        txt.grid(row=0,column=0)
        txt.focus_set()
        txt.insert("1.0",text)
        tkinter.Button(win,text="Save",command=lambda: self.save_comment(txt,win)).grid(row=1,column=0)

    def spawn_settings_window(self):
        f = tkinter.font.Font(family='Helvetica', size=8)
        win = tkinter.Toplevel(self,width = 200,height = 200,bg="white")
        win.wm_attributes("-topmost", 1)
        win.protocol("WM_DELETE_WINDOW", lambda: self.destroy__window(win))
        tkinter.Label(win,text = "Please Enter Your Name",font = f,bg="white").grid(row=0,column = 0,padx=10,pady=10)
        e = tkinter.Entry(win,width = 16,font = f)
        e.grid( row=0,column=1,padx=(0,10))
        tkinter.Label(win,text="Please Select Database file",font = f,bg="white").grid(row=1, column=0,padx=10,pady=10)
        txt = "None"
        l = tkinter.Label(win,text="",font = f,width = 20,bg="white",justify=tkinter.LEFT)

        self.user, file = self.load_settings()
        e.delete(0, tkinter.END)
        e.insert(0, self.user)
        if file != "":
            txt = self.Dbfile.split("/")[-1]

        l.bind("<Button-3>", self.open_database_file_location)
        l.bind("<Button-1>",self.get_database_file_location)
        l.grid(row=1, column=1)
        l.configure(text=txt)
        tkinter.Button(win, text="Create", font=f, command=lambda: self.create_database(e,l,win)).grid(row=2, column=0,padx=10,pady=10)
        tkinter.Button(win, text="Save", font=f,command=lambda:self.save_settings(e,l,win)).grid(row=2, column=1,padx=10,pady=10)

    def get_database_file_location(self,event):
        ###
        ### prompt the user with a file navigation dialog, to select the location of the job database
        ### display the selected location in a label in the settings window
        ###

        file = filedialog.askopenfilename()
        if file == "" or ".sqlite" not in file:
            messagebox.showinfo(message="You need to select a database file")
            event.widget.configure(text="None")
            return
        self.Dbfile = file
        myDB.set_file(file)
        event.widget.configure(text=file.split("/")[-1])


    def open_database_file_location(self,event):
        dirName = os.path.dirname(self.Dbfile)
        if os.path.isdir(dirName):
            p = os.path.normpath(dirName)
            subprocess.Popen('explorer "{0}"'.format(p))
        else:
            messagebox.showinfo(message="Database location doesnt exist")

    def create_database(self,e,l,win):
        fileName =myDB.create_Db()
        print(l.cget("text"))
        result = messagebox.askyesno(message="Do you want to set this new database as the working database?")
        if result:
            l.configure(text=fileName)

    def save_settings(self,e,l,win):
        name = e.get()
        print("name is",name,"file is",self.Dbfile)
        if (name == "") | (self.Dbfile == ""):
            messagebox.showinfo(message="You need to enter a name, and select a database location")
            return
        if ".sqlite" not in self.Dbfile:
            messagebox.showinfo(message="The selected database file must be a .sqlite file")
            l.configure(text="None")
            self.Dbfile  = ""
            return
        dir = os.getcwd()
        print("dir is",dir)
        f = open("settings.txt","w")
        f.write(name + "\n")
        f.write(self.Dbfile +  "\n")
        self.user = name
        self.destroy__window(win)
        myDB.set_file(self.Dbfile)
        self.spawn_survey_setup_screen()

    def load_settings(self):
        try:
            f = open("settings.txt", "r")
        except FileNotFoundError as e:
            f = open("settings.txt", "w")
            f.close()
            f = open("settings.txt", "r")
        try:
            name = f.readline().rstrip()
            file = f.readline().rstrip()
            val = f.readline().rstrip()
        except Exception as e:
            print(e)
            return ["",""]
        myDB.set_file(file)
        return([name,file])


    def save_comment(self,txt,win):
        ###
        ### get the comment entered by the user in the pop up text box, save it to the Db
        ### then destroy the pop up window
        ###
        contents = txt.get("1.0",tkinter.END).strip().rstrip("\n").replace("\n"," ")
        row = self.summaryTree.item(self.summaryTree.selection())
        values = row["values"]
        siteNo = values[1]
        move  = values[0]
        myDB.update_comment(self.currentJob["id"], siteNo,move,contents)
        values[-1]=contents
        index = self.summaryTree.index(self.summaryTree.selection())
        self.currentJob["comments"][index] = contents
        self.summaryTree.item(self.summaryTree.selection(),values=values)
        win.destroy()

    def destroy__window(self,win):
        win.destroy()

    def set_up_comparison_display(self):
        self.comparisonDataStructure = [] ### stores the widgets which display the data

        ###
        ### set up a list containing lists of vehicle classes, so we can set up the correct number of columns
        ### in each section
        ###

        classes = []
        seen = set()
        OVClasses = [x for i, x in enumerate(self.currentJob["classification"].split(",")) if
                     i % 2 == 0 and x not in seen and not seen.add(x)]
        classes.append(OVClasses)
        seen = set()
        ANPRClasses = [x for i, x in enumerate(self.currentJob["classification"].split(",")) if
                       i % 2 == 1 and x not in seen and not seen.add(x)]
        classes.append(ANPRClasses)
        classes.append(ANPRClasses)

        ###
        ### destroy any widgets in the window
        ###
        for child in self.winfo_children():
            if type(self.nametowidget(child))!= tkinter.Toplevel:
                child.destroy()


        win = tkinter.Frame(self, width=self.winfo_screenwidth(), height=self.winfo_screenheight(), bg="white", padx=20,
                            pady=20)
        win.columnconfigure(0,weight=1)
        f = tkinter.font.Font(family='Helvetica', size=10, weight='bold')
        appHighlightFont = font.Font(family='Helvetica', size=8)
        boldFont = font.Font(family='Helvetica', size=8, weight="bold")
        timeFont = tkinter.font.Font(family='Courier', size=9, weight='bold')
        frame = tkinter.Frame(win,bg ="white", width=self.winfo_screenwidth(), height=self.winfo_screenheight())
        frame.columnconfigure(0,weight = 1)
        tkinter.Label(frame,text="Comparison Type", bg="white",font = f).grid(row = 0,column = 0,padx=(20,10),pady=(10,10))
        box = ttk.Combobox(frame, width=20,font = f)
        box["values"] = ("Unclassed","Classed")
        box.bind("<<ComboboxSelected>>", self.boxChanged)
        box.current(self.box1Value)
        tkinter.Button(frame,text="Summary",command=self.spawn_summary_window,font = f).grid(row = 0,column = 2,padx = 20)
        self.revertButton = tkinter.Button(frame, text="Revert", font=f)
        self.revertButton.grid(row=0, column=3, padx=20,sticky="e")
        self.revertButton.bind("<Button-1>",self.revert)
        tkinter.Button(frame, text="Back", command=self.spawn_home_window, font=f).grid(row=0, column=4, padx=20)
        box.grid(row=0, column=1)

        tkinter.Label(frame, text="VRNs", bg="white",width = 14,anchor=tkinter.E,font = f).grid(row=1, column=0,padx=(20,10),pady=(10,10))
        box = ttk.Combobox(frame, width=20,font = f)
        box["values"] = ("Original VRNs", "Duplicates Removed")
        box.bind("<<ComboboxSelected>>", self.boxChanged)
        box.current(self.box2Value)
        box.grid(row=1, column=1)
        self.siteLabel = tkinter.Label(frame, text="Site 1", bg="white", width=14, anchor=tkinter.E, font=f)
        self.siteLabel.grid(row=1, column=2,padx=(50, 10),pady=(10, 10))


        tkinter.Label(frame, text="Select Mvmt", bg="white", width=14, anchor=tkinter.E, font=f).grid(row=1, column=4,
                                                                                                      padx=(20, 10),
                                                                                                      pady=(10, 10))
        ###
        ### get list of movements
        ###
        mvmts = []
        for site in self.dataList:
            [mvmts.append("Movement " + str(key)) for key, movement in sorted(site["movements"].items())]
        self.movementBox = ttk.Combobox(frame, width=15, font=f)
        self.movementBox["values"] = mvmts
        self.movementBox.grid(row=1, column=5)
        self.movementBox.current(0)
        self.movementBox.bind("<<ComboboxSelected>>", self.movementChanged)
        self.currentSelected[1] = int(mvmts[0].replace("Movement",""))
        tkinter.Button(frame,text="<",command=lambda :self.scroll_through_movements("left")).grid(row = 1,column = 7,padx = 10, pady=10)
        tkinter.Button(frame, text=">", command=lambda :self.scroll_through_movements("right")).grid(row=1, column=8, padx=10, pady=10)
        frame.grid(row=0,column=0,sticky = "w")
        win.grid(row=0, column=0)

        vcmd = (self.register(self.validate_edit), "%d", "%s", "%S")
        site = self.dataList[0]
        l = [movement for key, movement in sorted(site["movements"].items())]
        movement = l[0]
        displayedData = []
        displayedData.append(movement["data"][1])
        displayedData.append(movement["data"][2])
        displayedData.append(movement["data"][2]) ### we dont actually display this data, this is just used to build the structure of the 3rd section
        innerFrame = tkinter.Frame(win)
        scrollframe = VerticalScrolledFrame(innerFrame, bg="white")
        scrollframe.grid(row=0, column=0)
        frame = tkinter.Frame(scrollframe.interior, bg="beige", width=820)
        frame.grid(row=0, column=0)
        innerFrame.grid(row=1,column=0,padx =450,pady = (30,0))
        col = 0
        tkinter.Label(frame, text="OV Count", font=appHighlightFont, bg="beige").grid(row=1, column=col,
                                                                                      columnspan=len(
                                                                                          classes[0]) + 4,
                                                                                      padx=2, pady=20)

        tkinter.Label(frame, text="ANPR Count", font=appHighlightFont, bg="beige").grid(row=1,
                                                                                        column=col + len(
                                                                                            classes[0]) + len(
                                                                                            classes[1]),
                                                                                        columnspan=len(
                                                                                            classes[1]) + 4,
                                                                                        padx=2, pady=20)
        tkinter.Label(frame, text="Comparison", font=appHighlightFont, bg="beige").grid(row=1,
                                                                                        column=col + (2 * len(
                                                                                            classes[1])) + len(
                                                                                            classes[0]),
                                                                                        columnspan=len(
                                                                                            classes[1]) + 8,
                                                                                        padx=2, pady=20)

        if col != 0:
            ttk.Separator(frame, orient="vertical", style="sep.TSeparator").grid(row=0, column=col,
                                                                                 rowspan=1000,
                                                                                 padx=60, pady=4, sticky="ns")
            col += 1

        for i, d in enumerate(displayedData):
            print(i,"th block of data is",d)
            block = []
            first = True
            rowNo = 4
            for index, cl in enumerate(classes[i]):
                tkinter.Label(frame, text=cl, font=appHighlightFont, bg="beige").grid(row=2,
                                                                                      column=col + index + 2,
                                                                                      padx=2, pady=2)
            ttk.Separator(frame, orient="horizontal").grid(row=3, column=col, pady=4, columnspan=10,
                                                           sticky="ew")
            ttk.Separator(frame, orient="vertical").grid(row=3, column=col + 1, rowspan=1000, pady=4, padx=0,
                                                         sticky="ns")
            ttk.Separator(frame, orient="vertical").grid(row=3, column=col + 2 + len(classes[i]), rowspan=1000,
                                                         padx=0, pady=4, sticky="ns")
            tkinter.Label(frame, text="Total", font=appHighlightFont, bg="beige").grid(row=2,
                                                                                       column=col + 3 + len(
                                                                                           classes[i])
                                                                                       , padx=(2, 20), pady=2)

            for row in d:
                entryValueRow = []
                if row[0] == "1 Hr" :
                    ttk.Separator(frame, orient="horizontal").grid(row=rowNo, column=col, pady=4, columnspan=10,
                                                                   sticky="ew")
                    ttk.Separator(frame, orient="horizontal").grid(row=rowNo + 2, column=col, pady=4,
                                                                   columnspan=10, sticky="ew")
                    rowNo += 1
                    for j, item in enumerate(row):
                        if j == 0:
                            s = tkinter.StringVar()
                            lbl = tkinter.Label(frame, font=timeFont, fg="black",textvariable = s,
                                                bg="lemon chiffon")
                            lbl.grid(row=rowNo, column=col, padx=2, pady=2)
                            entryValueRow.append(s)
                            s.set(item)
                        elif j == len(row) - 1:
                            s = tkinter.StringVar()
                            lbl = tkinter.Label(frame, font=boldFont,textvariable = s,
                                                bg="beige")
                            lbl.grid(row=rowNo, column=col + j + 2, padx=(2, 20), pady=2)
                            entryValueRow.append(s)
                            s.set(int(item))
                            if i == 2:
                                self.colourLabels.append(lbl)
                        else:
                            s = tkinter.StringVar()
                            lbl = tkinter.Label(frame, font=boldFont,textvariable = s,
                                                bg="beige")
                            lbl.grid(row=rowNo, column=col + j + 1, padx=2, pady=2)
                            entryValueRow.append(s)
                            s.set(int(item))
                            if i == 2:
                                self.colourLabels.append(lbl)
                    rowNo += 2
                else:
                    for j, item in enumerate(row):
                        if j == 0:
                            s = tkinter.StringVar()
                            lbl = tkinter.Label(frame, font=timeFont, fg="black",textvariable = s,bg="lemon chiffon")
                            lbl.grid(row=rowNo, column=col, padx=2, pady=2)
                            s.set(item)
                            entryValueRow.append(s)
                        elif j == len(row) - 1:
                            s = tkinter.StringVar()
                            lbl = tkinter.Label(frame, font=boldFont,textvariable = s,bg="beige")
                            lbl.grid(row=rowNo, column=col + j + 2, padx=(2, 20), pady=2)
                            entryValueRow.append(s)
                            s.set(int(item))
                            if i == 2:
                                self.colourLabels.append(lbl)
                        else:
                            if i == 0:
                                s = tkinter.StringVar()
                                e = tkinter.Entry(frame, width=4, font=appHighlightFont, textvariable = s,
                                                  bg="beige", validate="key", validatecommand=vcmd)
                                e.grid(row=rowNo, column=col + j + 1, padx=2, pady=2)
                                s.set(int(item))
                                entryValueRow.append(s)
                                e.bind("<Return>", self.edit_cell)
                                e.bind("<Tab>", self.edit_cell)
                                e.bind("<FocusOut>",self.edit_cell)
                            else:
                                s = tkinter.StringVar()
                                lbl = tkinter.Label(frame, font=appHighlightFont,textvariable = s,bg="beige")
                                lbl.grid(row=rowNo, column=col + j + 1, padx=2, pady=2)
                                if i ==2 :
                                    self.colourLabels.append(lbl)
                                entryValueRow.append(s)
                                s.set(int(item))
                    rowNo += 1
                block.append(entryValueRow)
            self.comparisonDataStructure.append(block)
            self.update()
            self.update_idletasks()
            col += 4 + len(OVClasses)
        self.update_comparison_display()

    def scroll_through_movements(self,dir):
        ###
        ### using the "arrow" buttons on the comparison screen to move one by one through the
        ### movements , either up or down the movements
        ###
        index = self.movementBox.current()
        values = self.movementBox.cget("values")
        if dir=="left":
            if index >0:
                self.movementBox.current(index-1)
                self.movementBox.event_generate("<<ComboboxSelected>>",when="tail")
        else:
            if index<len(values)-1:
                self.movementBox.current(index + 1)
                self.movementBox.event_generate("<<ComboboxSelected>>", when="tail")

    def edit_cell(self,event):
        ###
        ### deals with when an entry box is edited, it updates the data in the data structure, and saves the data
        ### and then displays the updated data in the comparison display

        entry = event.widget

        ###
        ### find which widget triggered the event
        ###

        for r,row in enumerate(self.comparisonDataStructure[0]):
            for c, item in enumerate(row):
                if entry.cget("textvariable") == item._name:
                    for site in self.dataList:
                        for key, m in site["movements"].items():
                            if self.currentSelected[1] == key:
                                print("selected site is", site["siteNo"])
                                selectedSite = site
                    l = [movement for key, movement in sorted(selectedSite["movements"].items()) if
                         key == self.currentSelected[1]]
                    movement = l[0]
                    if entry.get().isdigit():
                        if movement["data"][1][r][c] != int(entry.get()):

                            ###
                            ### we found the widget, and its value has changed, update the stored data, update the display
                            ### and dump the data to file
                            ###
                            self.displayStatus = "edited"
                            self.revertButton.configure(text="Revert")
                            movement["data"][1][r][c] = int(entry.get())
                            self.update_comparison_display()
                            with open(self.currentJob["folder"] + '/comparisondata.pkl', 'wb') as handle:
                                pickle.dump(self.dataList, handle)
                    else:
                        entry.delete(0,tkinter.END)
                        entry.insert(0,movement["data"][1][r][c])

        try:
            if event.keycode == 13:
                event.widget.tk_focusNext().focus()
        except Exception as e:
            print(e)

    def update_comparison_display(self):
        ###
        ### update whats shown on the comparison screen depending on the selection, eg classed, unclassed etc
        ### and depending on what site and movement is selected
        ###

        for site in self.dataList:
            for key, m in site["movements"].items():
                if self.currentSelected[1] == key:
                    print("selected site is", site["siteNo"])
                    selectedSite = site
        l = [movement for key, movement in sorted(selectedSite["movements"].items()) if key == self.currentSelected[1]]
        movement = l[0]
        self.calculate_display()
        dataIndex = [(0,0),(1,0),(0,1),(1,1)].index((self.box1Value, self.box2Value)) + 2

        ###
        ### set up which data we are going to display
        ### data for each movement is in the form [OVdata,Edited OVdata,ANPRuc/orig,ANPRuc/dupremoved,ANPRc/orig,ANPRc/dupremoved]
        ### dataindex gives us the index of the ANPR data
        ###
        #print("selected boxes",(self.box1Value, self.box2Value))
        #print("site",selectedSite["siteNo"],"movement",self.currentSelected[1])
        #for item in movement["data"]:
            #print(item)
        displayedData=[movement["data"][1],movement["data"][dataIndex]]
        #print("we are actually displaying index",dataIndex)
        #print(displayedData)
        for index,block in enumerate(displayedData):
            vars = self.comparisonDataStructure[index]
            for i,row in enumerate(block):
                for j,item in enumerate(row[1:]):
                    vars[i][j+1].set(int(item))
                vars[i][0].set(row[0])
        self.update_summary_screen()

    def boxChanged(self,event):
        ###
        ### keep track of the combo box selections on the comparison display sheet
        ### box1Value tracks classed or unclassed
        ### box2Value tracks original or duplicate VRNs
        ###

        box = event.widget
        current = event.widget.current()
        text = box.get()
        if text in ["Unclassed" ,"Classed"]:
            if self.box1Value==current:
                return
            self.box1Value = current
        else:
            if self.box2Value == current:
                return
            self.box2Value = current
        print(self.box1Value,self.box2Value)
        self.update_comparison_display()

    def revert(self,event):
        ###
        ### the user can edit some of the data displayed on the comparison screen. They can also revert back to the
        ### original data. If the user reverts,we keep the edited data stored temporarily, which we can switch back
        ### in if requested. If the user reverts, and then edits the original data, the old edited data is discarded
        ### and the new edit becomes the stock edited data.
        ### self.displayStatus keeps track of which state we are in, "edited" or "base". "base" refers to displaying
        ### the base data

        print("disply status is",self.displayStatus)
        for site in self.dataList:
            for key, m in site["movements"].items():
                if self.currentSelected[1] == key:
                    print("selected site is", site["siteNo"])
                    selectedSite = site
        l = [movement for key, movement in sorted(selectedSite["movements"].items()) if key == self.currentSelected[1]]
        movement = l[0]
        print("selected movement is",movement)
        if self.displayStatus == "base":
            ### we are in reverted state, and user requested to go back to edited state
            self.displayStatus = "edited"
            event.widget.configure(text = "Revert")
            movement["data"][1] =self.tempEditedDataStore
            self.tempEditedDataStore= []
        else:
            self.displayStatus = "base"
            event.widget.configure(text="Load")
            self.tempEditedDataStore = list(movement["data"][1])
            movement["data"][1] = list(movement["data"][0])
        self.update_comparison_display()
        return "break"

    def validate_edit(self,action,text,char):
        #print("action is", action, type(action))
        print("char is",char,text)
        if action == "0":
            #print("yes")
            return True
        print("checking", char)
        return char.isdigit()

    def update_summary_screen(self):
        seen = set()
        ANPRClasses = [x for i, x in enumerate(self.currentJob["classification"].split(",")) if
                       i % 2 == 1 and x not in seen and not seen.add(x)]
        seen = set()
        OVClasses = [x for i, x in enumerate(self.currentJob["classification"].split(",")) if
                     i % 2 == 0 and x not in seen and not seen.add(x)]
        ANPRtoOVdict = {}  ### this will hold a dictionary of how we combine the OV classes into the ANPR classes
        for cl in ANPRClasses:
            # print("looing for ", cl)
            ANPRtoOVdict[cl] = []
            for item in [i for i, x in enumerate(self.currentJob["classification"].split(",")) if
                         x.lower() == cl.lower() and i % 2 == 1]:
                ANPRtoOVdict[cl].append(OVClasses.index(self.currentJob["classification"].split(",")[item - 1]))
        for site in self.dataList:
            #print("site is", site)
            l = [movement for key, movement in sorted(site["movements"].items())]
            for movement in l:
                movement["summary"] = {}
                movement["summary"]["OVTotal"] = 0
                movement["summary"]["ANPRTotal"] = 0
                movement["summary"]["AvgCapture"] = 0
                movement["summary"]["MinCapture"] = 1000
                movement["summary"]["MaxCapture"] = 0
                movement["summary"]["TimeLessThan"] = 0

                #print("site no", site["siteNo"], ",currently selected movement is", movement)


                OVdata = movement["data"][1]
                dataIndex = [(0,0),(1,0),(0,1),(1,1)].index((self.box1Value, self.box2Value)) + 2
                ANPRdata = movement["data"][dataIndex]

                for i, item in enumerate(OVdata):
                    #print("processing item", item)
                    if item[0] == "1 Hr":
                        movement["summary"]["OVTotal"] = movement["summary"]["OVTotal"] + int(item[-1])
                        #print("OVTotal is", movement["summary"]["OVTotal"])

                for row in ANPRdata:
                    if row[0] == "1 Hr":
                        movement["summary"]["ANPRTotal"] = movement["summary"]["ANPRTotal"] + row[-1]

                ###
                ### set up the comparison data for display
                ###

                compData = []
                for i, row in enumerate(OVdata):
                    compRowData = []
                    OVrow = row[1:]
                    ANPRrow = ANPRdata[i][1:]
                    timestamp = row[0]
                    for index, item in enumerate(ANPRrow[:-1]):
                        cl = ANPRClasses[index]
                        total = sum([OVrow[j] for j in ANPRtoOVdict[cl]])
                        if total == 0:
                            compRowData.append(0)
                        else:
                            compRowData.append(int(item * 100 / total))
                    if OVrow[-1] == 0:
                        compRowData.append(0)
                    else:
                        value = int(ANPRrow[-1] * 100 / OVrow[-1])
                        if timestamp != "1 Hr":
                            if value < movement["summary"]["MinCapture"]:
                                movement["summary"]["MinCapture"] = value
                            if value > movement["summary"]["MaxCapture"]:
                                movement["summary"]["MaxCapture"] = value
                            if value < 85:
                                movement["summary"]["TimeLessThan"] += 1
                        compRowData.append(value)
                    compRowData.insert(0, timestamp)
                    compData.append(compRowData)

                movement["summary"]["TimeLessThan"] = datetime.timedelta(
                    seconds=movement["summary"]["TimeLessThan"] * self.currentJob["interval"] * 60)
                if movement["summary"]["OVTotal"] != 0:
                    # print("ovtotal",site["summary"]["OVTotal"])
                    movement["summary"]["AvgCapture"] = int(
                        movement["summary"]["ANPRTotal"] * 100 / movement["summary"]["OVTotal"])
        if self.summaryTree is None:
            return
        if not self.summaryTree is None:
            for child in self.summaryTree.get_children():
                self.summaryTree.delete(child)
            self.summaryTree.tag_configure("tree", font="courier 10")
            #print("data list is", self.dataList)
            count = 0
            for site in self.dataList:
                #print("site is", site["siteNo"])
                keylist = [s for s in sorted(site["movements"])]
                #print("keylist is", keylist)
                for index, key in enumerate(keylist):
                    mvt = site["movements"][key]
                    item = []
                    item.append(key)
                    item.append(site["siteNo"])
                    summary = mvt["summary"]
                    item.append(summary["OVTotal"])
                    item.append(summary["ANPRTotal"])
                    item.append(str(summary["AvgCapture"]) + "%")
                    if summary["MinCapture"] ==1000:
                        item.append("0%")
                    else:
                        item.append(str(summary["MinCapture"]) + "%")
                    item.append(str(summary["MaxCapture"]) + "%")
                    item.append(summary["TimeLessThan"])
                    if self.currentJob["comments"][count] is None:
                        item.append("")
                    else:
                        item.append(self.currentJob["comments"][count])
                    count += 1
                    self.summaryTree.insert("", "end", values=item, tags=("tree",))
            for i in range(8):
                self.summaryTree.column(i, width=110, anchor=tkinter.CENTER, stretch=tkinter.NO)
            self.summaryTree.column(8, width=700, anchor=tkinter.CENTER, stretch=tkinter.NO)

    def calculate_display(self):

        ###
        ### this function re calculates the comparison display each time an entry box is edited and then
        ###  loses focus
        ###

        seen = set()
        ANPRClasses = [x for i, x in enumerate(self.currentJob["classification"].split(",")) if
                       i % 2 == 1 and x not in seen and not seen.add(x)]
        seen = set()
        OVClasses = [x for i, x in enumerate(self.currentJob["classification"].split(",")) if
                     i % 2 == 0 and x not in seen and not seen.add(x)]
        ANPRtoOVdict = {}  ### this will hold a dictionary of how we combine the OV classes into the ANPR classes
        for cl in ANPRClasses:
            # print("looing for ", cl)
            ANPRtoOVdict[cl] = []
            for item in [i for i, x in enumerate(self.currentJob["classification"].split(",")) if x.lower() == cl.lower() and i % 2 == 1]:
                ANPRtoOVdict[cl].append(OVClasses.index(self.currentJob["classification"].split(",")[item - 1]))

        rowList = []  ### holds the blocks of data that we want to sum by column


        for site in self.dataList:
            for key, m in site["movements"].items():
                if self.currentSelected[1] == key:
                    print("selected site is", site["siteNo"])
                    selectedSite = site
        l = [movement for key, movement in sorted(selectedSite["movements"].items()) if key == self.currentSelected[1]]
        movement = l[0]

        movement["summary"] = {}
        movement["summary"]["OVTotal"] = 0
        movement["summary"]["ANPRTotal"] = 0
        movement["summary"]["AvgCapture"] = 0
        movement["summary"]["MinCapture"] = 1000
        movement["summary"]["MaxCapture"] = 0
        movement["summary"]["TimeLessThan"] = 0

        #print("site no", site["siteNo"], ",currently selected movement is",movement)

        OVdata = list(movement["data"][1])
        dataIndex = [(0,0),(1,0),(0,1),(1,1)].index((self.box1Value,self.box2Value)) + 2
        ANPRdata = movement["data"][dataIndex]

        ###
        ### set up the OV data for display
        ###

        newList = []
        for i, item in enumerate(OVdata):
            #print("processing item", item)

            if item[0] == "1 Hr":
                rowList = [int(sum(r)) for r in zip(*rowList)]
                movement["summary"]["OVTotal"] = movement["summary"]["OVTotal"] + int(item[-1])
                rowList.insert(0, "1 Hr")
                newList.append(list(rowList))
                rowList = []
            else:
                item[-1] = (int(sum(item[1:-1])))
                rowList.append(list(item[1:]))
                newList.append(list(item))

        movement["data"][1] = list(newList)

        ###
        ### set up the ANPR data for display
        ###

        for row in ANPRdata:
            if row[0] == "1 Hr":
                movement["summary"]["ANPRTotal"] = movement["summary"]["ANPRTotal"] + row[-1]

        ###
        ### set up the comparison data for display
        ###

        compData = []
        for i, row in enumerate(OVdata):
            compRowData = []
            OVrow = row[1:]
            ANPRrow = ANPRdata[i][1:]
            timestamp = row[0]
            print("ANPRrow is",ANPRrow)
            for index, item in enumerate(ANPRrow[:-1]):
                print(index,item)
                cl = ANPRClasses[index]
                total = int(sum([OVrow[j] for j in ANPRtoOVdict[cl]]))
                if total == 0:
                    compRowData.append(int(0))
                else:
                    compRowData.append(int(item * 100 / total))
            if OVrow[-1] == 0:
                compRowData.append(0)
            else:
                value = int(ANPRrow[-1] * 100 / OVrow[-1])
                if timestamp != "1 Hr":
                    if value < movement["summary"]["MinCapture"]:
                        movement["summary"]["MinCapture"] = value
                    if value > movement["summary"]["MaxCapture"]:
                        movement["summary"]["MaxCapture"] = value
                    if value < 85:
                        movement["summary"]["TimeLessThan"] += 1
                compRowData.append(value)
            compRowData = [str(item) + "%" for item in compRowData]
            compRowData.insert(0, timestamp)
            compData.append(compRowData)

        movement["summary"]["TimeLessThan"] = datetime.timedelta(
            seconds=movement["summary"]["TimeLessThan"] * self.currentJob["interval"] * 60)
        if movement["summary"]["OVTotal"] != 0:
            # print("ovtotal",site["summary"]["OVTotal"])
            movement["summary"]["AvgCapture"] = int(movement["summary"]["ANPRTotal"] * 100 / movement["summary"]["OVTotal"])

        vars = self.comparisonDataStructure[2]
        #print(vars)
        for i,row in enumerate(compData):
            #print("row is",row)
            v = vars[i]
            for j, item in enumerate(row):
                v[j].set(item)

        for label in self.colourLabels:
            val = label.cget("text")
            print("val is",val)
            if val =="":
                val = 0
            else:
                val = int(val.replace("%",""))
            label.configure(fg=get_colour(val))

    def movementChanged(self,event):


        if self.displayStatus == "base":
            ###
            ### if the user has reverted, then not edited anything, then changed movement, we need to
            ### deal with the reverted data, putting it back into the previously viewed movement
            ###
            for site in self.dataList:
                for key, m in site["movements"].items():
                    if self.currentSelected[1] == key:
                        print("selected site is", site["siteNo"])
                        selectedSite = site
            l = [movement for key, movement in sorted(selectedSite["movements"].items()) if key == self.currentSelected[1]]
            movement = l[0]
            movement["data"][1] = self.tempEditedDataStore
            self.tempEditedDataStore = []

        mvmnt = int(event.widget.get().replace("Movement", ""))
        for site in self.dataList:
            for key, m in site["movements"].items():
                if mvmnt == key:
                    print("selected site is", site["siteNo"])
                    selectedSite = site
        self.siteLabel.configure(text="Site " + str(selectedSite["siteNo"]))
        self.currentSelected[0] = selectedSite["siteNo"]
        self.currentSelected[1] = mvmnt
        self.displayStatus = "edited"
        self.revertButton.configure(text="Revert")

        self.update_comparison_display()
        self.update()
        self.update_idletasks()

    def load_job(self,event):
        inMov = []
        outMov = []
        self.currentJob = None
        self.durationsDictionary = None
        self.selectedDuplicates = None
        self.tempEditedDataStore = []
        self.overtakingPairsDict = {}
        self.comparisonDataStructure = []
        self.dataList = []
        jobname = self.tree.item(self.tree.selection()[0])
        self.currentJob = myDB.load_job(jobname["values"][0],jobname["values"][1], datetime.datetime.strptime(jobname["values"][2],"%d/%m/%y").date())
        print("current job folder is2,",self.currentJob["folder"])
        title = self.currentJob["jobno"] + " " + self.currentJob["jobname"]
        self.wm_title(title)
        for site, details in self.currentJob["sites"].items():
            for mvmtNo, mvmt in details.items():
                if mvmt["newmovement"] not in inMov:
                    inMov.append(mvmt["newmovement"])
                if mvmt["newmovement"] not in outMov:
                    outMov.append(mvmt["newmovement"])
        inMov = sorted(inMov)
        outMov = sorted(outMov)


        ###
        ### make sure the data and output folders exist
        ###

        try:
            os.mkdir(self.currentJob["folder"] + "/output")
        except Exception as e:
            print(e, type(e))
        try:
            os.mkdir(self.currentJob["folder"] + "/data")
        except Exception as e:
            print(e, type(e))


        ###
        ### each movement may have a slightly out time, depending on the settings on the camera
        ### so we may need to add or subtract a time delta from all reocords from a certain movement
        ### timeAdjustmentDictionary is to keep a record of any entered adjustments, and apply them on loading
        ### the job

        timeAdjustmentsDictionary = {}
        try:
            with open(self.currentJob["folder"] + "/data/timeAdjustments.pkl", "rb") as f:
                timeAdjustmentsDictionary = pickle.load(f)
        except IOError as e:
            print("setting up time adjustments dictionary")
            for i in inMov:
                timeAdjustmentsDictionary[i] = 0
            print("dict is", timeAdjustmentsDictionary)


        ###
        ### make sure that all movements have an associated time adjustment value
        ###

        for i in inMov:
            try:
                print(timeAdjustmentsDictionary[i])
            except KeyError as e:
                timeAdjustmentsDictionary[i] = 0
        with open(self.currentJob["folder"] + "/data/timeAdjustments.pkl", "wb") as f:
            pickle.dump(timeAdjustmentsDictionary, f)
        self.currentJob["timeAdjustmentsDictionary"] = timeAdjustmentsDictionary

        self.durationsDictionary = {}

        try:
            with open(self.currentJob["folder"] + "/data/durations.pkl", "rb") as f:
                self.durationsDictionary = pickle.load(f)
        except IOError as e:
            pass
        startTime = self.currentJob["timeperiod1"].split("-")[0]
        endTime = self.currentJob["timeperiod1"].split("-")[1]
        for i in range(2,5):
            if self.currentJob["timeperiod" + str(i)].split("-")[1] != "":
                endTime = self.currentJob["timeperiod" + str(i)].split("-")[1]
        print(startTime,endTime)

        t = datetime.datetime.strptime(startTime, "%H:%M")
        t1 = datetime.datetime.strptime(endTime, "%H:%M")
        t =  format_timedelta(t1 - t)


        ###
        ### make sure that all movement pairs have an associated duration value
        ###

        for i in inMov:
            for o in outMov:
                try:
                    print(self.durationsDictionary[i, o])
                except KeyError as e:
                    print("key error in durations dictionary",i,o)
                    self.durationsDictionary[i, o] = t

        self.currentJob["durationsDictionary"] = self.durationsDictionary
        with open(self.currentJob["folder"] + "/data/durations.pkl", "wb") as f:
            pickle.dump(self.durationsDictionary, f)

        ###
        ### try and load the data , classes, plates etc, related to the current job
        ###

        if not self.loadJobFunction(self.currentJob):
            return

        self.spawn_home_window()

    def delete_job(self):
        if self.tree.selection() == "":
            return
        jobname = self.tree.item(self.tree.selection()[0])
        result = messagebox.askyesno(message="Are you sure you want to delete this project?")
        if not result:
            return
        myDB.delete_job(jobname["values"][0], jobname["values"][1], jobname["values"][2])
        self.spawn_survey_setup_screen()

    def export_OVTemplate(self):
        wb = openpyxl.load_workbook("OV template.xlsm", keep_vba=True)
        try:
            sheet = wb.get_sheet_by_name("Temp")
        except Exception as e:
            messagebox.showinfo(message="Trying to export to excel, sheet Temp doesnt exist in the template file,cannot export")
            return
        classes = self.currentJob["classification"].split(",")
        classes = [x for i,x in enumerate(classes) if i % 2 == 0]
        for i,c in enumerate(classes):
            sheet.cell(row = 1 + i,column=1).value = c
        times = self.currentJob["timeperiod1"].split("-") + self.currentJob["timeperiod2"].split("-") + self.currentJob["timeperiod3"].split("-") + self.currentJob["timeperiod4"].split("-")
        print("times are ",times)
        for i,t in enumerate(times):
            sheet.cell(row=1 + i, column=2).value = t
        sheet.cell(row=1,column=3).value = self.currentJob["interval"]
        col = 4
        for k,v in self.currentJob["sites"].items():
            row = 1
            sheet.cell(row=row,column=col).value = int(k)
            for key,value in v.items():
                row+=1
                sheet.cell(row=row, column=col).value = int(key)
            col+=1
        file = self.currentJob["folder"] + "/" + self.currentJob["jobno"] +  " " + self.currentJob["jobname"] + " - OV Template " + self.currentJob["surveydate"].strftime("%d-%m-%y") +  ".xlsm"
        print("file is",file)
        wb.save(file)
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=file, ReadOnly=1)
        xl.Application.Run("create_template")
        xl.Workbooks(1).Close(SaveChanges=1)
        xl.Application.Quit()
        myDB.update_job_with_progress(self.currentJob["id"], "OVTemplate")

    def load_unclassed_plates(self):
        self.loadUnclassedFunction(self.currentJob)

    def get_comparison_data(self):
        self.dataList = self.getComparisonFunction(self.currentJob)
        if not self.dataList is None:
            self.set_up_comparison_display()

    def load_OV_counts(self):
        self.loadOVCountsFunction(self.currentJob)
        myDB.update_job_with_progress(self.currentJob["id"], "OVCounts")

    def load_classes(self):
        self.loadClassedFunction(self.currentJob)

    def setCallbackFunction(self, text, fun):
        if text == "load unclassed":
            self.loadUnclassedFunction = fun
        if text == "load classed":
            self.loadClassedFunction = fun
        if text == "load job":
            self.loadJobFunction = fun
        if text == "load overview count":
            self.loadOVCountsFunction = fun
        if text == "get unclassed comparison":
            self.getComparisonFunction = fun
        if text == "reprocess data":
            self.reprocessDataFunction = fun
        if text == "set duplicates":
            self.setDuplicatesFunction = fun
        if text == "get cordon in out only data":
            self.getCordonFunction = fun
        if text == "get cordon non directional data":
            self.getNonDirectionalCordonFunction = fun
        if text == "get fs-ls data":
            self.getRouteAssignmentFsLsFunction = fun
        if text == "get journey pairs":
            self.getJourneyPairsFunction = fun
        if text == "get overtaking data":
             self.getOvertakingDataFunction = fun
        if text == "resample overtaking data":
            self.resampleOvertakingDataFunction = fun
        if text == "update data after job save":
            self.updateDataFunction = fun
        if text == "recalculate platooning":
            self.recalcuatePlatooningfunction = fun
        if text == "filtered matching":
            self.filteredMatchingfunction = fun

    def on_label_entry(self,event):
        event.widget.configure(fg=self.tracsisGrey)

    def on_label_exit(self,event):
        event.widget.configure(fg=self.tracsisBlue)

    def find_in_grid(self,frame, row, column):
        if frame is None:
            return
        for children in frame.children.values():
            info = children.grid_info()
            print("info is",info)
            # note that rows and column numbers are stored as string
            if info['row'] == str(row) and info['column'] == str(column):
                return children
        return None

def format_timedelta(td):
    minutes, seconds = divmod(td.seconds + td.days * 86400, 60)
    hours, minutes = divmod(minutes, 60)
    return '{:d}:{:02d}'.format(hours, minutes)

def get_colour(value):
    if isinstance(value,str):
        return "black"
    value = int(value)
    if value == 0:
        return "black"
    if value > 100:
        return "blue"
    if value < 85:
        return  "red"
    return "black"

class VerticalScrolledFrame(tkinter.Frame):
    """A pure Tkinter scrollable frame that actually works!
    * Use the 'interior' attribute to place widgets inside the scrollable frame
    * Construct and pack/place/grid normally
    * This frame only allows vertical scrolling

    """
    def __init__(self, parent, *args, **kw):
        tkinter.Frame.__init__(self, parent, *args, **kw)
        print("height of scroll",parent.winfo_height())
        parentOfparent = self.nametowidget(parent.winfo_parent())
        #parentOfparent.configure(bg="black")
        print("parent of parent height is",parentOfparent.winfo_height())
        # create a canvas object and a vertical scrollbar for scrolling it
        vscrollbar = tkinter.Scrollbar(self, orient=tkinter.VERTICAL,bg="white")
        vscrollbar.pack(fill=tkinter.Y, side=tkinter.RIGHT, expand=tkinter.TRUE)
        self.canvas = tkinter.Canvas(self, bd=0, highlightthickness=0,bg="white",
                        yscrollcommand=vscrollbar.set,height = parent.winfo_height()-64)
        self.canvas.bind_all("<MouseWheel>",self.on_mousewheel)
        self.canvas.bind("<Enter>",self.on_entry)
        self.canvas.bind("<Leave>", self.on_exit)
        self.canvas.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=tkinter.TRUE)
        vscrollbar.config(command=self.canvas.yview)

        # reset the view
        self.canvas.xview_moveto(0)
        self.canvas.yview_moveto(0)

        # create a frame inside the canvas which will be scrolled with it
        self.interior = interior = tkinter.Frame(self.canvas,bg="white",width = 100)
        interior_id = self.canvas.create_window(0, 0, window=interior,
                                           anchor=tkinter.NW)

        # track changes to the canvas and frame width and sync them,
        # also updating the scrollbar
        def _configure_interior(event):
            # update the scrollbars to match the size of the inner frame
            size = (interior.winfo_reqwidth(), interior.winfo_reqheight())
            #print("required size is",size)
            #print("actual size is",canvas.winfo_width(),canvas.winfo_height())
            self.canvas.config(scrollregion="0 0 %s %s" % size)
            if interior.winfo_reqwidth() != self.canvas.winfo_width():
                # update the canvas's width to fit the inner frame
                self.canvas.config(width=interior.winfo_reqwidth())
            #if interior.winfo_reqheight() != self.canvas.winfo_height():
                # update the canvas's height to fit the inner frame
                #self.canvas.config(height=interior.winfo_reqheight())
            #print("actual size is", canvas.winfo_width(), canvas.winfo_height())
        interior.bind('<Configure>', _configure_interior)

        def _configure_canvas(event):
            if interior.winfo_reqwidth() != self.canvas.winfo_width():
                # update the inner frame's width to fill the canvas
                self.canvas.itemconfigure(interior_id, width=self.canvas.winfo_width())
                self.canvas.bind('<Configure>', _configure_canvas)

    def on_exit(self,event):
        print("left canvas")
        self.canvas.unbind_all("<MouseWheel>")

    def on_entry(self,event):
        print("entered canvas")
        self.canvas.bind_all("<MouseWheel>", self.on_mousewheel)

    def on_mousewheel(self,event):
        print("psdf")
        self.canvas.yview_scroll(-1 * int(event.delta / 120), "units")

